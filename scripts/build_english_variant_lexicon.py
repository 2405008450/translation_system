from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


LEXICON_COLUMNS = (
    "british",
    "american",
    "category",
    "form",
    "to_american_enabled",
    "to_british_enabled",
    "source_refs",
    "notes",
)

CATEGORY_BY_SHEET_NAME = {
    "名词": "noun",
    "动词": "verb",
    "形容词": "adjective",
    "副词": "adverb",
}


@dataclass
class LexiconRow:
    british: str
    american: str
    categories: set[str] = field(default_factory=set)
    forms: set[str] = field(default_factory=set)
    source_refs: set[str] = field(default_factory=set)
    to_american_enabled: bool = True
    to_british_enabled: bool = True
    notes: set[str] = field(default_factory=set)


def _clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def _form_name(header: str, fallback: str) -> str:
    normalized = header.replace("英式", "").replace("美式", "").strip()
    return normalized or fallback


def _merge_row(merged: dict[tuple[str, str], LexiconRow], row: LexiconRow) -> None:
    key = (row.british.casefold(), row.american.casefold())
    item = merged.setdefault(
        key,
        LexiconRow(british=row.british, american=row.american),
    )
    item.categories.update(row.categories)
    item.forms.update(row.forms)
    item.source_refs.update(row.source_refs)


def read_csv_rows(path: Path) -> list[LexiconRow]:
    """读取已有运行时词库，用于增量合并；启用状态和冲突备注会重新计算。"""
    rows: list[LexiconRow] = []
    with path.open("r", encoding="utf-8", newline="") as stream:
        reader = csv.DictReader(stream)
        if tuple(reader.fieldnames or ()) != LEXICON_COLUMNS:
            raise ValueError(f"已有词库列结构不符合要求：{path}")
        for row_number, values in enumerate(reader, start=2):
            british = _clean(values.get("british"))
            american = _clean(values.get("american"))
            if not british or not american or british.casefold() == american.casefold():
                raise ValueError(f"已有词库第 {row_number} 行包含无效词对")
            rows.append(
                LexiconRow(
                    british=british,
                    american=american,
                    categories=set(filter(None, _clean(values.get("category")).split("|"))),
                    forms=set(filter(None, _clean(values.get("form")).split("|"))),
                    source_refs=set(filter(None, _clean(values.get("source_refs")).split("|"))),
                )
            )
    return rows


def _detect_column_pairs(headers: list[str], worksheet_title: str) -> list[tuple[int, int, str]]:
    category = CATEGORY_BY_SHEET_NAME.get(worksheet_title)
    if category is None:
        return []

    british_columns = [index for index, header in enumerate(headers) if header.startswith("英式")]
    american_columns = [index for index, header in enumerate(headers) if header.startswith("美式")]
    if not british_columns or len(british_columns) != len(american_columns):
        raise ValueError(f"工作表“{worksheet_title}”的英式/美式列无法配对")

    column_pairs: list[tuple[int, int, str]] = []
    for pair_index, (british_column, american_column) in enumerate(
        zip(british_columns, american_columns),
        start=1,
    ):
        british_form = _form_name(headers[british_column], "基本形式")
        american_form = _form_name(headers[american_column], "基本形式")
        if british_form != american_form:
            raise ValueError(
                f"工作表“{worksheet_title}”第 {pair_index} 组英式/美式词形不一致："
                f"{headers[british_column]!r} / {headers[american_column]!r}"
            )
        column_pairs.append((british_column, american_column, british_form))
    return column_pairs


def read_excel_rows(path: Path, base_rows: list[LexiconRow] | None = None) -> list[LexiconRow]:
    merged: dict[tuple[str, str], LexiconRow] = {}
    for row in base_rows or ():
        _merge_row(merged, row)

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            header_values = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                (),
            )
            headers = [_clean(value) for value in header_values]
            category = CATEGORY_BY_SHEET_NAME.get(worksheet.title)
            column_pairs = _detect_column_pairs(headers, worksheet.title)
            if not column_pairs:
                continue

            for row_number, values in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True),
                start=2,
            ):
                for british_column, american_column, form_name in column_pairs:
                    if max(british_column, american_column) >= len(values):
                        continue
                    british = _clean(values[british_column])
                    american = _clean(values[american_column])
                    if not british or not american or british.casefold() == american.casefold():
                        continue

                    _merge_row(
                        merged,
                        LexiconRow(
                            british=british,
                            american=american,
                            categories={category} if category else set(),
                            forms={form_name},
                            source_refs={
                                f"{worksheet.title}!{get_column_letter(british_column + 1)}{row_number}:"
                                f"{get_column_letter(american_column + 1)}{row_number}"
                            },
                        ),
                    )
    finally:
        workbook.close()

    rows = list(merged.values())
    _disable_ambiguous_directions(rows)
    return sorted(rows, key=lambda item: (item.british.casefold(), item.american.casefold()))


def _disable_ambiguous_directions(rows: list[LexiconRow]) -> None:
    british_targets: dict[str, set[str]] = defaultdict(set)
    american_targets: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        british_targets[row.british.casefold()].add(row.american.casefold())
        american_targets[row.american.casefold()].add(row.british.casefold())

    for row in rows:
        if len(british_targets[row.british.casefold()]) > 1:
            row.to_american_enabled = False
            row.notes.add("英式源词存在多个美式目标，已禁用英转美")
        if len(american_targets[row.american.casefold()]) > 1:
            row.to_british_enabled = False
            row.notes.add("美式源词存在多个英式目标，已禁用美转英")

    # 已是目标方言的词必须能够被保护。若它同时还是本方向的源词，无法在不理解
    # 语义的情况下安全判断，因此禁用该源词方向，保证重复执行不会继续改写。
    changed = True
    while changed:
        changed = False
        american_targets_enabled = {
            row.american.casefold()
            for row in rows
            if row.to_american_enabled
        }
        british_targets_enabled = {
            row.british.casefold()
            for row in rows
            if row.to_british_enabled
        }
        for row in rows:
            if row.to_american_enabled and row.british.casefold() in american_targets_enabled:
                row.to_american_enabled = False
                row.notes.add("英式源词同时是英转美目标，为保证幂等已禁用英转美")
                changed = True
            if row.to_british_enabled and row.american.casefold() in british_targets_enabled:
                row.to_british_enabled = False
                row.notes.add("美式源词同时是美转英目标，为保证幂等已禁用美转英")
                changed = True


def validate_rows(rows: list[LexiconRow]) -> None:
    if not rows:
        raise ValueError("未从工作簿中识别到英美词汇")

    seen_pairs: set[tuple[str, str]] = set()
    to_american: dict[str, str] = {}
    to_british: dict[str, str] = {}
    for row in rows:
        british = row.british.casefold()
        american = row.american.casefold()
        if not british or not american or british == american:
            raise ValueError(f"存在无效词对：{row.british!r} / {row.american!r}")
        pair = (british, american)
        if pair in seen_pairs:
            raise ValueError(f"存在重复词对：{row.british!r} / {row.american!r}")
        seen_pairs.add(pair)
        if row.to_american_enabled:
            previous = to_american.setdefault(british, american)
            if previous != american:
                raise ValueError(f"启用的英转美映射冲突：{row.british!r}")
        if row.to_british_enabled:
            previous = to_british.setdefault(american, british)
            if previous != british:
                raise ValueError(f"启用的美转英映射冲突：{row.american!r}")

    if set(to_american).intersection(to_american.values()):
        raise ValueError("启用的英转美映射存在源词/目标词重叠，无法保证幂等")
    if set(to_british).intersection(to_british.values()):
        raise ValueError("启用的美转英映射存在源词/目标词重叠，无法保证幂等")


def write_csv(rows: list[LexiconRow], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=LEXICON_COLUMNS, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "british": row.british,
                    "american": row.american,
                    "category": "|".join(sorted(row.categories)),
                    "form": "|".join(sorted(row.forms)),
                    "to_american_enabled": str(row.to_american_enabled).lower(),
                    "to_british_enabled": str(row.to_british_enabled).lower(),
                    "source_refs": "|".join(sorted(row.source_refs)),
                    "notes": "；".join(sorted(row.notes)),
                }
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从 Excel 构建英美英语运行时词库")
    parser.add_argument("input", type=Path, help="英美词汇 Excel 文件")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("app/resources/english_variant_lexicon.csv"),
        help="输出 CSV 路径",
    )
    parser.add_argument(
        "--merge-existing",
        type=Path,
        help="可选：先载入已有 CSV，再将工作簿内容合并进去",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_rows = None
    if args.merge_existing is not None:
        base_rows = read_csv_rows(args.merge_existing.expanduser().resolve())
    rows = read_excel_rows(args.input.expanduser().resolve(), base_rows=base_rows)
    validate_rows(rows)
    output_path = args.output.expanduser().resolve()
    write_csv(rows, output_path)
    print(f"已生成 {output_path}，共 {len(rows)} 条词对")


if __name__ == "__main__":
    main()
