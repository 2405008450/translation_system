"""
API 路由模块 - 文件上传、解析和导出接口

支持多种文档格式的上传、解析和导出。
"""
import asyncio
import json
import logging
import re
from datetime import datetime
from functools import partial
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional
from urllib.parse import quote, unquote, urlparse
from uuid import UUID, uuid4

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, Header, HTTPException, Query, Request, UploadFile
from fastapi.responses import JSONResponse, Response, StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import case, func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.auth import (
    USER_ROLE,
    can_access_all_projects,
    get_current_user,
    get_user_by_id,
    get_user_display_name,
    is_admin_role,
    is_external_translator,
    require_admin,
    serialize_user,
)
from app.config import get_settings
from app.database import SessionLocal, get_db
from app.models import (
    FileRecord,
    IssueMarker,
    MemoryBase,
    MemoryEntry,
    Project,
    Segment,
    SegmentRevision,
    TMCollection,
    TermBase,
    TermEntry,
    TranslationMemory,
    User,
)
from app.services.adapters import (
    DocumentAST,
    ExportService,
    FileTooLargeError,
    ParseError,
    UnsupportedFormatError,
    get_registry,
)
from app.services.analytics_service import (
    count_source_words,
    get_dashboard_payload,
    record_translation_metric_event,
    run_analytics_backfill_once,
)
from app.services.auto_tm_sync import (
    AutoTMEnqueueSummary,
    enqueue_confirmed_segments_for_auto_tm,
    run_auto_tm_background_once,
)
from app.services.adapters.dita_exporter import DitaExporter
from app.services.adapters.svg_exporter import SvgExporter
from app.services.adapters.tmx_exporter import TmxExporter
from app.services.adapters.xliff_exporter import XliffExporter, XliffImporter
from app.services.comment_service import (
    create_segment_comment,
    create_segment_comment_reply,
    delete_segment_comment,
    list_segment_comments_for_file_record,
    serialize_segment_comment,
    update_segment_comment,
)
from app.services.document_statistics import compute_docx_statistics, serialize_document_statistics
from app.services.issue_marker_service import (
    create_issue_marker,
    delete_issue_marker,
    list_issue_markers_for_project,
    serialize_issue_marker,
    update_issue_marker,
)
from app.services.cache import get_json as cache_get_json
from app.services.cache import set_json as cache_set_json
from app.services.file_record_service import (
    backfill_file_record_source_html,
    batch_update_segments,
    calculate_file_record_progress,
    create_file_record_with_segments,
    delete_file_record,
    duplicate_file_record,
    get_file_record as get_file_record_model,
    get_file_record_document_statistics,
    get_file_record_source_filename,
    get_file_record_with_segments,
    get_tm_target_text_map,
    list_file_records,
    list_segments_for_file_record,
    load_file_record_source,
    resolve_file_record_status,
    sync_file_record_status,
    update_segment_by_sentence_id,
    update_segment_source_text,
)
from app.services.file_operation_lock_service import (
    FILE_OPERATION_TOKEN_HEADER,
    PRE_TRANSLATE_OPERATION,
    acquire_file_operation_lock,
    clear_stale_file_operation_lock,
    ensure_file_record_write_allowed,
    heartbeat_file_operation_lock,
    release_file_operation_lock,
    serialize_file_operation_state,
)
from app.services.guideline_repository import (
    GuidelineTemplate,
    delete_guideline_template,
    list_guideline_templates,
    read_guideline_template,
    save_guideline_template,
    update_guideline_template,
)
from app.services.llm_service import (
    LLMConfigurationError,
    LLMRequestError,
    LLMResponseValidationError,
    LLMTranslationFailure,
    LLMTranslationTask,
    iter_batch_translate,
    validate_provider_choice,
)
from app.services.term_entry_service import build_term_entry_conflict_items
from app.services.term_extraction_service import (
    TERM_EXTRACTION_MODEL,
    TERM_EXTRACTION_MODEL_OPTIONS,
    ExtractedTerm,
    TermExtractionError,
    extract_terms_from_segments,
    merge_extracted_terms,
)
from app.services.language_detection import detect_upload_language
from app.services.language_pairs import require_language_pair
from app.services.matcher import get_tm_candidates_for_text, match_sentences_with_stats
from app.services.normalizer import build_source_hash, normalize_match_text, normalize_text
from app.services.revision_service import (
    accept_revision,
    batch_accept_revisions,
    batch_reject_revisions,
    get_revision_or_404,
    list_revisions,
    reject_revision,
    serialize_segment_revision,
)
from app.services.resource_export_queue import (
    ResourceExportFormat,
    build_resource_export_download_response,
    ensure_export_task_status,
    queue_resource_export,
)
from app.services.slate_parser import parse_docx_for_slate
from app.services.task_file_service import (
    BILINGUAL_DOCX_LAYOUT_EXPORT_ORDERS,
    DOCUMENT_PARSE_MODE_FULL,
    build_task_preview_html,
    build_task_workspace,
    can_export_task_file,
    export_bilingual_task_docx_with_layout,
    export_translated_task_file,
    get_upload_capabilities,
    get_supported_task_extensions,
    get_task_file_extension,
    normalize_document_parse_options,
    normalize_document_parse_mode,
    supports_task_file,
)
from app.services.tm_importer import (
    SDLTM_EXTENSIONS,
    TM_IMPORT_EXTENSIONS,
    XLSX_EXTENSIONS,
    import_tm_from_sdltm_upload,
    import_tm_from_xlsx_upload,
    preview_sdltm_metadata,
)
from app.services.tm_vector import sync_tm_embeddings
from app.services.translation_memory_service import TMUpsertEntry, batch_upsert_tm_entries
from app.services.xlsx_exporter import build_tabular_xlsx, build_xlsx_download_response

try:
    from arq import create_pool as arq_create_pool
    from arq.connections import RedisSettings
except ModuleNotFoundError:  # pragma: no cover - 本地未安装 ARQ 时使用 FastAPI 后台任务兜底
    arq_create_pool = None
    RedisSettings = None


logger = logging.getLogger(__name__)
router = APIRouter(dependencies=[Depends(get_current_user)])


def _build_task_workspace_with_new_session(
    raw_bytes: bytes,
    filename: str,
    similarity_threshold: float,
    collection_ids: list[UUID] | None,
    document_parse_mode: str,
    document_parse_options: dict[str, object] | str | None = None,
) -> dict:
    with SessionLocal() as workspace_db:
        return build_task_workspace(
            db=workspace_db,
            raw_bytes=raw_bytes,
            filename=filename,
            similarity_threshold=similarity_threshold,
            collection_ids=collection_ids,
            document_parse_mode=document_parse_mode,
            document_parse_options=document_parse_options,
        )


async def _build_task_workspace_async(
    raw_bytes: bytes,
    filename: str,
    similarity_threshold: float,
    collection_ids: list[UUID] | None = None,
    document_parse_mode: str = DOCUMENT_PARSE_MODE_FULL,
    document_parse_options: dict[str, object] | str | None = None,
) -> dict:
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(
        None,
        partial(
            _build_task_workspace_with_new_session,
            raw_bytes=raw_bytes,
            filename=filename,
            similarity_threshold=similarity_threshold,
            collection_ids=collection_ids,
            document_parse_mode=document_parse_mode,
            document_parse_options=document_parse_options,
        ),
    )


def _begin_repeatable_read_snapshot(db: Session) -> None:
    if db.get_bind().dialect.name == "postgresql":
        db.connection(execution_options={"isolation_level": "REPEATABLE READ"})


IMPORT_TASK_TTL_SECONDS = 24 * 60 * 60


def _import_task_cache_key(task_id: str) -> str:
    return f"import-task:{task_id}"


def _set_import_task_status(
    task_id: str,
    status: Literal["queued", "running", "completed", "failed"],
    *,
    progress: int = 0,
    message: str = "",
    result: dict[str, Any] | None = None,
    error: str | None = None,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "task_id": task_id,
        "status": status,
        "progress": max(0, min(100, int(progress))),
        "message": message,
        "result": result,
        "error": error,
        "updated_at": datetime.utcnow().isoformat(),
    }
    cache_set_json(_import_task_cache_key(task_id), payload, ttl_seconds=IMPORT_TASK_TTL_SECONDS)
    return payload


def _get_import_task_status(task_id: str) -> dict[str, Any] | None:
    payload = cache_get_json(_import_task_cache_key(task_id))
    return payload if isinstance(payload, dict) else None


def _build_arq_redis_settings(redis_url: str):
    if RedisSettings is None:
        return None

    parsed = urlparse(redis_url)
    database = int((parsed.path or "/0").lstrip("/") or "0")
    kwargs = {
        "host": parsed.hostname or "localhost",
        "port": parsed.port or 6379,
        "database": database,
        "password": unquote(parsed.password) if parsed.password else None,
        "ssl": parsed.scheme == "rediss",
    }
    if parsed.username:
        kwargs["username"] = unquote(parsed.username)
    try:
        return RedisSettings(**kwargs)
    except TypeError:
        kwargs.pop("username", None)
        return RedisSettings(**kwargs)


async def _close_arq_pool(redis_pool) -> None:
    close = getattr(redis_pool, "close", None)
    if close is not None:
        close_result = close()
        if asyncio.iscoroutine(close_result):
            await close_result

    wait_closed = getattr(redis_pool, "wait_closed", None)
    if wait_closed is not None:
        wait_result = wait_closed()
        if asyncio.iscoroutine(wait_result):
            await wait_result


async def _enqueue_arq_import_task(task_id: str, payload: dict[str, Any]) -> bool:
    settings = get_settings()
    if settings.import_queue_backend.lower() != "arq":
        return False
    if not settings.redis_url or arq_create_pool is None:
        return False

    redis_settings = _build_arq_redis_settings(settings.redis_url)
    if redis_settings is None:
        return False

    redis_pool = None
    try:
        redis_pool = await arq_create_pool(redis_settings)
        await redis_pool.enqueue_job("process_import_task_job", task_id, payload)
        return True
    except Exception:
        logger.warning("enqueue ARQ import task failed, fallback to local background task", exc_info=True)
        return False
    finally:
        if redis_pool is not None:
            await _close_arq_pool(redis_pool)


async def _queue_import_task(
    background_tasks: BackgroundTasks,
    payload: dict[str, Any],
) -> JSONResponse:
    task_id = str(uuid4())
    _set_import_task_status(task_id, "queued", progress=0, message="任务已进入导入队列。")
    if not await _enqueue_arq_import_task(task_id, payload):
        background_tasks.add_task(_run_import_task, task_id, payload)

    return JSONResponse(
        status_code=202,
        content={
            "task_id": task_id,
            "status": "queued",
            "progress": 0,
            "message": "任务已进入导入队列。",
        },
    )


def _uuid_list(values: list[str] | None) -> list[UUID]:
    return [UUID(value) for value in values or []]


def _load_file_record_term_base_ids(file_record: FileRecord) -> list[UUID]:
    raw_ids = getattr(file_record, "term_base_ids", "") or "[]"
    parsed_ids: list[UUID] = []
    try:
        values = json.loads(raw_ids)
    except (TypeError, ValueError):
        values = []
    if isinstance(values, list):
        for value in values:
            try:
                parsed_ids.append(value if isinstance(value, UUID) else UUID(str(value)))
            except (TypeError, ValueError):
                continue
    if not parsed_ids and file_record.term_base_id:
        parsed_ids.append(file_record.term_base_id)
    return list(dict.fromkeys(parsed_ids))


def _store_file_record_term_base_ids(file_record: FileRecord, term_base_ids: list[UUID]) -> None:
    normalized_ids = list(dict.fromkeys(term_base_ids))
    file_record.term_base_id = normalized_ids[0] if normalized_ids else None
    file_record.term_base_ids = json.dumps([str(term_base_id) for term_base_id in normalized_ids])


def _serialize_file_record_upload_result(file_record: FileRecord) -> dict[str, Any]:
    return {
        "id": str(file_record.id),
        "filename": file_record.filename,
        "status": file_record.status,
        "document_parse_mode": file_record.document_parse_mode,
        "document_parse_options": _get_file_record_document_parse_options(file_record),
        "document_statistics": get_file_record_document_statistics(file_record),
        "created_at": file_record.created_at.isoformat(),
    }


def _json_ready_project_file_payload(payload: dict[str, Any]) -> dict[str, Any]:
    for key in ("collection_id", "term_base_id"):
        if payload.get(key) is not None:
            payload[key] = str(payload[key])
    return payload


def _get_file_record_document_parse_options(file_record: FileRecord) -> dict[str, bool]:
    return normalize_document_parse_options(
        getattr(file_record, "document_parse_options", None),
        getattr(file_record, "document_parse_mode", DOCUMENT_PARSE_MODE_FULL),
    )


def _process_file_record_import(db: Session, payload: dict[str, Any]) -> dict[str, Any]:
    file_payload = payload["file"]
    raw_bytes = file_payload["content"]
    filename = file_payload["filename"] or "untitled.txt"
    selected_collection_ids = _uuid_list(payload.get("collection_ids"))
    term_base_id = UUID(payload["term_base_id"]) if payload.get("term_base_id") else None
    document_parse_mode = payload.get("document_parse_mode") or DOCUMENT_PARSE_MODE_FULL
    document_parse_options = payload.get("document_parse_options")
    threshold = float(payload.get("threshold", 0.6))

    primary_collection = _get_collection_or_404(
        db,
        selected_collection_ids[0] if selected_collection_ids else None,
    )
    resolved_source_language, resolved_target_language = _resolve_upload_language_pair(
        payload.get("source_language"),
        payload.get("target_language"),
        primary_collection,
    )

    term_base = None
    if term_base_id is not None:
        term_base = db.query(TermBase).filter(TermBase.id == term_base_id).first()
        if term_base is None:
            raise HTTPException(status_code=404, detail="术语库不存在。")
        _ensure_resource_language_pair_matches(
            term_base,
            resolved_source_language,
            resolved_target_language,
            "术语库",
        )

    workspace_data = build_task_workspace(
        db=db,
        raw_bytes=raw_bytes,
        filename=filename,
        similarity_threshold=threshold,
        collection_ids=selected_collection_ids,
        document_parse_mode=document_parse_mode,
        document_parse_options=document_parse_options,
    )
    file_record = create_file_record_with_segments(
        db=db,
        raw_bytes=raw_bytes,
        filename=filename,
        similarity_threshold=threshold,
        workspace_data=workspace_data,
        collection_ids=selected_collection_ids,
        document_parse_mode=document_parse_mode,
        document_parse_options=document_parse_options,
    )
    if selected_collection_ids:
        file_record.collection_id = selected_collection_ids[0]
        file_record.collection_ids_json = json.dumps([str(cid) for cid in selected_collection_ids])
        _apply_collection_language_pair(file_record, primary_collection)
    file_record.source_language = resolved_source_language
    file_record.target_language = resolved_target_language
    if payload.get("creator_id"):
        file_record.creator_id = UUID(payload["creator_id"])
    if term_base is not None:
        _store_file_record_term_base_ids(file_record, [term_base_id])

    db.commit()
    db.refresh(file_record)
    return _serialize_file_record_upload_result(file_record)


def _expand_archive_payloads(file_payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """展开压缩包文件为独立文件列表。

    如果上传的文件是 .zip 或 .rar，则解压其中支持的文件，
    每个文件作为独立的 file_payload 返回。非压缩包文件原样保留。
    """
    import zipfile
    from io import BytesIO

    expanded: list[dict[str, Any]] = []

    for file_payload in file_payloads:
        filename = file_payload.get("filename") or ""
        ext = Path(filename).suffix.lower()

        if ext == ".zip":
            extracted = _extract_zip_files(file_payload["content"], filename)
            if extracted:
                expanded.extend(extracted)
            else:
                # 如果解压后没有支持的文件，保留原始 zip 处理方式
                expanded.append(file_payload)
        elif ext == ".rar":
            extracted = _extract_rar_files(file_payload["content"], filename)
            if extracted:
                expanded.extend(extracted)
            else:
                expanded.append(file_payload)
        else:
            expanded.append(file_payload)

    return expanded


def _extract_zip_files(raw_bytes: bytes, archive_name: str) -> list[dict[str, Any]]:
    """从 ZIP 文件中提取支持的文件。"""
    import zipfile
    from io import BytesIO

    try:
        zf = zipfile.ZipFile(BytesIO(raw_bytes), 'r')
    except zipfile.BadZipFile:
        return []

    extracted: list[dict[str, Any]] = []

    for info in zf.infolist():
        # 跳过目录
        if info.is_dir():
            continue

        # 获取文件名，处理编码问题
        raw_filename = info.filename
        decoded_filename = _decode_zip_filename(raw_filename, info)

        # 跳过隐藏文件和系统文件
        basename = Path(decoded_filename).name
        if basename.startswith('__') or basename.startswith('.'):
            continue

        # 检查是否是支持的格式
        if not supports_task_file(basename):
            continue

        try:
            file_bytes = zf.read(info.filename)
            if file_bytes:
                extracted.append({
                    "filename": basename,
                    "content": file_bytes,
                })
        except Exception:
            continue

    zf.close()
    return extracted


def _extract_rar_files(raw_bytes: bytes, archive_name: str) -> list[dict[str, Any]]:
    """从 RAR 文件中提取支持的文件。"""
    try:
        import rarfile
        from io import BytesIO

        rf = rarfile.RarFile(BytesIO(raw_bytes))
    except Exception:
        return []

    extracted: list[dict[str, Any]] = []

    try:
        for info in rf.infolist():
            if info.is_dir():
                continue

            decoded_filename = info.filename
            basename = Path(decoded_filename).name

            if basename.startswith('__') or basename.startswith('.'):
                continue

            if not supports_task_file(basename):
                continue

            try:
                file_bytes = rf.read(info.filename)
                if file_bytes:
                    extracted.append({
                        "filename": basename,
                        "content": file_bytes,
                    })
            except Exception:
                continue
    finally:
        rf.close()

    return extracted


def _decode_zip_filename(raw_filename: str, info: Any) -> str:
    """解码 ZIP 文件名，处理中文等非 ASCII 编码问题。

    Python zipfile 模块对非 UTF-8 文件名使用 CP437 解码，
    这会导致中文文件名乱码。此函数尝试修复。
    """
    # 如果 flag_bits 的 bit 11 设置了，说明文件名是 UTF-8 编码
    if info.flag_bits & 0x800:
        return raw_filename

    # 尝试将 CP437 解码的字符串重新编码回字节，再用正确编码解码
    try:
        raw_bytes = raw_filename.encode('cp437')
        # 尝试 UTF-8
        try:
            return raw_bytes.decode('utf-8')
        except UnicodeDecodeError:
            pass
        # 尝试 GBK（中文 Windows 常用）
        try:
            return raw_bytes.decode('gbk')
        except UnicodeDecodeError:
            pass
        # 尝试 Shift-JIS（日文）
        try:
            return raw_bytes.decode('shift_jis')
        except UnicodeDecodeError:
            pass
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass

    return raw_filename


def _process_project_source_import(db: Session, task_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    project_id = UUID(payload["project_id"])
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在。")

    selected_collection_ids = _uuid_list(payload.get("collection_ids"))
    term_base_id = UUID(payload["term_base_id"]) if payload.get("term_base_id") else None
    document_parse_mode = payload.get("document_parse_mode") or DOCUMENT_PARSE_MODE_FULL
    document_parse_options = payload.get("document_parse_options")
    threshold = float(payload.get("threshold", 0.6))
    primary_collection = _get_collection_or_404(
        db,
        selected_collection_ids[0] if selected_collection_ids else None,
    )
    resolved_source_language, resolved_target_language = _resolve_upload_language_pair(
        payload.get("source_language"),
        payload.get("target_language"),
        primary_collection,
    )

    term_base = None
    if term_base_id is not None:
        term_base = db.query(TermBase).filter(TermBase.id == term_base_id).first()
        if term_base is None:
            raise HTTPException(status_code=404, detail="术语库不存在。")
        _ensure_resource_language_pair_matches(
            term_base,
            resolved_source_language,
            resolved_target_language,
            "术语库",
        )

    file_payloads = payload["files"]

    # 展开压缩包：将 zip/rar 文件解压为独立文件
    expanded_payloads = _expand_archive_payloads(file_payloads)

    created_files: list[FileRecord] = []
    for index, file_payload in enumerate(expanded_payloads, start=1):
        filename = file_payload["filename"] or "source.txt"
        _set_import_task_status(
            task_id,
            "running",
            progress=10 + int((index - 1) / max(len(expanded_payloads), 1) * 80),
            message=f"正在解析 {filename}",
        )
        workspace_data = build_task_workspace(
            db=db,
            raw_bytes=file_payload["content"],
            filename=filename,
            similarity_threshold=threshold,
            collection_ids=selected_collection_ids,
            document_parse_mode=document_parse_mode,
            document_parse_options=document_parse_options,
        )
        file_record = create_file_record_with_segments(
            db=db,
            raw_bytes=file_payload["content"],
            filename=filename,
            similarity_threshold=threshold,
            workspace_data=workspace_data,
            collection_ids=selected_collection_ids,
            document_parse_mode=document_parse_mode,
            document_parse_options=document_parse_options,
        )
        file_record.project_id = project.id
        file_record.creator_id = project.creator_id
        file_record.deadline = project.deadline
        file_record.access_level = project.access_level
        file_record.source_language = resolved_source_language
        file_record.target_language = resolved_target_language
        if selected_collection_ids:
            file_record.collection_id = selected_collection_ids[0]
            file_record.collection_ids_json = json.dumps([str(cid) for cid in selected_collection_ids])
        if term_base is not None:
            _store_file_record_term_base_ids(file_record, [term_base_id])
        created_files.append(file_record)

    project.status = "in_progress"
    db.commit()
    for file_record in created_files:
        db.refresh(file_record)

    file_stats = _get_file_segment_stats(db, [file_record.id for file_record in created_files])
    return {
        "id": str(project.id),
        "status": project.status,
        "filename": project.name,
        "uploaded_count": len(created_files),
        "files": [
            _json_ready_project_file_payload(
                _build_project_file_payload(
                    file_record=file_record,
                    total_segments=file_stats.get(file_record.id, {"total": 0})["total"],
                    translated_segments=file_stats.get(file_record.id, {"filled": 0})["filled"],
                )
            )
            for file_record in created_files
        ],
    }


def _run_import_task(task_id: str, payload: dict[str, Any]) -> None:
    _set_import_task_status(task_id, "running", progress=5, message="导入任务开始处理。")
    with SessionLocal() as db:
        try:
            if payload.get("kind") == "project_source_document":
                result = _process_project_source_import(db, task_id, payload)
            else:
                result = _process_file_record_import(db, payload)
            _set_import_task_status(
                task_id,
                "completed",
                progress=100,
                message="导入完成。",
                result=result,
            )
        except HTTPException as exc:
            db.rollback()
            _set_import_task_status(
                task_id,
                "failed",
                progress=100,
                message="导入失败。",
                error=str(exc.detail),
            )
        except Exception as exc:
            db.rollback()
            logger.exception("import task failed task_id=%s", task_id)
            _set_import_task_status(
                task_id,
                "failed",
                progress=100,
                message="导入失败。",
                error=str(exc),
            )


async def process_import_task_job(ctx, task_id: str, payload: dict[str, Any]) -> None:
    await asyncio.to_thread(_run_import_task, task_id, payload)


class WorkerSettings:
    functions = [process_import_task_job]
    redis_settings = _build_arq_redis_settings(get_settings().redis_url or "redis://localhost:6379/0")


class SegmentUpdate(BaseModel):
    sentence_id: str
    target_text: str
    target_html: str | None = None
    source: str = "manual"
    track_revision: bool = True
    base_version: int | None = None


class SegmentSourceUpdate(BaseModel):
    source_text: str


class BatchSegmentUpdate(BaseModel):
    updates: list[SegmentUpdate]


class SegmentConfirmationBatchUpdate(BaseModel):
    action: Literal["confirm", "cancel"]


class SegmentSplitRequest(BaseModel):
    """拆分句段请求"""
    split_offset: int = Field(..., ge=1, description="在 source_text 中的拆分位置（字符偏移）")


class SegmentMergeRequest(BaseModel):
    """合并句段请求"""
    target_sentence_id: str = Field(..., description="要合并的下一个句段的 sentence_id")


class SegmentReplaceRequest(BaseModel):
    scope: str = "all"
    source_query: str = ""
    target_query: str
    replace_text: str = ""
    search_fuzzy: bool = False
    replace_all: bool = True


class RevisionResolvePayload(BaseModel):
    status: Literal["accepted", "rejected"]


class FileOperationLockRequest(BaseModel):
    operation: Literal["pre_translate"] = PRE_TRANSLATE_OPERATION


class LLMTranslateRequest(BaseModel):
    scope: Literal["fuzzy_only", "none_only", "empty_target_only", "all", "all_with_exact"] = "all"
    provider: Literal["auto", "deepseek", "openrouter"] = "deepseek"
    model: str | None = Field(default=None, max_length=120)
    translation_unit: Literal["paragraph", "sentence"] = "paragraph"
    translation_guidelines: str = ""
    guideline_template_id: str | None = None
    temporary_prompt: str = ""


class TermExtractionRequest(BaseModel):
    term_base_id: UUID | None = None
    max_terms: int = Field(default=150, ge=1, le=300)
    models: list[str] = Field(default_factory=lambda: [TERM_EXTRACTION_MODEL])
    extraction_prompt: str = Field(default="", max_length=4000)


class GuidelineTemplateUpdateRequest(BaseModel):
    content: str


class RematchRequest(BaseModel):
    collection_ids: list[UUID]
    threshold: float = 0.75
    skip_confirmed: bool = True
    overwrite_fuzzy: bool = True
    auto_confirm_exact: bool = True


class FileRecordBindingsRequest(BaseModel):
    term_base_id: UUID | None = None
    term_base_ids: list[UUID] | None = None
    collection_id: UUID | None = None


class FileRecordDuplicateRequest(BaseModel):
    filename: str | None = None


class FileRecordAssignmentRequest(BaseModel):
    assignee_id: UUID | None = None


class SaveToTMRequest(BaseModel):
    collection_id: UUID | None = None
    collection_mode: Literal["existing", "new"] = "existing"
    collection_name: str | None = None
    scope: Literal["confirmed", "translated", "all"] = "translated"


class MemoryBasePayload(BaseModel):
    name: str
    description: str | None = None
    source_language: str
    target_language: str


class TMCollectionMergePayload(BaseModel):
    source_collection_ids: list[UUID]
    name: str
    description: str | None = None


class TermBasePayload(BaseModel):
    name: str
    description: str | None = None
    source_language: str = "zh"
    target_language: str = "en"


class TermPayload(BaseModel):
    source_text: str
    target_text: str
    collection_id: UUID | None = None


class CommentCreateRequest(BaseModel):
    sentence_id: str | None = None
    segment_id: UUID | None = None
    anchor_mode: Literal["sentence", "range"] = "range"
    range_start_offset: int | None = None
    range_end_offset: int | None = None
    anchor_text: str | None = None
    body: str


class CommentUpdateRequest(BaseModel):
    body: str | None = None
    status: Literal["open", "resolved"] | None = None


class CommentReplyRequest(BaseModel):
    body: str


class IssueMarkerCreateRequest(BaseModel):
    file_record_id: UUID | None = None
    title: str | None = None
    description: str
    category: Literal["bug", "translation", "format", "performance", "data", "other"] = "other"
    severity: Literal["low", "medium", "high", "critical"] = "medium"
    page_url: str | None = None
    user_agent: str | None = None


class IssueMarkerUpdateRequest(BaseModel):
    title: str | None = None
    description: str | None = None
    category: Literal["bug", "translation", "format", "performance", "data", "other"] | None = None
    severity: Literal["low", "medium", "high", "critical"] | None = None
    status: Literal["open", "resolved"] | None = None


class ProjectCreatePayload(BaseModel):
    name: str
    source_language: str | None = None
    target_language: str | None = None
    deadline: str | None = None
    access_level: Literal["team", "private", "public"] = "team"
    collection_id: UUID | None = None
    term_base_id: UUID | None = None


class ProjectUpdatePayload(BaseModel):
    name: str | None = None
    source_language: str | None = None
    target_language: str | None = None
    deadline: str | None = None
    access_level: Literal["team", "private", "public"] | None = None
    translation_guidelines: str | None = None


class ProjectDocumentStatisticsPayload(BaseModel):
    file_ids: list[UUID] = Field(default_factory=list)


def _can_manage_workflow(current_user: User | None) -> bool:
    return is_admin_role(getattr(current_user, "role", None))


def _can_read_file_record(file_record: FileRecord, current_user: User | None) -> bool:
    if current_user is None:
        return False
    if can_access_all_projects(current_user):
        return True
    return file_record.assignee_id == current_user.id


def _can_write_file_record(file_record: FileRecord, current_user: User | None) -> bool:
    return _can_read_file_record(file_record, current_user)


def _require_file_record_read_access(file_record: FileRecord, current_user: User) -> None:
    if not _can_read_file_record(file_record, current_user):
        raise HTTPException(status_code=404, detail="任务不存在或未分配给当前用户。")


def _require_file_record_work_access(file_record: FileRecord, current_user: User) -> None:
    if not _can_write_file_record(file_record, current_user):
        raise HTTPException(status_code=403, detail="当前账号没有处理该任务的权限。")


def _apply_project_visibility_filter(query, db: Session, current_user: User):
    if can_access_all_projects(current_user):
        return query
    assigned_project_ids = (
        db.query(FileRecord.project_id)
        .filter(
            FileRecord.assignee_id == current_user.id,
            FileRecord.project_id.is_not(None),
        )
        .distinct()
    )
    return query.filter(Project.id.in_(assigned_project_ids))


def _visible_project_files(files: list[FileRecord], current_user: User) -> list[FileRecord]:
    if can_access_all_projects(current_user):
        return files
    return [file_record for file_record in files if file_record.assignee_id == current_user.id]


def _serialize_assignee(user: User | None) -> dict[str, Any] | None:
    if user is None:
        return None
    return serialize_user(user)


@router.get("/analytics/dashboard")
def get_analytics_dashboard(
    background_tasks: BackgroundTasks,
    granularity: Literal["day", "month"] = "day",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not can_access_all_projects(current_user):
        raise HTTPException(status_code=403, detail="当前账号无权查看全局数据看板。")
    background_tasks.add_task(run_analytics_backfill_once)
    return get_dashboard_payload(db, granularity=granularity)


def _build_binary_download_response(
    filename: str,
    content: bytes,
    media_type: str,
) -> StreamingResponse:
    ascii_filename = filename.encode("ascii", "ignore").decode("ascii").strip() or "translated.bin"
    ascii_filename = ascii_filename.replace('"', "")
    quoted_filename = quote(filename)

    return StreamingResponse(
        BytesIO(content),
        media_type=media_type,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{ascii_filename}"; '
                f"filename*=UTF-8''{quoted_filename}"
            )
        },
    )


def _sse_event(event: str, payload: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"


def _serialize_guideline_template(template: GuidelineTemplate, include_content: bool = False) -> dict:
    payload = {
        "id": template.id,
        "name": template.name,
        "filename": template.filename,
        "size_bytes": template.size_bytes,
        "updated_at": template.updated_at.isoformat(),
        "content_preview": template.content[:180],
    }
    if include_content:
        payload["content"] = template.content
    return payload


def _resolve_llm_guidelines(
    db: Session,
    file_record: FileRecord,
    body: LLMTranslateRequest,
) -> str:
    project_guidelines = ""
    if file_record.project_id:
        project = db.query(Project).filter(Project.id == file_record.project_id).first()
        if project:
            project_guidelines = (project.translation_guidelines or "").strip()

    parts: list[str] = []
    if project_guidelines:
        parts.append(project_guidelines)

    if body.guideline_template_id:
        try:
            template = read_guideline_template(body.guideline_template_id)
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail="选择的翻译细则模板不存在。") from exc
        parts.append(f"【可复用细则：{template.name}】\n{template.content.strip()}")

    temporary_prompt = (body.temporary_prompt or "").strip()
    legacy_guidelines = (body.translation_guidelines or "").strip()
    if legacy_guidelines and not temporary_prompt:
        temporary_prompt = legacy_guidelines

    if temporary_prompt and temporary_prompt != project_guidelines:
        parts.append(f"【本次临时提示词】\n{temporary_prompt}")

    return "\n\n".join(part for part in parts if part.strip())


def _parse_optional_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    normalized = value.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        return None


def _build_llm_translation_tasks(
    db: Session,
    file_record_id: UUID,
    scope: Literal["fuzzy_only", "none_only", "empty_target_only", "all", "all_with_exact"],
    source_language: str | None = None,
    target_language: str | None = None,
    collection_id: UUID | None = None,
    include_context: bool = False,
) -> list[LLMTranslationTask]:
    statuses_by_scope = {
        "fuzzy_only": {"fuzzy"},
        "none_only": {"none"},
        "all": {"fuzzy", "none"},
        "all_with_exact": {"exact", "fuzzy", "none"},
    }
    target_statuses: set[str] | None = None
    if scope != "empty_target_only":
        target_statuses = statuses_by_scope.get(scope)
        if target_statuses is None:
            raise ValueError(f"不支持的 scope: {scope}")
    segments = list_segments_for_file_record(db, file_record_id)
    tm_target_text_map = get_tm_target_text_map(
        db,
        [segment.matched_source_text for segment in segments if segment.matched_source_text],
        collection_id=collection_id,
        source_language=source_language,
        target_language=target_language,
    )

    tasks: list[LLMTranslationTask] = []
    for segment in segments:
        should_translate = True
        if scope == "empty_target_only":
            if normalize_text(segment.target_text):
                should_translate = False
        elif target_statuses is None or segment.status not in target_statuses:
            should_translate = False

        if not should_translate and not include_context:
            continue

        segment_source = getattr(segment, "source", "none")
        matched_source_text = getattr(segment, "matched_source_text", None)
        segment_tm_target_text = segment.target_text if segment_source == "tm" and normalize_text(segment.target_text) else ""
        tm_target_text = segment_tm_target_text or tm_target_text_map.get(matched_source_text or "", "")

        tasks.append(
            LLMTranslationTask(
                sentence_id=segment.sentence_id,
                status=segment.status,
                source_text=segment.source_text,
                source_language=source_language,
                target_language=target_language,
                block_type=getattr(segment, "block_type", "paragraph") or "paragraph",
                block_index=int(getattr(segment, "block_index", 0) or 0),
                row_index=getattr(segment, "row_index", None),
                cell_index=getattr(segment, "cell_index", None),
                matched_source_text=matched_source_text,
                tm_target_text=tm_target_text,
                should_translate=should_translate,
            )
        )

    return tasks


@router.get("/guideline-templates")
def list_translation_guideline_templates():
    """列出仓库中可复用的 Markdown 翻译细则模板。"""
    return [
        _serialize_guideline_template(template)
        for template in list_guideline_templates()
    ]


@router.post("/guideline-templates/import")
async def import_translation_guideline_template(file: UploadFile = File(...)):
    """导入 .md/.txt 翻译细则，并统一保存为仓库内 UTF-8 Markdown。"""
    raw_bytes = await file.read()
    try:
        template = save_guideline_template(file.filename or "", raw_bytes)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _serialize_guideline_template(template, include_content=True)


@router.get("/guideline-templates/{template_id}")
def get_translation_guideline_template(template_id: str):
    try:
        template = read_guideline_template(template_id)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="翻译细则模板不存在。") from exc
    return _serialize_guideline_template(template, include_content=True)


@router.put("/guideline-templates/{template_id}")
def update_translation_guideline_template(
    template_id: str,
    payload: GuidelineTemplateUpdateRequest,
):
    try:
        template = update_guideline_template(template_id, payload.content)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="翻译细则模板不存在。") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _serialize_guideline_template(template, include_content=True)


@router.delete("/guideline-templates/{template_id}", status_code=204)
def delete_translation_guideline_template(template_id: str):
    try:
        delete_guideline_template(template_id)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="翻译细则模板不存在。") from exc
    return Response(status_code=204)


# 支持的文件扩展名（30种格式）
SUPPORTED_EXTENSIONS = {
    # 办公文档
    ".docx", ".txt", ".dat", ".pdf", ".pptx", ".xlsx",
    # 本地化文件
    ".properties", ".po", ".pot", ".strings", ".yaml", ".yml", ".json", ".php",
    # 网页/排版
    ".html", ".htm", ".md", ".markdown", ".csv", ".srt",
    # 技术写作
    ".dita", ".ditamap", ".xml", ".svg",
    # 双语文件
    ".sdlxliff", ".txml",
    # 工程/设计
    ".dxf", ".idml", ".mif",
    # 压缩包
    ".zip", ".rar",
}


def _get_file_extension(filename: str) -> str:
    """获取文件扩展名（小写）"""
    return Path(filename or "").suffix.lower()


def _validate_file_upload(file: UploadFile, allowed_extensions: set[str] | None = None) -> str:
    """验证上传的文件

    Args:
        file: 上传的文件
        allowed_extensions: 允许的扩展名集合，None 表示使用默认支持的扩展名

    Returns:
        str: 文件扩展名

    Raises:
        HTTPException: 当文件格式不支持时
    """
    ext = _get_file_extension(file.filename)
    allowed = allowed_extensions or SUPPORTED_EXTENSIONS

    if ext not in allowed:
        supported_list = ", ".join(sorted(allowed))
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件格式 '{ext}'。支持的格式: {supported_list}"
        )

    return ext


def _validate_docx_upload(file: UploadFile) -> None:
    """验证 DOCX 文件上传（向后兼容）"""
    _validate_file_upload(file, {".docx"})


def _validate_task_upload(file: UploadFile) -> None:
    if supports_task_file(file.filename or ""):
        return

    supported_extensions = ", ".join(get_supported_task_extensions())
    raise HTTPException(
        status_code=400,
        detail=f"暂不支持该文件格式。当前支持：{supported_extensions}",
    )


def _normalize_upload_document_parse_mode(document_parse_mode: str | None) -> str:
    try:
        return normalize_document_parse_mode(document_parse_mode)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


def _normalize_upload_document_parse_options(
    document_parse_options: str | None,
    document_parse_mode: str,
) -> dict[str, bool]:
    try:
        return normalize_document_parse_options(document_parse_options, document_parse_mode)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


def _normalize_collection_name(name: str) -> str:
    return " ".join(name.strip().split())


def _build_default_save_to_tm_collection_name(file_record: FileRecord) -> str:
    filename = (file_record.filename or "").strip()
    stem = Path(filename).stem.strip() if filename else ""
    base_name = stem or filename or "任务记忆库"
    return _normalize_collection_name(f"{base_name} 记忆库")


def _build_unique_memory_base_name(db: Session, name: str) -> str:
    base_name = _normalize_collection_name(name) or "任务记忆库"
    if len(base_name) > 255:
        base_name = base_name[:255].rstrip() or "任务记忆库"

    candidate = base_name
    suffix_index = 2
    while db.query(MemoryBase.id).filter(MemoryBase.name == candidate).first() is not None:
        suffix = f" ({suffix_index})"
        candidate_base = base_name[: 255 - len(suffix)].rstrip() or "任务记忆库"
        candidate = f"{candidate_base}{suffix}"
        suffix_index += 1

    return candidate


def _require_tm_language_pair(
    source_language: str | None,
    target_language: str | None,
) -> tuple[str, str]:
    try:
        return require_language_pair(source_language, target_language)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


def _resolve_collection_language_pair(
    collection: TMCollection | None,
    source_language: str | None,
    target_language: str | None,
) -> tuple[str, str]:
    if collection and collection.source_language and collection.target_language:
        if source_language or target_language:
            normalized_source_language, normalized_target_language = _require_tm_language_pair(
                source_language,
                target_language,
            )
            if (
                normalized_source_language != collection.source_language
                or normalized_target_language != collection.target_language
            ):
                raise HTTPException(status_code=400, detail="所选记忆库的语言对与本次提交不一致。")
        return collection.source_language, collection.target_language

    normalized_source_language, normalized_target_language = _require_tm_language_pair(
        source_language,
        target_language,
    )
    if collection is not None:
        collection.source_language = normalized_source_language
        collection.target_language = normalized_target_language
    return normalized_source_language, normalized_target_language


def _resolve_file_record_language_pair(file_record: FileRecord) -> tuple[str, str]:
    source_language = file_record.source_language
    target_language = file_record.target_language

    if (not source_language or not target_language) and file_record.collection:
        source_language = source_language or file_record.collection.source_language
        target_language = target_language or file_record.collection.target_language

    try:
        return require_language_pair(source_language, target_language)
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail="当前项目缺少源语言或目标语言，请先在项目任务中设置语言对。",
        ) from exc


def _apply_collection_language_pair(file_record: FileRecord, collection: TMCollection | None) -> None:
    if not collection:
        return
    if not file_record.source_language:
        file_record.source_language = collection.source_language
    if not file_record.target_language:
        file_record.target_language = collection.target_language


def _ensure_resource_language_pair_matches(
    resource,
    source_language: str,
    target_language: str,
    resource_label: str,
) -> None:
    resource_source_language = getattr(resource, "source_language", None)
    resource_target_language = getattr(resource, "target_language", None)
    if not resource_source_language or not resource_target_language:
        return
    if (
        resource_source_language != source_language
        or resource_target_language != target_language
    ):
        raise HTTPException(status_code=400, detail=f"所选{resource_label}的语言对与项目任务语言对不一致。")


def _validate_collection_ids(
    db: Session,
    collection_ids: list[UUID] | None,
) -> list[UUID] | None:
    if not collection_ids:
        return None

    normalized_ids = list(dict.fromkeys(collection_ids))
    existing_collections = (
        db.query(MemoryBase)
        .filter(MemoryBase.id.in_(normalized_ids))
        .all()
    )
    existing_ids = {collection.id for collection in existing_collections}
    missing_ids = [collection_id for collection_id in normalized_ids if collection_id not in existing_ids]
    if missing_ids:
        raise HTTPException(status_code=404, detail="选择的 TM 记忆库不存在。")

    return normalized_ids


def _validate_term_base_ids(
    db: Session,
    term_base_ids: list[UUID] | None,
) -> list[TermBase]:
    if not term_base_ids:
        return []

    normalized_ids = list(dict.fromkeys(term_base_ids))
    term_bases = (
        db.query(TermBase)
        .filter(TermBase.id.in_(normalized_ids))
        .all()
    )
    term_base_by_id = {term_base.id: term_base for term_base in term_bases}
    missing_ids = [term_base_id for term_base_id in normalized_ids if term_base_id not in term_base_by_id]
    if missing_ids:
        raise HTTPException(status_code=404, detail="选择的术语库不存在。")

    return [term_base_by_id[term_base_id] for term_base_id in normalized_ids]


def _require_selected_collection_ids(
    collection_ids: list[UUID] | None,
) -> list[UUID]:
    if not collection_ids:
        raise HTTPException(
            status_code=400,
            detail="请至少选择一个 TM 记忆库，避免全库模糊匹配拖慢处理进程。",
        )
    return collection_ids


def _resolve_upload_language_pair(
    source_language: str | None,
    target_language: str | None,
    primary_collection: TMCollection | None = None,
) -> tuple[str, str]:
    if source_language or target_language:
        resolved_source_language, resolved_target_language = _require_tm_language_pair(
            source_language,
            target_language,
        )
        _ensure_resource_language_pair_matches(
            primary_collection,
            resolved_source_language,
            resolved_target_language,
            "记忆库",
        )
        return resolved_source_language, resolved_target_language

    if primary_collection and primary_collection.source_language and primary_collection.target_language:
        return primary_collection.source_language, primary_collection.target_language

    raise HTTPException(status_code=400, detail="上传文件前请选择源语言和目标语言。")


def _get_collection_or_404(db: Session, collection_id: UUID | None) -> TMCollection | None:
    if collection_id is None:
        return None

    collection = db.query(MemoryBase).filter(MemoryBase.id == collection_id).first()
    if collection is None:
        raise HTTPException(status_code=404, detail="TM 记忆库不存在。")
    return collection


def _filter_tm_collection(
    query,
    collection_id: UUID | None,
    source_language: str | None = None,
    target_language: str | None = None,
):
    if collection_id is None:
        query = query.filter(TranslationMemory.collection_id.is_(None))
    else:
        query = query.filter(TranslationMemory.collection_id == collection_id)
    if source_language:
        query = query.filter(TranslationMemory.source_language == source_language)
    if target_language:
        query = query.filter(TranslationMemory.target_language == target_language)
    return query


def _serialize_tm_collection(collection: MemoryBase, entry_count: int = 0) -> dict:
    return {
        "id": collection.id,
        "name": collection.name,
        "description": collection.description,
        "source_language": collection.source_language,
        "target_language": collection.target_language,
        "created_at": collection.created_at.isoformat(),
        "updated_at": collection.updated_at.isoformat(),
        "entry_count": entry_count,
    }


def _require_same_tm_collection_language_pair(
    collections: list[MemoryBase],
) -> tuple[str, str]:
    try:
        source_language, target_language = require_language_pair(
            collections[0].source_language,
            collections[0].target_language,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="选中的记忆库缺少语言对，无法合并。") from exc

    for collection in collections[1:]:
        try:
            candidate_source_language, candidate_target_language = require_language_pair(
                collection.source_language,
                collection.target_language,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="选中的记忆库缺少语言对，无法合并。") from exc
        if (
            candidate_source_language != source_language
            or candidate_target_language != target_language
        ):
            raise HTTPException(status_code=400, detail="只能合并语言对完全一致的记忆库。")

    return source_language, target_language


@router.post("/parser/slate")
async def upload_for_slate(
    file: UploadFile = File(...),
    threshold: float = Form(default=0.6),
    collection_ids: list[UUID] | None = Form(default=None),
    db: Session = Depends(get_db),
):
    """上传文件并解析为 Slate 编辑器格式

    目前仅支持 DOCX 格式。
    """
    _validate_docx_upload(file)

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="文件为空。")

    selected_collection_ids = _validate_collection_ids(db, collection_ids)
    required_collection_ids = _require_selected_collection_ids(selected_collection_ids)
    try:
        result = parse_docx_for_slate(
            db=db,
            raw_bytes=raw_bytes,
            similarity_threshold=threshold,
            collection_ids=required_collection_ids,
        )
        return result
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.post("/parser/workspace")
async def upload_for_workspace(
    file: UploadFile = File(...),
    threshold: float = Form(default=0.6),
    collection_ids: list[UUID] | None = Form(default=None),
    db: Session = Depends(get_db),
):
    _validate_task_upload(file)

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="文件为空。")

    selected_collection_ids = _validate_collection_ids(db, collection_ids)
    required_collection_ids = _require_selected_collection_ids(selected_collection_ids)
    try:
        return await _build_task_workspace_async(
            raw_bytes=raw_bytes,
            filename=file.filename or "untitled.txt",
            similarity_threshold=threshold,
            collection_ids=required_collection_ids,
            document_parse_mode=DOCUMENT_PARSE_MODE_FULL,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.post("/parser/parse")
async def parse_document(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """通用文档解析接口

    使用适配器系统解析多种格式的文档。
    """
    ext = _validate_file_upload(file)

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="文件为空。")

    try:
        registry = get_registry()
        adapter = registry.get_adapter(file.filename)
        result = adapter.parse_with_validation(raw_bytes, file.filename)

        return {
            "filename": file.filename,
            "format": ext,
            "ast": result.ast.to_dict(),
            "segments": [seg.to_dict() for seg in result.segments],
            "metadata": result.metadata,
        }
    except UnsupportedFormatError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except FileTooLargeError as e:
        raise HTTPException(status_code=413, detail=str(e)) from e
    except ParseError as e:
        raise HTTPException(status_code=422, detail=str(e)) from e
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"解析失败: {str(exc)}") from exc


@router.get("/parser/formats")
async def get_supported_formats():
    """获取支持的文件格式列表"""
    registry = get_registry()

    formats = []
    for ext in sorted(registry.list_supported_extensions()):
        adapter = registry.get_adapter(f"test{ext}")
        formats.append({
            "extension": ext,
            "adapter": adapter.__class__.__name__,
            "max_size_mb": adapter.get_max_file_size() / (1024 * 1024),
        })

    return {
        "formats": formats,
        "total": len(formats),
    }


@router.get("/import-tasks/{task_id}")
def get_import_task(task_id: str):
    status = _get_import_task_status(task_id)
    if status is None:
        raise HTTPException(status_code=404, detail="导入任务不存在或已过期。")
    return status


# ============== 适配器导出相关模型 ==============

class AdapterExportRequest(BaseModel):
    """导出请求模型"""
    segments: List[dict]
    format: str = "txt"
    bilingual: bool = False
    filename: Optional[str] = None


class DitaExportRequest(BaseModel):
    """DITA 导出请求模型"""
    ast: dict
    translations: Dict[str, str]
    original_content: Optional[str] = None


class SvgExportRequest(BaseModel):
    """SVG 导出请求模型"""
    original_content: str
    translations: Dict[str, str]
    bilingual: bool = False


class TmxExportRequest(BaseModel):
    """TMX 导出请求模型"""
    segments: List[dict]
    source_lang: str = "zh-CN"
    target_lang: str = "en-US"
    filename: Optional[str] = None


class XliffExportRequest(BaseModel):
    """XLIFF 导出请求模型"""
    segments: List[dict]
    source_lang: str = "zh-CN"
    target_lang: str = "en-US"
    filename: str = "document"
    version: str = "1.2"


# ============== 适配器导出接口 ==============

@router.post("/export/txt")
async def export_txt(request: AdapterExportRequest):
    """导出为 TXT 格式"""
    try:
        service = ExportService()
        from app.services.adapters.models import BlockNode, NodeType
        nodes = []
        for seg in request.segments:
            text = seg.get("target_text") or seg.get("source_text", "")
            if text:
                nodes.append(BlockNode(node_type=NodeType.PARAGRAPH, text_content=text))

        ast = DocumentAST(nodes=nodes, source_format=".txt")
        translations = {
            seg.get("segment_id", f"seg_{i}"): seg.get("target_text", "")
            for i, seg in enumerate(request.segments)
        }

        if request.bilingual:
            nodes_bilingual = []
            for seg in request.segments:
                source = seg.get("source_text", "")
                if source:
                    nodes_bilingual.append(BlockNode(node_type=NodeType.PARAGRAPH, text_content=source))
            ast_bilingual = DocumentAST(nodes=nodes_bilingual, source_format=".txt")
            content = service.export_bilingual(ast_bilingual, translations, format="txt")
            filename = "bilingual_export.txt"
        else:
            content = service.export_txt(ast, translations)
            filename = "export.txt"

        return Response(
            content=content,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}") from e


@router.post("/export/docx")
async def export_adapter_docx(request: AdapterExportRequest):
    """通过适配器导出为 DOCX 格式"""
    try:
        service = ExportService()
        from app.services.adapters.models import BlockNode, NodeType
        nodes = []
        for seg in request.segments:
            text = seg.get("target_text") or seg.get("source_text", "")
            if text:
                nodes.append(BlockNode(node_type=NodeType.PARAGRAPH, text_content=text))

        ast = DocumentAST(nodes=nodes, source_format=".docx")
        translations = {
            seg.get("segment_id", f"seg_{i}"): seg.get("target_text", "")
            for i, seg in enumerate(request.segments)
        }

        if request.bilingual:
            nodes_bilingual = []
            for seg in request.segments:
                source = seg.get("source_text", "")
                if source:
                    nodes_bilingual.append(BlockNode(node_type=NodeType.PARAGRAPH, text_content=source))
            ast_bilingual = DocumentAST(nodes=nodes_bilingual, source_format=".docx")
            content = service.export_bilingual(ast_bilingual, translations, format="docx")
            filename = "bilingual_export.docx"
        else:
            content = service.export_docx(ast, translations)
            filename = "export.docx"

        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}") from e


@router.post("/export/dita")
async def export_dita(request: DitaExportRequest):
    """导出为 DITA 格式"""
    try:
        import base64
        exporter = DitaExporter()
        ast = DocumentAST.from_dict(request.ast)
        original_bytes = None
        if request.original_content:
            original_bytes = base64.b64decode(request.original_content)
        content = exporter.export(ast, request.translations, original_bytes)
        return Response(
            content=content,
            media_type="application/xml",
            headers={"Content-Disposition": 'attachment; filename="export.dita"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DITA 导出失败: {str(e)}") from e


@router.post("/export/svg")
async def export_svg(request: SvgExportRequest):
    """导出为 SVG 格式"""
    try:
        import base64
        exporter = SvgExporter()
        original_bytes = base64.b64decode(request.original_content)
        if request.bilingual:
            content, warnings = exporter.export_bilingual(original_bytes, request.translations)
            filename = "bilingual_export.svg"
        else:
            content, warnings = exporter.export(original_bytes, request.translations)
            filename = "export.svg"
        return Response(
            content=content,
            media_type="image/svg+xml",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "X-Export-Warnings": str(len(warnings)),
            },
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SVG 导出失败: {str(e)}") from e


@router.get("/export/formats")
async def get_export_formats():
    """获取支持的导出格式列表"""
    return {
        "formats": [
            {"id": "txt", "name": "纯文本 (TXT)", "extension": ".txt", "bilingual": True},
            {"id": "docx", "name": "Word 文档 (DOCX)", "extension": ".docx", "bilingual": True},
            {"id": "dita", "name": "DITA XML", "extension": ".dita", "bilingual": False},
            {"id": "svg", "name": "SVG 矢量图", "extension": ".svg", "bilingual": True},
            {"id": "tmx", "name": "翻译记忆库 (TMX)", "extension": ".tmx", "bilingual": False},
            {"id": "xliff", "name": "XLIFF 离线文件", "extension": ".xlf", "bilingual": False},
        ]
    }


@router.post("/export/tmx")
async def export_tmx(request: TmxExportRequest):
    """导出为 TMX 格式"""
    try:
        exporter = TmxExporter(source_lang=request.source_lang, target_lang=request.target_lang)
        content = exporter.export(request.segments, request.filename)
        filename = "export.tmx"
        if request.filename:
            base_name = request.filename.rsplit(".", 1)[0]
            filename = f"{base_name}.tmx"
        return Response(
            content=content,
            media_type="application/x-tmx+xml",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"TMX 导出失败: {str(e)}") from e


@router.post("/export/xliff")
async def export_xliff(request: XliffExportRequest):
    """导出为 XLIFF 格式"""
    try:
        exporter = XliffExporter(
            source_lang=request.source_lang,
            target_lang=request.target_lang,
            version=request.version,
        )
        original_format = "plaintext"
        if request.filename:
            ext = request.filename.rsplit(".", 1)[-1].lower()
            format_map = {
                "docx": "winword", "pdf": "pdf", "pptx": "powerpoint",
                "txt": "plaintext", "xml": "xml", "dita": "xml",
            }
            original_format = format_map.get(ext, "plaintext")
        content = exporter.export(request.segments, request.filename or "document", original_format)
        filename = "export.xlf"
        if request.filename:
            base_name = request.filename.rsplit(".", 1)[0]
            filename = f"{base_name}.xlf"
        return Response(
            content=content,
            media_type="application/xliff+xml",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"XLIFF 导出失败: {str(e)}") from e


@router.post("/import/xliff")
async def import_xliff(file: UploadFile = File(...)):
    """导入 XLIFF 文件"""
    if not file.filename or not file.filename.lower().endswith((".xlf", ".xliff")):
        raise HTTPException(status_code=400, detail="请上传 XLIFF 文件 (.xlf 或 .xliff)")
    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="文件为空")
    try:
        importer = XliffImporter()
        segments = importer.import_xliff(raw_bytes)
        return {"filename": file.filename, "segments": segments, "count": len(segments)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"XLIFF 导入失败: {str(e)}") from e


# ========== 文档管理 API ==========

@router.post("/file-records")
@router.post("/documents", include_in_schema=False)
async def create_file_record(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    threshold: float = Form(default=0.6),
    collection_ids: list[UUID] | None = Form(default=None),
    term_base_id: UUID | None = Form(default=None),
    source_language: str | None = Form(default=None),
    target_language: str | None = Form(default=None),
    document_parse_mode: str = Form(default=DOCUMENT_PARSE_MODE_FULL),
    document_parse_options: str | None = Form(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """上传文档并创建持久化记录"""
    _validate_task_upload(file)
    document_parse_mode = _normalize_upload_document_parse_mode(document_parse_mode)
    normalized_parse_options = _normalize_upload_document_parse_options(document_parse_options, document_parse_mode)

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="文件为空。")

    selected_collection_ids = _validate_collection_ids(db, collection_ids) or []
    primary_collection = _get_collection_or_404(
        db,
        selected_collection_ids[0] if selected_collection_ids else None,
    )
    resolved_source_language, resolved_target_language = _resolve_upload_language_pair(
        source_language,
        target_language,
        primary_collection,
    )

    # 验证术语库是否存在
    term_base = None
    if term_base_id is not None:
        term_base = db.query(TermBase).filter(TermBase.id == term_base_id).first()
        if term_base is None:
            raise HTTPException(status_code=404, detail="术语库不存在。")
        _ensure_resource_language_pair_matches(
            term_base,
            resolved_source_language,
            resolved_target_language,
            "术语库",
        )

    payload = {
        "kind": "file_record",
        "file": {
            "filename": file.filename or "untitled.txt",
            "content": raw_bytes,
        },
        "threshold": threshold,
        "collection_ids": [str(collection_id) for collection_id in selected_collection_ids],
        "term_base_id": str(term_base_id) if term_base_id is not None else None,
        "source_language": resolved_source_language,
        "target_language": resolved_target_language,
        "document_parse_mode": document_parse_mode,
        "document_parse_options": normalized_parse_options,
        "creator_id": str(current_user.id),
    }
    return await _queue_import_task(background_tasks, payload)


@router.post("/projects")
def create_project(
    payload: ProjectCreatePayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """仅填写基础信息创建项目，文档导入在项目详情页完成"""
    from datetime import datetime as _dt

    deadline_dt = None
    if payload.deadline:
        try:
            deadline_dt = _dt.fromisoformat(payload.deadline)
        except ValueError:
            raise HTTPException(status_code=400, detail="截止期限格式不正确，请使用 ISO 格式。")

    project_name = payload.name.strip()
    if not project_name:
        raise HTTPException(status_code=400, detail="项目名称不能为空。")

    source_language = None
    target_language = None
    if payload.source_language or payload.target_language:
        source_language, target_language = _require_tm_language_pair(
            payload.source_language,
            payload.target_language,
        )

    project = Project(
        name=project_name,
        status="draft",
        source_language=source_language,
        target_language=target_language,
        creator_id=current_user.id,
        deadline=deadline_dt,
        access_level=payload.access_level,
    )
    db.add(project)
    db.commit()
    db.refresh(project)

    return {
        "id": str(project.id),
        "name": project.name,
        "filename": project.name,
        "status": project.status,
        "source_language": project.source_language,
        "target_language": project.target_language,
        "creator": get_user_display_name(current_user),
        "deadline": project.deadline.isoformat() if project.deadline else None,
        "access_level": project.access_level,
        "created_at": project.created_at.isoformat(),
    }


def _build_project_summary_payload(
    project: Project,
    total_segments: int,
    translated_segments: int,
    file_count: int,
    creator_name: str | None = None,
    issue_stats: dict[str, int] | None = None,
    current_user: User | None = None,
) -> dict:
    progress = calculate_file_record_progress(total_segments, translated_segments)
    effective_status = (
        resolve_file_record_status("in_progress", total_segments, translated_segments)
        if total_segments > 0
        else project.status
    )
    issue_stats = issue_stats or {"issue_count": 0, "open_issue_count": 0}

    return {
        "id": str(project.id),
        "name": project.name,
        "filename": project.name,
        "status": effective_status,
        "progress": progress,
        "total_segments": total_segments,
        "translated_segments": translated_segments,
        "source_language": project.source_language,
        "target_language": project.target_language,
        "creator": creator_name,
        "deadline": project.deadline.isoformat() if project.deadline else None,
        "access_level": project.access_level,
        "translation_guidelines": project.translation_guidelines or "",
        "file_count": file_count,
        "issue_count": issue_stats.get("issue_count", 0),
        "open_issue_count": issue_stats.get("open_issue_count", 0),
        "can_manage": _can_manage_workflow(current_user),
        "can_write": bool(current_user) and (can_access_all_projects(current_user) or file_count > 0),
        "created_at": project.created_at.isoformat(),
        "updated_at": project.updated_at.isoformat(),
    }


def _build_project_file_payload(
    file_record: FileRecord,
    total_segments: int,
    translated_segments: int,
    issue_stats: dict[str, int] | None = None,
    current_user: User | None = None,
) -> dict:
    source_bytes = load_file_record_source(file_record)
    operation_state = serialize_file_operation_state(file_record)
    progress = calculate_file_record_progress(total_segments, translated_segments)
    effective_status = resolve_file_record_status(
        file_record.status,
        total_segments=total_segments,
        translated_segments=translated_segments,
    )
    issue_stats = issue_stats or {"issue_count": 0, "open_issue_count": 0}

    return {
        "id": str(file_record.id),
        "project_id": str(file_record.project_id) if file_record.project_id else None,
        "filename": file_record.filename,
        "status": effective_status,
        "document_parse_mode": getattr(file_record, "document_parse_mode", DOCUMENT_PARSE_MODE_FULL),
        "document_parse_options": _get_file_record_document_parse_options(file_record),
        "document_statistics": get_file_record_document_statistics(file_record),
        "progress": progress,
        "total_segments": total_segments,
        "translated_segments": translated_segments,
        "source_language": file_record.source_language,
        "target_language": file_record.target_language,
        "creator": get_user_display_name(file_record.creator) if file_record.creator else None,
        "assignee_id": str(file_record.assignee_id) if file_record.assignee_id else None,
        "assignee": _serialize_assignee(file_record.assignee),
        "assigned_at": file_record.assigned_at.isoformat() if file_record.assigned_at else None,
        "deadline": file_record.deadline.isoformat() if file_record.deadline else None,
        "access_level": file_record.access_level,
        "created_at": file_record.created_at.isoformat(),
        "updated_at": file_record.updated_at.isoformat(),
        "has_source_document": source_bytes is not None,
        "file_size_bytes": len(source_bytes) if source_bytes is not None else None,
        "collection_id": file_record.collection_id,
        "term_base_id": file_record.term_base_id,
        "issue_count": issue_stats.get("issue_count", 0),
        "open_issue_count": issue_stats.get("open_issue_count", 0),
        "can_manage": _can_manage_workflow(current_user),
        "can_write": _can_write_file_record(file_record, current_user),
        **operation_state,
    }


def _get_file_segment_stats(db: Session, file_record_ids: list[UUID]) -> dict[UUID, dict]:
    if not file_record_ids:
        return {}

    from sqlalchemy import case as sql_case

    stats_rows = (
        db.query(
            Segment.file_record_id,
            func.count(Segment.id).label("total"),
            func.count(sql_case((Segment.target_text != "", 1))).label("filled"),
        )
        .filter(Segment.file_record_id.in_(file_record_ids))
        .group_by(Segment.file_record_id)
        .all()
    )
    return {
        row.file_record_id: {
            "total": row.total,
            "filled": row.filled,
        }
        for row in stats_rows
    }


def _get_project_stats(
    db: Session,
    project_ids: list[UUID],
    current_user: User | None = None,
) -> dict[UUID, dict]:
    if not project_ids:
        return {}

    from sqlalchemy import case as sql_case

    query = (
        db.query(
            FileRecord.project_id,
            func.count(func.distinct(FileRecord.id)).label("file_count"),
            func.count(Segment.id).label("total"),
            func.count(sql_case((Segment.target_text != "", 1))).label("filled"),
        )
        .outerjoin(Segment, Segment.file_record_id == FileRecord.id)
        .filter(FileRecord.project_id.in_(project_ids))
    )
    if current_user is not None and is_external_translator(current_user):
        query = query.filter(FileRecord.assignee_id == current_user.id)
    stats_rows = query.group_by(FileRecord.project_id).all()
    return {
        row.project_id: {
            "file_count": row.file_count,
            "total": row.total,
            "filled": row.filled,
        }
        for row in stats_rows
    }


def _get_project_issue_stats(
    db: Session,
    project_ids: list[UUID],
    current_user: User | None = None,
) -> dict[UUID, dict[str, int]]:
    if not project_ids:
        return {}

    from sqlalchemy import case as sql_case

    query = (
        db.query(
            IssueMarker.project_id,
            func.count(IssueMarker.id).label("issue_count"),
            func.count(sql_case((IssueMarker.status == "open", 1))).label("open_issue_count"),
        )
        .filter(IssueMarker.project_id.in_(project_ids))
    )
    if current_user is not None and is_external_translator(current_user):
        assigned_file_ids = (
            db.query(FileRecord.id)
            .filter(
                FileRecord.project_id.in_(project_ids),
                FileRecord.assignee_id == current_user.id,
            )
        )
        query = query.filter(IssueMarker.file_record_id.in_(assigned_file_ids))
    stats_rows = query.group_by(IssueMarker.project_id).all()
    return {
        row.project_id: {
            "issue_count": int(row.issue_count or 0),
            "open_issue_count": int(row.open_issue_count or 0),
        }
        for row in stats_rows
    }


def _get_file_issue_stats(db: Session, file_record_ids: list[UUID]) -> dict[UUID, dict[str, int]]:
    if not file_record_ids:
        return {}

    from sqlalchemy import case as sql_case

    stats_rows = (
        db.query(
            IssueMarker.file_record_id,
            func.count(IssueMarker.id).label("issue_count"),
            func.count(sql_case((IssueMarker.status == "open", 1))).label("open_issue_count"),
        )
        .filter(IssueMarker.file_record_id.in_(file_record_ids))
        .group_by(IssueMarker.file_record_id)
        .all()
    )
    return {
        row.file_record_id: {
            "issue_count": int(row.issue_count or 0),
            "open_issue_count": int(row.open_issue_count or 0),
        }
        for row in stats_rows
        if row.file_record_id is not None
    }


def _build_project_detail_payload(
    project: Project,
    files: list[FileRecord],
    file_stats: dict[UUID, dict],
    issue_markers: list[IssueMarker] | None = None,
    project_issue_stats: dict[str, int] | None = None,
    file_issue_stats: dict[UUID, dict[str, int]] | None = None,
    current_user: User | None = None,
) -> dict:
    total_segments = sum(file_stats.get(file.id, {"total": 0})["total"] for file in files)
    translated_segments = sum(file_stats.get(file.id, {"filled": 0})["filled"] for file in files)
    file_issue_stats = file_issue_stats or {}
    payload = _build_project_summary_payload(
        project=project,
        total_segments=total_segments,
        translated_segments=translated_segments,
        file_count=len(files),
        creator_name=get_user_display_name(project.creator),
        issue_stats=project_issue_stats,
        current_user=current_user,
    )
    payload["files"] = [
        _build_project_file_payload(
            file_record=file_record,
            total_segments=file_stats.get(file_record.id, {"total": 0})["total"],
            translated_segments=file_stats.get(file_record.id, {"filled": 0})["filled"],
            issue_stats=file_issue_stats.get(file_record.id),
            current_user=current_user,
        )
        for file_record in files
    ]
    payload["issue_markers"] = [
        serialize_issue_marker(marker)
        for marker in (issue_markers or [])
    ]
    payload["has_source_document"] = any(file_item["has_source_document"] for file_item in payload["files"])
    payload["file_size_bytes"] = sum(
        file_item["file_size_bytes"] or 0
        for file_item in payload["files"]
    ) or None
    return payload


@router.get("/projects/{project_id}")
def get_project_detail(
    project_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from sqlalchemy.orm import joinedload

    project = (
        db.query(Project)
        .options(joinedload(Project.creator))
        .filter(Project.id == project_id)
        .first()
    )
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在。")

    files = (
        db.query(FileRecord)
        .options(joinedload(FileRecord.assignee))
        .filter(FileRecord.project_id == project_id)
        .order_by(FileRecord.created_at.asc(), FileRecord.id.asc())
        .all()
    )
    files = _visible_project_files(files, current_user)
    if is_external_translator(current_user) and not files:
        raise HTTPException(status_code=404, detail="项目不存在或没有分配给当前用户的任务。")
    stale_lock_cleared = False
    for file_record in files:
        stale_lock_cleared = clear_stale_file_operation_lock(db, file_record) or stale_lock_cleared
    if stale_lock_cleared:
        db.commit()
    file_stats = _get_file_segment_stats(db, [file_record.id for file_record in files])
    if is_external_translator(current_user):
        issue_markers = []
        for file_record in files:
            issue_markers.extend(list_issue_markers_for_project(
                db,
                project_id,
                file_record_id=file_record.id,
            ))
    else:
        issue_markers = list_issue_markers_for_project(db, project_id)
    file_issue_stats = _get_file_issue_stats(db, [file_record.id for file_record in files])
    if is_external_translator(current_user):
        project_issue_stats = {
            "issue_count": sum(item.get("issue_count", 0) for item in file_issue_stats.values()),
            "open_issue_count": sum(item.get("open_issue_count", 0) for item in file_issue_stats.values()),
        }
    else:
        project_issue_stats = _get_project_issue_stats(db, [project_id]).get(project_id)
    return _build_project_detail_payload(
        project,
        files,
        file_stats,
        issue_markers=issue_markers,
        project_issue_stats=project_issue_stats,
        file_issue_stats=file_issue_stats,
        current_user=current_user,
    )


@router.post("/projects/{project_id}/document-statistics")
def compute_project_document_statistics(
    project_id: UUID,
    payload: ProjectDocumentStatisticsPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在。")

    file_ids = list(dict.fromkeys(payload.file_ids))
    if not file_ids:
        raise HTTPException(status_code=400, detail="请先选择要统计的文件。")

    files = (
        db.query(FileRecord)
        .filter(FileRecord.project_id == project_id, FileRecord.id.in_(file_ids))
        .order_by(FileRecord.created_at.asc(), FileRecord.id.asc())
        .all()
    )
    if len(files) != len(file_ids):
        raise HTTPException(status_code=404, detail="部分文件不存在或不属于当前项目。")

    unavailable_statistics = {
        "source": "unavailable",
        "engine": None,
        "license_status": None,
        "include_textboxes_footnotes_endnotes": None,
        "pages": None,
        "words": None,
        "characters": None,
        "characters_with_spaces": None,
        "paragraphs": None,
        "lines": None,
    }

    for file_record in files:
        source_bytes = load_file_record_source(file_record)
        source_filename = get_file_record_source_filename(file_record)
        if source_bytes and Path(source_filename).suffix.lower() == ".docx":
            statistics = compute_docx_statistics(source_bytes)
        else:
            statistics = unavailable_statistics
        file_record.document_statistics = serialize_document_statistics(statistics)

    db.commit()
    for file_record in files:
        db.refresh(file_record)

    file_stats = _get_file_segment_stats(db, [file_record.id for file_record in files])
    file_issue_stats = _get_file_issue_stats(db, [file_record.id for file_record in files])
    return {
        "files": [
            _build_project_file_payload(
                file_record=file_record,
                total_segments=file_stats.get(file_record.id, {"total": 0})["total"],
                translated_segments=file_stats.get(file_record.id, {"filled": 0})["filled"],
                issue_stats=file_issue_stats.get(file_record.id),
                current_user=current_user,
            )
            for file_record in files
        ]
    }


@router.delete("/projects/{project_id}")
def delete_project(
    project_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在。")

    file_record_ids = [
        row.id
        for row in db.query(FileRecord.id).filter(FileRecord.project_id == project_id).all()
    ]
    for file_record_id in file_record_ids:
        delete_file_record(db, file_record_id)

    project = db.query(Project).filter(Project.id == project_id).first()
    if project is not None:
        db.delete(project)
        db.commit()

    return {"message": "项目已删除。"}


@router.patch("/projects/{project_id}")
def update_project(
    project_id: UUID,
    payload: ProjectUpdatePayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在。")

    if payload.name is not None:
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="项目名称不能为空。")
        project.name = name

    if {"source_language", "target_language"} & payload.model_fields_set:
        raise HTTPException(status_code=400, detail="项目语言对不允许修改。")

    if payload.deadline is not None:
        project.deadline = _parse_optional_datetime(payload.deadline)

    if payload.access_level is not None:
        project.access_level = payload.access_level

    if payload.translation_guidelines is not None:
        project.translation_guidelines = payload.translation_guidelines

    db.commit()
    db.refresh(project)

    return {
        "id": str(project.id),
        "name": project.name,
        "filename": project.name,
        "source_language": project.source_language,
        "target_language": project.target_language,
        "deadline": project.deadline.isoformat() if project.deadline else None,
        "access_level": project.access_level,
        "translation_guidelines": project.translation_guidelines or "",
        "updated_at": project.updated_at.isoformat(),
    }


@router.get("/projects/{project_id}/issue-markers")
def list_project_issue_markers(
    project_id: UUID,
    status: Literal["open", "resolved"] | None = None,
    file_record_id: UUID | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在。")
    if is_external_translator(current_user):
        assigned_file_ids = [
            row.id
            for row in (
                db.query(FileRecord.id)
                .filter(FileRecord.project_id == project_id, FileRecord.assignee_id == current_user.id)
                .all()
            )
        ]
        if not assigned_file_ids:
            raise HTTPException(status_code=404, detail="项目不存在或没有分配给当前用户的任务。")
        if file_record_id is not None and file_record_id not in assigned_file_ids:
            raise HTTPException(status_code=404, detail="任务不存在或未分配给当前用户。")
        if file_record_id is None:
            markers = []
            for assigned_file_id in assigned_file_ids:
                markers.extend(list_issue_markers_for_project(
                    db,
                    project_id,
                    status=status,
                    file_record_id=assigned_file_id,
                ))
            return [serialize_issue_marker(marker) for marker in markers]

    markers = list_issue_markers_for_project(
        db,
        project_id,
        status=status,
        file_record_id=file_record_id,
    )
    return [serialize_issue_marker(marker) for marker in markers]


@router.post("/projects/{project_id}/issue-markers")
def create_project_issue_marker(
    project_id: UUID,
    payload: IssueMarkerCreateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if is_external_translator(current_user):
        if payload.file_record_id is None:
            raise HTTPException(status_code=403, detail="外部译者只能在已分配任务上提交问题标记。")
        assigned = (
            db.query(FileRecord.id)
            .filter(
                FileRecord.project_id == project_id,
                FileRecord.id == payload.file_record_id,
                FileRecord.assignee_id == current_user.id,
            )
            .first()
        )
        if assigned is None:
            raise HTTPException(status_code=404, detail="任务不存在或未分配给当前用户。")
    marker = create_issue_marker(
        db,
        project_id=project_id,
        file_record_id=payload.file_record_id,
        title=payload.title,
        description=payload.description,
        category=payload.category,
        severity=payload.severity,
        page_url=payload.page_url,
        user_agent=payload.user_agent,
        reporter=current_user,
    )
    return serialize_issue_marker(marker)


@router.patch("/issue-markers/{marker_id}")
def patch_issue_marker(
    marker_id: UUID,
    payload: IssueMarkerUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    marker = update_issue_marker(
        db,
        marker_id=marker_id,
        title=payload.title,
        description=payload.description,
        category=payload.category,
        severity=payload.severity,
        status=payload.status,
        current_user=current_user,
    )
    return serialize_issue_marker(marker)


@router.delete("/issue-markers/{marker_id}")
def remove_issue_marker(
    marker_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    delete_issue_marker(
        db,
        marker_id=marker_id,
        current_user=current_user,
    )
    return {"message": "问题标记已删除。"}


@router.post("/projects/{project_id}/detect-source-language")
async def detect_project_source_language(
    project_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在。")

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="文件为空，无法识别语言。")

    result = detect_upload_language(file.filename or "", raw_bytes)
    return result.to_dict()


@router.post("/projects/{project_id}/source-document")
async def upload_project_source_document(
    project_id: UUID,
    background_tasks: BackgroundTasks,
    files: list[UploadFile] | None = File(default=None),
    file: UploadFile | None = File(default=None),
    threshold: float = Form(default=0.6),
    collection_ids: list[UUID] | None = Form(default=None),
    term_base_id: UUID | None = Form(default=None),
    source_language: str | None = Form(default=None),
    target_language: str | None = Form(default=None),
    document_parse_mode: str = Form(default=DOCUMENT_PARSE_MODE_FULL),
    document_parse_options: str | None = Form(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在。")

    uploaded_files = list(files or [])
    if file is not None:
        uploaded_files.append(file)
    if not uploaded_files:
        raise HTTPException(status_code=400, detail="请选择要上传的文件。")

    for upload_file in uploaded_files:
        _validate_task_upload(upload_file)
    document_parse_mode = _normalize_upload_document_parse_mode(document_parse_mode)
    normalized_parse_options = _normalize_upload_document_parse_options(document_parse_options, document_parse_mode)

    selected_collection_ids = _validate_collection_ids(db, collection_ids) or []
    primary_collection = _get_collection_or_404(
        db,
        selected_collection_ids[0] if selected_collection_ids else None,
    )
    resolved_source_language, resolved_target_language = _resolve_upload_language_pair(
        source_language,
        target_language,
        primary_collection,
    )

    term_base = None
    if term_base_id is not None:
        term_base = db.query(TermBase).filter(TermBase.id == term_base_id).first()
        if term_base is None:
            raise HTTPException(status_code=404, detail="术语库不存在。")
        _ensure_resource_language_pair_matches(
            term_base,
            resolved_source_language,
            resolved_target_language,
            "术语库",
        )

    queued_files = []
    for upload_file in uploaded_files:
        raw_bytes = await upload_file.read()
        if not raw_bytes:
            raise HTTPException(status_code=400, detail=f"{upload_file.filename or '文件'} 为空。")
        queued_files.append(
            {
                "filename": upload_file.filename or "source.txt",
                "content": raw_bytes,
            }
        )

    payload = {
        "kind": "project_source_document",
        "project_id": str(project.id),
        "files": queued_files,
        "threshold": threshold,
        "collection_ids": [str(collection_id) for collection_id in selected_collection_ids],
        "term_base_id": str(term_base_id) if term_base_id is not None else None,
        "source_language": resolved_source_language,
        "target_language": resolved_target_language,
        "document_parse_mode": document_parse_mode,
        "document_parse_options": normalized_parse_options,
    }
    return await _queue_import_task(background_tasks, payload)


@router.get("/projects")
def list_projects(
    skip: int = 0,
    limit: int = 50,
    search: str = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取项目列表（分页、搜索），含翻译进度统计"""
    from sqlalchemy.orm import joinedload

    base_query = db.query(Project).options(joinedload(Project.creator))
    base_query = _apply_project_visibility_filter(base_query, db, current_user)
    if search.strip():
        base_query = base_query.filter(Project.name.ilike(f"%{search.strip()}%"))

    total = base_query.count()
    safe_skip = max(skip, 0)
    safe_limit = min(max(limit, 1), 200)
    projects = (
        base_query
        .order_by(Project.created_at.desc())
        .offset(safe_skip)
        .limit(safe_limit)
        .all()
    )

    project_ids = [project.id for project in projects]
    project_stats = _get_project_stats(db, project_ids, current_user=current_user)
    project_issue_stats = _get_project_issue_stats(db, project_ids, current_user=current_user)

    items = []
    for project in projects:
        st = project_stats.get(project.id, {"file_count": 0, "total": 0, "filled": 0})
        total_segs = st["total"]
        filled_segs = st["filled"]

        creator_name = None
        if project.creator:
            creator_name = get_user_display_name(project.creator)

        items.append(
            _build_project_summary_payload(
                project=project,
                total_segments=total_segs,
                translated_segments=filled_segs,
                file_count=st["file_count"],
                creator_name=creator_name,
                issue_stats=project_issue_stats.get(project.id),
                current_user=current_user,
            )
        )

    return {
        "items": items,
        "total": total,
        "skip": safe_skip,
        "limit": safe_limit,
    }


@router.get("/file-records/upload-capabilities")
def get_file_record_upload_capabilities():
    """返回任务上传入口真实支持的格式和解析能力。"""
    return get_upload_capabilities()


@router.get("/file-records")
@router.get("/documents", include_in_schema=False)
def get_file_records(
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取文档列表"""
    from sqlalchemy.orm import joinedload

    safe_skip = max(skip, 0)
    safe_limit = min(max(limit, 1), 200)
    query = db.query(FileRecord).options(joinedload(FileRecord.assignee))
    if not can_access_all_projects(current_user):
        query = query.filter(FileRecord.assignee_id == current_user.id)
    file_records = (
        query
        .order_by(FileRecord.created_at.desc())
        .offset(safe_skip)
        .limit(safe_limit)
        .all()
    )
    file_stats = _get_file_segment_stats(db, [file_record.id for file_record in file_records])
    file_issue_stats = _get_file_issue_stats(db, [file_record.id for file_record in file_records])
    return [
        {
            "id": file_record.id,
            "project_id": str(file_record.project_id) if file_record.project_id else None,
            "filename": file_record.filename,
            "status": file_record.status,
            "progress": calculate_file_record_progress(
                file_stats.get(file_record.id, {"total": 0})["total"],
                file_stats.get(file_record.id, {"filled": 0})["filled"],
            ),
            "total_segments": file_stats.get(file_record.id, {"total": 0})["total"],
            "translated_segments": file_stats.get(file_record.id, {"filled": 0})["filled"],
            "issue_count": file_issue_stats.get(file_record.id, {}).get("issue_count", 0),
            "open_issue_count": file_issue_stats.get(file_record.id, {}).get("open_issue_count", 0),
            "document_parse_mode": getattr(file_record, "document_parse_mode", DOCUMENT_PARSE_MODE_FULL),
            "document_parse_options": _get_file_record_document_parse_options(file_record),
            "document_statistics": get_file_record_document_statistics(file_record),
            "source_language": file_record.source_language,
            "target_language": file_record.target_language,
            "assignee_id": str(file_record.assignee_id) if file_record.assignee_id else None,
            "assignee": _serialize_assignee(file_record.assignee),
            "assigned_at": file_record.assigned_at.isoformat() if file_record.assigned_at else None,
            "can_manage": _can_manage_workflow(current_user),
            "can_write": _can_write_file_record(file_record, current_user),
            "created_at": file_record.created_at.isoformat(),
            "updated_at": file_record.updated_at.isoformat(),
        }
        for file_record in file_records
    ]


SEGMENT_PAGE_MAX_LIMIT = 500


def _normalize_segment_page_limit(limit: int) -> int:
    return min(max(int(limit), 1), SEGMENT_PAGE_MAX_LIMIT)


def _serialize_workbench_segment(seg: Segment, display_index: int | None = None) -> dict:
    payload = {
        "id": str(seg.id),
        "sentence_id": seg.sentence_id,
        "source_text": seg.source_text,
        "display_text": seg.display_text,
        "source_html": seg.source_html,
        "target_text": seg.target_text,
        "target_html": seg.target_html,
        "status": seg.status,
        "version": int(seg.version or 1),
        "score": seg.score,
        "matched_source_text": seg.matched_source_text,
        "matched_collection_name": seg.matched_collection_name,
        "matched_creator_name": seg.matched_creator_name,
        "matched_created_at": seg.matched_created_at.isoformat() if seg.matched_created_at else None,
        "matched_updated_at": seg.matched_updated_at.isoformat() if seg.matched_updated_at else None,
        "source": seg.source,
        "llm_provider": seg.llm_provider,
        "llm_model": seg.llm_model,
        "block_type": seg.block_type,
        "block_index": seg.block_index,
        "row_index": seg.row_index,
        "cell_index": seg.cell_index,
        "updated_at": seg.updated_at.isoformat() if seg.updated_at else None,
    }
    if display_index is not None:
        payload["display_index"] = display_index
    return payload


def _serialize_segment_update_conflict(conflict) -> dict:
    return {
        "sentence_id": conflict.sentence_id,
        "current_version": conflict.current_version,
        "attempted_version": conflict.attempted_version,
        "current_target_text": conflict.current_target_text,
    }


def _empty_auto_tm_summary() -> AutoTMEnqueueSummary:
    return AutoTMEnqueueSummary()


def _schedule_auto_tm_processing(
    background_tasks: BackgroundTasks | None,
    summary: AutoTMEnqueueSummary,
) -> None:
    if background_tasks is not None and summary.queued_count > 0:
        background_tasks.add_task(run_auto_tm_background_once)


def _resolve_unconfirmed_segment_status(segment: Segment) -> str:
    if not normalize_text(segment.target_text):
        return "none"

    score = float(segment.score or 0)
    source_text = normalize_match_text(segment.source_text or "")
    matched_source_text = normalize_match_text(segment.matched_source_text or "")
    if score >= 0.999 or (matched_source_text and matched_source_text == source_text):
        return "exact"
    if score > 0 or matched_source_text:
        return "fuzzy"
    return "none"


def _apply_segment_scope_filter(query, scope: str):
    normalized_scope = (scope or "all").strip().lower()
    if normalized_scope == "exact_only":
        return query.filter(Segment.status == "exact")
    if normalized_scope == "fuzzy_only":
        return query.filter(Segment.status == "fuzzy")
    if normalized_scope == "none_only":
        return query.filter(Segment.status == "none")
    if normalized_scope == "confirmed_only":
        return query.filter(Segment.status == "confirmed")
    if normalized_scope == "empty_target":
        return query.filter(func.trim(Segment.target_text) == "")
    return query


def _apply_segment_text_filters(
    query,
    source_query: str | None,
    target_query: str | None,
):
    source_keyword = (source_query or "").strip()
    target_keyword = (target_query or "").strip()
    if source_keyword:
        source_pattern = f"%{source_keyword}%"
        query = query.filter(
            or_(
                Segment.source_text.ilike(source_pattern),
                Segment.display_text.ilike(source_pattern),
            )
        )
    if target_keyword:
        query = query.filter(Segment.target_text.ilike(f"%{target_keyword}%"))
    return query


def _order_segment_query(query):
    return query.order_by(
        Segment.block_index.asc(),
        Segment.row_index.asc().nullsfirst(),
        Segment.cell_index.asc().nullsfirst(),
        Segment.sentence_id.asc(),
    )


def _get_segment_display_index_map(
    db: Session,
    file_record_id: UUID,
    segments: list[Segment],
) -> dict[UUID, int]:
    segment_ids = [segment.id for segment in segments]
    if not segment_ids:
        return {}

    ordered_segments = (
        db.query(
            Segment.id.label("id"),
            func.row_number()
            .over(
                order_by=(
                    Segment.block_index.asc(),
                    Segment.row_index.asc().nullsfirst(),
                    Segment.cell_index.asc().nullsfirst(),
                    Segment.sentence_id.asc(),
                )
            )
            .label("display_index"),
        )
        .filter(Segment.file_record_id == file_record_id)
        .subquery()
    )
    rows = (
        db.query(ordered_segments.c.id, ordered_segments.c.display_index)
        .filter(ordered_segments.c.id.in_(segment_ids))
        .all()
    )
    return {row.id: int(row.display_index) - 1 for row in rows}


def _generate_split_sentence_id(original_id: str, db: Session, file_record_id: UUID) -> str:
    """为拆分生成新的 sentence_id，使用子编号方式（如 "5" → "5.1"）。"""
    base = original_id
    suffix = 1
    while True:
        candidate = f"{base}.{suffix}"
        exists = (
            db.query(Segment.id)
            .filter(Segment.file_record_id == file_record_id, Segment.sentence_id == candidate)
            .first()
        )
        if not exists:
            return candidate
        suffix += 1


def _is_cjk_text(text: str) -> bool:
    """判断文本是否主要为中日韩文字（决定合并时是否加空格）。"""
    if not text:
        return False
    cjk_count = sum(1 for ch in text if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f')
    return cjk_count > len(text) * 0.3


def _build_preview_render_segments(segments: list[Segment], mode: str) -> list[dict]:
    if mode != "target":
        return [_serialize_workbench_segment(segment) for segment in segments]

    rendered: list[dict] = []
    for segment in segments:
        item = _serialize_workbench_segment(segment)
        item["display_text"] = segment.target_text or segment.display_text or segment.source_text
        rendered.append(item)
    return rendered


def _get_segment_status_stats(db: Session, file_record_id: UUID) -> dict[str, int]:
    empty_target_expr = func.trim(func.coalesce(Segment.target_text, "")) == ""
    row = (
        db.query(
            func.count(Segment.id).label("total"),
            func.coalesce(func.sum(case((Segment.status == "exact", 1), else_=0)), 0).label("exact"),
            func.coalesce(func.sum(case((Segment.status == "fuzzy", 1), else_=0)), 0).label("fuzzy"),
            func.coalesce(func.sum(case((Segment.status == "none", 1), else_=0)), 0).label("none"),
            func.coalesce(func.sum(case((Segment.status == "confirmed", 1), else_=0)), 0).label("confirmed"),
            func.coalesce(func.sum(case((empty_target_expr, 1), else_=0)), 0).label("empty_target"),
        )
        .filter(Segment.file_record_id == file_record_id)
        .one()
    )
    return {
        "total": int(row.total or 0),
        "exact": int(row.exact or 0),
        "fuzzy": int(row.fuzzy or 0),
        "none": int(row.none or 0),
        "confirmed": int(row.confirmed or 0),
        "empty_target": int(row.empty_target or 0),
    }


def _get_segment_page_sentence_ids(
    db: Session,
    file_record_id: UUID,
    *,
    skip: int,
    limit: int,
    scope: str = "all",
    source_query: str | None = None,
    target_query: str | None = None,
) -> list[str]:
    safe_skip = max(skip, 0)
    safe_limit = _normalize_segment_page_limit(limit)
    query = db.query(Segment.sentence_id).filter(Segment.file_record_id == file_record_id)
    query = _apply_segment_scope_filter(query, scope)
    query = _apply_segment_text_filters(
        query,
        source_query=source_query,
        target_query=target_query,
    )
    return [
        sentence_id
        for (sentence_id,) in (
            _order_segment_query(query)
            .offset(safe_skip)
            .limit(safe_limit)
            .all()
        )
        if sentence_id
    ]


@router.get("/file-records/{file_record_id}")
@router.get("/documents/{file_record_id}", include_in_schema=False)
def get_file_record(
    file_record_id: UUID,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取文档详情及片段，支持分页"""
    safe_skip = max(skip, 0)
    safe_limit = _normalize_segment_page_limit(limit)
    result = get_file_record_with_segments(
        db,
        file_record_id,
        skip=safe_skip,
        limit=safe_limit,
    )
    if not result:
        raise HTTPException(status_code=404, detail="\u4efb\u52a1\u4e0d\u5b58\u5728")

    file_record = result["file_record"]
    _require_file_record_read_access(file_record, current_user)
    if clear_stale_file_operation_lock(db, file_record):
        db.commit()
        db.refresh(file_record)
    if backfill_file_record_source_html(db, file_record):
        db.commit()
        result = get_file_record_with_segments(
            db,
            file_record_id,
            skip=safe_skip,
            limit=safe_limit,
        )
        if not result:
            raise HTTPException(status_code=404, detail="\u4efb\u52a1\u4e0d\u5b58\u5728")
        file_record = result["file_record"]
    segments = result["segments"]
    source_bytes = load_file_record_source(file_record)
    source_filename = get_file_record_source_filename(file_record)

    # 获取绑定的库信息
    collection_name = None
    if file_record.collection:
        collection_name = file_record.collection.name
    term_base_name = None
    if file_record.term_base:
        term_base_name = file_record.term_base.name
    term_base_ids = _load_file_record_term_base_ids(file_record)
    term_base_names: list[str] = []
    if term_base_ids:
        term_bases = (
            db.query(TermBase)
            .filter(TermBase.id.in_(term_base_ids))
            .all()
        )
        term_base_by_id = {term_base.id: term_base for term_base in term_bases}
        term_base_names = [
            term_base_by_id[term_base_id].name
            for term_base_id in term_base_ids
            if term_base_id in term_base_by_id
        ]
        if term_base_name is None and term_base_names:
            term_base_name = term_base_names[0]

    project_guidelines = ""
    if file_record.project_id:
        project = db.query(Project).filter(Project.id == file_record.project_id).first()
        if project:
            project_guidelines = project.translation_guidelines or ""

    issue_stats = _get_file_issue_stats(db, [file_record.id]).get(
        file_record.id,
        {"issue_count": 0, "open_issue_count": 0},
    )

    return {
        "id": file_record.id,
        "project_id": str(file_record.project_id) if file_record.project_id else None,
        "filename": file_record.filename,
        "status": file_record.status,
        "document_parse_mode": getattr(file_record, "document_parse_mode", DOCUMENT_PARSE_MODE_FULL),
        "document_parse_options": _get_file_record_document_parse_options(file_record),
        "document_statistics": get_file_record_document_statistics(file_record),
        **serialize_file_operation_state(file_record),
        "source_language": file_record.source_language,
        "target_language": file_record.target_language,
        "assignee_id": str(file_record.assignee_id) if file_record.assignee_id else None,
        "assignee": _serialize_assignee(file_record.assignee),
        "assigned_at": file_record.assigned_at.isoformat() if file_record.assigned_at else None,
        "collection_id": file_record.collection_id,
        "collection_name": collection_name,
        "term_base_id": file_record.term_base_id,
        "term_base_name": term_base_name,
        "term_base_ids": [str(term_base_id) for term_base_id in term_base_ids],
        "term_base_names": term_base_names,
        "translation_guidelines": project_guidelines,
        "created_at": file_record.created_at.isoformat(),
        "updated_at": file_record.updated_at.isoformat(),
        "server_time": datetime.utcnow().isoformat(),
        "total_segments": result["total_segments"],
        "skip": result["skip"],
        "limit": result["limit"],
        "source_extension": get_task_file_extension(source_filename),
        "has_source_document": source_bytes is not None,
        "can_export": can_export_task_file(source_filename, has_source_file=source_bytes is not None),
        "can_manage": _can_manage_workflow(current_user),
        "can_write": _can_write_file_record(file_record, current_user),
        "issue_count": issue_stats["issue_count"],
        "open_issue_count": issue_stats["open_issue_count"],
        "status_stats": _get_segment_status_stats(db, file_record_id),
        "segments": [
            _serialize_workbench_segment(seg, display_index=result["skip"] + index)
            for index, seg in enumerate(segments)
        ],
    }


@router.post("/file-records/{file_record_id}/operation-lock")
def acquire_file_record_operation_lock(
    file_record_id: UUID,
    payload: FileOperationLockRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    existing_file_record = get_file_record_model(db, file_record_id)
    if not existing_file_record:
        raise HTTPException(status_code=404, detail="任务不存在。")
    _require_file_record_work_access(existing_file_record, current_user)
    file_record, token = acquire_file_operation_lock(
        db,
        file_record_id,
        operation=payload.operation,
        current_user=current_user,
    )
    return {
        "id": str(file_record.id),
        "token": token,
        **serialize_file_operation_state(file_record),
    }


@router.patch("/file-records/{file_record_id}/operation-lock")
def heartbeat_file_record_operation_lock(
    file_record_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文件不存在。")

    _require_file_record_work_access(file_record, current_user)
    heartbeat_file_operation_lock(
        db,
        file_record,
        operation_token=operation_token,
    )
    return {
        "id": str(file_record.id),
        **serialize_file_operation_state(file_record),
    }


@router.delete("/file-records/{file_record_id}/operation-lock")
def release_file_record_operation_lock(
    file_record_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文件不存在。")

    _require_file_record_work_access(file_record, current_user)
    release_file_operation_lock(
        db,
        file_record,
        operation_token=operation_token,
    )
    return {
        "id": str(file_record.id),
        **serialize_file_operation_state(file_record),
    }


@router.get("/file-records/{file_record_id}/segments")
def get_file_record_segments(
    file_record_id: UUID,
    skip: int = 0,
    limit: int = 100,
    scope: str = "all",
    source_query: str | None = None,
    target_query: str | None = None,
    search_fuzzy: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """分页获取工作台句段；搜索/筛选在服务端执行，避免前端加载全文。"""
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_read_access(file_record, current_user)
    safe_skip = max(skip, 0)
    safe_limit = _normalize_segment_page_limit(limit)
    base_query = db.query(Segment).filter(Segment.file_record_id == file_record_id)
    total_segments = base_query.count()
    filtered_query = _apply_segment_scope_filter(base_query, scope)
    filtered_query = _apply_segment_text_filters(
        filtered_query,
        source_query=source_query,
        target_query=target_query,
    )
    matched_segments = filtered_query.count()
    page_segments = (
        _order_segment_query(filtered_query)
        .offset(safe_skip)
        .limit(safe_limit)
        .all()
    )
    display_index_map = _get_segment_display_index_map(db, file_record_id, page_segments)

    return {
        "file_record_id": str(file_record_id),
        "total_segments": total_segments,
        "matched_segments": matched_segments,
        "status_stats": _get_segment_status_stats(db, file_record_id),
        "skip": safe_skip,
        "limit": safe_limit,
        "filters": {
            "scope": scope,
            "source_query": source_query or "",
            "target_query": target_query or "",
            "search_fuzzy": search_fuzzy,
        },
        "server_time": datetime.utcnow().isoformat(),
        "segments": [
            _serialize_workbench_segment(seg, display_index=display_index_map.get(seg.id))
            for seg in page_segments
        ],
    }


@router.get("/file-records/{file_record_id}/segments/changes")
def get_file_record_segment_changes(
    file_record_id: UUID,
    since: str,
    limit: int = 500,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文件不存在。")

    _require_file_record_read_access(file_record, current_user)
    try:
        since_dt = datetime.fromisoformat(since.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="since 参数不是有效时间。") from exc

    safe_limit = _normalize_segment_page_limit(limit)
    changed_segments = (
        _order_segment_query(
            db.query(Segment).filter(
                Segment.file_record_id == file_record_id,
                Segment.updated_at > since_dt,
            )
        )
        .limit(safe_limit)
        .all()
    )
    return {
        "file_record_id": str(file_record_id),
        "server_time": datetime.utcnow().isoformat(),
        "segments": [
            _serialize_workbench_segment(segment)
            for segment in changed_segments
        ],
    }


@router.post("/file-records/{file_record_id}/duplicate")
def duplicate_file_record_task(
    file_record_id: UUID,
    payload: FileRecordDuplicateRequest | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """复制一个文件任务，保留源文件、句段和当前译文状态。"""
    duplicate = duplicate_file_record(
        db,
        file_record_id,
        current_user=current_user,
        filename=payload.filename if payload else None,
    )
    if duplicate is None:
        raise HTTPException(status_code=404, detail="文档不存在。")

    db.commit()
    db.refresh(duplicate)
    file_stats = _get_file_segment_stats(db, [duplicate.id]).get(
        duplicate.id,
        {"total": 0, "filled": 0},
    )
    return _build_project_file_payload(
        duplicate,
        total_segments=file_stats["total"],
        translated_segments=file_stats["filled"],
        current_user=current_user,
    )


@router.patch("/file-records/{file_record_id}/assignment")
def assign_file_record_task(
    file_record_id: UUID,
    payload: FileRecordAssignmentRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="任务不存在。")

    assignee: User | None = None
    if payload.assignee_id is not None:
        assignee = get_user_by_id(db, payload.assignee_id)
        if assignee is None or not assignee.is_active or assignee.role != USER_ROLE:
            raise HTTPException(status_code=400, detail="只能指派给启用中的普通译者账号。")

    file_record.assignee_id = assignee.id if assignee else None
    file_record.assigned_by_id = current_user.id if assignee else None
    file_record.assigned_at = datetime.utcnow() if assignee else None
    db.commit()
    db.refresh(file_record)

    file_stats = _get_file_segment_stats(db, [file_record.id]).get(
        file_record.id,
        {"total": 0, "filled": 0},
    )
    issue_stats = _get_file_issue_stats(db, [file_record.id]).get(file_record.id)
    return _build_project_file_payload(
        file_record,
        total_segments=file_stats["total"],
        translated_segments=file_stats["filled"],
        issue_stats=issue_stats,
        current_user=current_user,
    )


@router.get("/file-records/{file_record_id}/preview")
@router.get("/documents/{file_record_id}/preview", include_in_schema=False)
def get_file_record_preview(
    file_record_id: UUID,
    skip: int = 0,
    limit: int = 100,
    mode: Literal["source", "target"] = "source",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_read_access(file_record, current_user)
    source_filename = get_file_record_source_filename(file_record)
    safe_skip = max(skip, 0)
    safe_limit = _normalize_segment_page_limit(limit)
    page_segments = (
        _order_segment_query(
            db.query(Segment).filter(Segment.file_record_id == file_record_id)
        )
        .offset(safe_skip)
        .limit(safe_limit)
        .all()
    )
    render_segments = _build_preview_render_segments(page_segments, mode)
    preview_html = build_task_preview_html(
        filename=source_filename,
        segments=render_segments,
        source_bytes=None,
        document_parse_mode=getattr(file_record, "document_parse_mode", DOCUMENT_PARSE_MODE_FULL),
        document_parse_options=_get_file_record_document_parse_options(file_record),
    )

    return {
        "id": str(file_record.id),
        "filename": file_record.filename,
        "source_extension": get_task_file_extension(source_filename),
        "supports_preview": bool(preview_html),
        "preview_html": preview_html,
        "preview_mode": "window",
        "render_mode": mode,
        "skip": safe_skip,
        "limit": safe_limit,
        "supports_full_preview": False,
    }


@router.get("/file-records/{file_record_id}/segments/{segment_ref}/tm-candidates")
def get_segment_tm_candidates(
    file_record_id: UUID,
    segment_ref: str,
    threshold: float = 0.6,
    max_candidates: int = 5,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取指定句段的 TM 匹配候选列表，兼容句段 UUID 和 sentence_id。"""
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_read_access(file_record, current_user)
    segment = None
    try:
        segment_uuid = UUID(segment_ref)
    except ValueError:
        segment_uuid = None

    if segment_uuid is not None:
        segment = db.query(Segment).filter(
            Segment.id == segment_uuid,
            Segment.file_record_id == file_record_id,
        ).first()

    if segment is None:
        segment = db.query(Segment).filter(
            Segment.file_record_id == file_record_id,
            Segment.sentence_id == segment_ref,
        ).first()

    if not segment:
        raise HTTPException(status_code=404, detail="句段不存在。")

    # 优先使用 collection_ids_json（多记忆库），回退到 collection_id（单记忆库）
    collection_ids: list[UUID] = []
    if file_record.collection_ids_json and file_record.collection_ids_json != "[]":
        try:
            parsed_ids = json.loads(file_record.collection_ids_json)
            collection_ids = [UUID(cid) for cid in parsed_ids if cid]
        except (json.JSONDecodeError, ValueError):
            pass
    if not collection_ids and file_record.collection_id:
        collection_ids = [file_record.collection_id]

    candidates = get_tm_candidates_for_text(
        db=db,
        source_text=segment.source_text,
        similarity_threshold=threshold,
        collection_ids=collection_ids,
        top_n=max_candidates,
    )

    return {
        "segment_id": str(segment.id),
        "sentence_id": segment.sentence_id,
        "source_text": segment.source_text,
        "candidates": [
            {
                "source_text": c.source_text,
                "target_text": c.target_text,
                "score": c.score,
                "diff_html": c.diff_html,
                "collection_name": c.collection_name,
                "creator_name": c.creator_name,
                "created_at": c.created_at,
                "updated_at": c.updated_at,
            }
            for c in candidates
        ],
    }


@router.post("/file-records/{file_record_id}/rematch")
def rematch_file_record(
    file_record_id: UUID,
    payload: RematchRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_work_access(file_record, current_user)
    ensure_file_record_write_allowed(db, file_record, operation_token=operation_token)
    selected_collection_ids = _require_selected_collection_ids(
        _validate_collection_ids(db, payload.collection_ids)
    )
    source_language, target_language = _resolve_file_record_language_pair(file_record)
    collections = (
        db.query(MemoryBase)
        .filter(MemoryBase.id.in_(selected_collection_ids))
        .all()
    )
    for collection in collections:
        _ensure_resource_language_pair_matches(collection, source_language, target_language, "记忆库")

    threshold = float(payload.threshold)
    if threshold < 0.5 or threshold > 1:
        raise HTTPException(status_code=400, detail="TM 匹配阈值必须在 0.50 到 1.00 之间。")

    segments = list_segments_for_file_record(db, file_record_id)
    if not segments:
        return {"exact": 0, "fuzzy": 0, "skipped": 0, "updated": 0}

    source_sentences = [segment.source_text for segment in segments]
    auxiliary_sentences = [segment.display_text for segment in segments]
    matches, _ = match_sentences_with_stats(
        db=db,
        sentences=source_sentences,
        auxiliary_sentences=auxiliary_sentences,
        similarity_threshold=threshold,
        collection_ids=selected_collection_ids,
    )

    exact_count = 0
    fuzzy_count = 0
    skipped_count = 0
    updated_count = 0

    for segment, match in zip(segments, matches, strict=False):
        before = (
            segment.target_text,
            segment.status,
            segment.score,
            segment.source,
            segment.matched_source_text,
            segment.matched_collection_name,
            segment.matched_creator_name,
            segment.matched_created_at,
            segment.matched_updated_at,
        )
        if payload.skip_confirmed and segment.status == "confirmed":
            skipped_count += 1
            continue

        segment.score = float(match.score or 0)
        segment.matched_source_text = match.matched_source_text
        segment.matched_collection_name = match.matched_collection_name
        segment.matched_creator_name = match.matched_creator_name
        segment.matched_created_at = _parse_optional_datetime(match.matched_created_at)
        segment.matched_updated_at = _parse_optional_datetime(match.matched_updated_at)

        if match.status == "exact" and match.target_text is not None:
            segment.target_text = match.target_text
            segment.source = "tm"
            segment.status = "confirmed" if payload.auto_confirm_exact else "exact"
            exact_count += 1
        elif match.status == "fuzzy" and match.target_text is not None:
            can_overwrite = payload.overwrite_fuzzy or (
                segment.status in {"none", "fuzzy"} and segment.source != "user"
            )
            if can_overwrite:
                segment.target_text = match.target_text
                segment.source = "tm"
                segment.status = "fuzzy"
                fuzzy_count += 1

        after = (
            segment.target_text,
            segment.status,
            segment.score,
            segment.source,
            segment.matched_source_text,
            segment.matched_collection_name,
            segment.matched_creator_name,
            segment.matched_created_at,
            segment.matched_updated_at,
        )
        if before != after:
            segment.version = int(segment.version or 1) + 1
            updated_count += 1

    # 保留当前文档绑定的记忆库，供右侧匹配面板查询 TM 候选。
    if selected_collection_ids:
        file_record.collection_id = selected_collection_ids[0]
        file_record.collection_ids_json = json.dumps([str(cid) for cid in selected_collection_ids])

    db.commit()
    return {
        "exact": exact_count,
        "fuzzy": fuzzy_count,
        "skipped": skipped_count,
        "updated": updated_count,
    }


@router.patch("/file-records/{file_record_id}/bindings")
def patch_file_record_bindings(
    file_record_id: UUID,
    payload: FileRecordBindingsRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_work_access(file_record, current_user)
    ensure_file_record_write_allowed(db, file_record, operation_token=operation_token)
    source_language, target_language = _resolve_file_record_language_pair(file_record)

    if "collection_id" in payload.model_fields_set:
        if payload.collection_id is None:
            file_record.collection_id = None
            file_record.collection_ids_json = "[]"
        else:
            collection = _get_collection_or_404(db, payload.collection_id)
            _ensure_resource_language_pair_matches(collection, source_language, target_language, "记忆库")
            file_record.collection_id = payload.collection_id
            # 同步更新 collection_ids_json，保持单记忆库场景的一致性
            file_record.collection_ids_json = json.dumps([str(payload.collection_id)])

    if "term_base_ids" in payload.model_fields_set:
        term_bases = _validate_term_base_ids(db, payload.term_base_ids)
        for term_base in term_bases:
            _ensure_resource_language_pair_matches(term_base, source_language, target_language, "术语库")
        _store_file_record_term_base_ids(file_record, [term_base.id for term_base in term_bases])
    elif "term_base_id" in payload.model_fields_set:
        if payload.term_base_id is None:
            _store_file_record_term_base_ids(file_record, [])
        else:
            term_base = db.query(TermBase).filter(TermBase.id == payload.term_base_id).first()
            if not term_base:
                raise HTTPException(status_code=404, detail="术语库不存在。")
            _ensure_resource_language_pair_matches(term_base, source_language, target_language, "术语库")
            _store_file_record_term_base_ids(file_record, [payload.term_base_id])

    db.commit()
    db.refresh(file_record)
    term_base_ids = _load_file_record_term_base_ids(file_record)
    return {
        "id": str(file_record.id),
        "collection_id": str(file_record.collection_id) if file_record.collection_id else None,
        "term_base_id": str(file_record.term_base_id) if file_record.term_base_id else None,
        "term_base_ids": [str(term_base_id) for term_base_id in term_base_ids],
    }


@router.get("/file-records/{file_record_id}/export")
@router.get("/documents/{file_record_id}/export", include_in_schema=False)
@router.get("/file-records/{file_record_id}/export-docx")
@router.get("/documents/{file_record_id}/export-docx", include_in_schema=False)
def export_file_record_docx(
    file_record_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _begin_repeatable_read_snapshot(db)
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="File record not found.")

    _require_file_record_read_access(file_record, current_user)
    raw_bytes = load_file_record_source(file_record)
    source_filename = get_file_record_source_filename(file_record)
    if not can_export_task_file(source_filename, has_source_file=raw_bytes is not None):
        raise HTTPException(status_code=400, detail="Current file format does not support original export yet.")

    segments = list_segments_for_file_record(db, file_record_id)
    try:
        exported_file = export_translated_task_file(
            raw_bytes=raw_bytes,
            filename=source_filename,
            segments=segments,
            document_parse_mode=getattr(file_record, "document_parse_mode", DOCUMENT_PARSE_MODE_FULL),
            document_parse_options=_get_file_record_document_parse_options(file_record),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return _build_binary_download_response(
        filename=exported_file.filename,
        content=exported_file.content,
        media_type=exported_file.media_type,
    )


@router.get("/file-records/{file_record_id}/export-options")
@router.get("/documents/{file_record_id}/export-options", include_in_schema=False)
def get_file_record_export_options(
    file_record_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取文件支持的导出格式选项"""
    from app.services.adapters import get_export_options_for_file

    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_read_access(file_record, current_user)
    options = get_export_options_for_file(file_record.filename)

    return {
        "file_record_id": str(file_record_id),
        "filename": file_record.filename,
        "export_options": options,
    }


@router.get("/file-records/{file_record_id}/export/{export_type}")
@router.get("/documents/{file_record_id}/export/{export_type}", include_in_schema=False)
def export_file_record_with_type(
    file_record_id: UUID,
    export_type: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """多格式导出接口 - 支持原格式、双语、TMX、XLIFF 等导出类型

    Args:
        file_record_id: 文件记录 ID
        export_type: 导出类型 (original, bilingual, bilingual_txt, tmx, xliff, xliff2)
    """
    from app.services.adapters import export_file

    _begin_repeatable_read_snapshot(db)
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_read_access(file_record, current_user)
    raw_bytes = load_file_record_source(file_record)
    segments = list_segments_for_file_record(db, file_record_id)

    # 原格式导出需要按原文件重新写回，避免走通用导出器丢失格式。
    if export_type == "original":
        source_filename = get_file_record_source_filename(file_record)
        if not can_export_task_file(source_filename, has_source_file=raw_bytes is not None):
            raise HTTPException(status_code=400, detail="Current file format does not support original export yet.")
        try:
            exported_file = export_translated_task_file(
                raw_bytes=raw_bytes,
                filename=source_filename,
                segments=segments,
                document_parse_mode=getattr(file_record, "document_parse_mode", DOCUMENT_PARSE_MODE_FULL),
                document_parse_options=_get_file_record_document_parse_options(file_record),
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"导出失败: {str(exc)}") from exc

        return _build_binary_download_response(
            filename=exported_file.filename,
            content=exported_file.content,
            media_type=exported_file.media_type,
        )

    if export_type in BILINGUAL_DOCX_LAYOUT_EXPORT_ORDERS:
        source_filename = get_file_record_source_filename(file_record)
        if get_task_file_extension(source_filename) != ".docx":
            raise HTTPException(status_code=400, detail="Only DOCX source files support layout-preserving bilingual Word export.")
        try:
            exported_file = export_bilingual_task_docx_with_layout(
                raw_bytes=raw_bytes,
                filename=source_filename,
                segments=segments,
                order=BILINGUAL_DOCX_LAYOUT_EXPORT_ORDERS[export_type],
                document_parse_mode=getattr(file_record, "document_parse_mode", DOCUMENT_PARSE_MODE_FULL),
                document_parse_options=_get_file_record_document_parse_options(file_record),
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"导出失败: {str(exc)}") from exc

        return _build_binary_download_response(
            filename=exported_file.filename,
            content=exported_file.content,
            media_type=exported_file.media_type,
        )

    # 其他导出格式使用通用句段列表。
    segment_dicts = [
        {
            "segment_id": seg.sentence_id,
            "source_text": seg.source_text,
            "target_text": seg.target_text,
            "status": seg.status,
            "matched_source_text": seg.matched_source_text,
        }
        for seg in segments
    ]

    try:
        exported_bytes, mime_type, export_filename = export_file(
            export_type=export_type,
            segments=segment_dicts,
            filename=file_record.filename,
            original_bytes=raw_bytes,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(exc)}") from exc

    # 构建下载响应
    ascii_filename = export_filename.encode("ascii", "ignore").decode("ascii").strip() or "exported"
    ascii_filename = ascii_filename.replace('"', "")
    quoted_filename = quote(export_filename)

    return StreamingResponse(
        BytesIO(exported_bytes),
        media_type=mime_type,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{ascii_filename}"; '
                f"filename*=UTF-8''{quoted_filename}"
            )
        },
    )


def _require_file_record_write_access(
    db: Session,
    file_record_id: UUID,
    current_user: User,
    operation_token: str | None = None,
) -> FileRecord:
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文件不存在。")
    _require_file_record_work_access(file_record, current_user)
    ensure_file_record_write_allowed(
        db,
        file_record,
        operation_token=operation_token,
    )
    return file_record


@router.put("/file-records/{file_record_id}/segments/{sentence_id}")
@router.put("/documents/{file_record_id}/segments/{sentence_id}", include_in_schema=False)
def update_segment(
    file_record_id: UUID,
    sentence_id: str,
    update: SegmentUpdate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    """更新单个片段的译文"""
    file_record = _require_file_record_write_access(db, file_record_id, current_user, operation_token)
    if update.base_version is not None:
        current_segment = (
            db.query(Segment)
            .filter(Segment.file_record_id == file_record_id, Segment.sentence_id == sentence_id)
            .first()
        )
        if not current_segment:
            raise HTTPException(status_code=404, detail="片段不存在。")
        current_version = int(current_segment.version or 1)
        if current_version != update.base_version:
            return {
                "updated_count": 0,
                "conflicts": [
                    {
                        "sentence_id": current_segment.sentence_id,
                        "current_version": current_version,
                        "attempted_version": update.base_version,
                        "current_target_text": current_segment.target_text or "",
                    }
                ],
                "auto_tm": _empty_auto_tm_summary().to_dict(),
                "segments": [_serialize_workbench_segment(current_segment)],
            }
    segment = update_segment_by_sentence_id(
        db=db,
        file_record_id=file_record_id,
        sentence_id=sentence_id,
        target_text=update.target_text,
        target_html=update.target_html,
        source=update.source,
        current_user=current_user,
        track_revision=update.track_revision,
    )
    if not segment:
        raise HTTPException(status_code=404, detail="片段不存在。")

    auto_tm_summary = _empty_auto_tm_summary()
    if segment.status == "confirmed":
        auto_tm_summary = enqueue_confirmed_segments_for_auto_tm(
            db,
            file_record=file_record,
            segments=[segment],
            current_user=current_user,
        )
        if auto_tm_summary.queued_count > 0:
            db.commit()
            _schedule_auto_tm_processing(background_tasks, auto_tm_summary)

    return {
        "id": segment.id,
        "sentence_id": segment.sentence_id,
        "target_text": segment.target_text,
        "status": segment.status,
        "source": segment.source,
        "version": int(segment.version or 1),
        "updated_at": segment.updated_at.isoformat() if segment.updated_at else None,
        "auto_tm": auto_tm_summary.to_dict(),
        "updated_count": 1,
        "conflicts": [],
        "segments": [_serialize_workbench_segment(segment)],
    }


@router.put("/file-records/{file_record_id}/segments/{sentence_id}/source")
@router.put("/documents/{file_record_id}/segments/{sentence_id}/source", include_in_schema=False)
def update_segment_source(
    file_record_id: UUID,
    sentence_id: str,
    update: SegmentSourceUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    """更新单个片段的原文"""
    _require_file_record_write_access(db, file_record_id, current_user, operation_token)
    segment = update_segment_source_text(
        db=db,
        file_record_id=file_record_id,
        sentence_id=sentence_id,
        source_text=update.source_text,
    )
    if not segment:
        raise HTTPException(status_code=404, detail="片段不存在。")

    return {
        "id": segment.id,
        "sentence_id": segment.sentence_id,
        "source_text": segment.source_text,
        "display_text": segment.display_text,
        "source_html": segment.source_html,
    }


@router.post("/file-records/{file_record_id}/segments/{sentence_id}/split")
def split_segment(
    file_record_id: UUID,
    sentence_id: str,
    payload: SegmentSplitRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    """拆分句段：在指定偏移位置将一个句段拆为两个。"""
    _require_file_record_write_access(db, file_record_id, current_user, operation_token)
    segment = (
        db.query(Segment)
        .filter(Segment.file_record_id == file_record_id, Segment.sentence_id == sentence_id)
        .first()
    )
    if not segment:
        raise HTTPException(status_code=404, detail="片段不存在。")

    source_text = segment.source_text or ""
    if payload.split_offset <= 0 or payload.split_offset >= len(source_text):
        raise HTTPException(status_code=400, detail="拆分位置无效，必须在文本范围内。")

    # 拆分 source_text
    first_source = source_text[:payload.split_offset].rstrip()
    second_source = source_text[payload.split_offset:].lstrip()
    if not first_source or not second_source:
        raise HTTPException(status_code=400, detail="拆分后不能产生空句段。")

    # 拆分 target_text（按同比例偏移，如果有译文的话）
    target_text = segment.target_text or ""
    target_html = segment.target_html
    first_target = ""
    second_target = ""
    if target_text.strip():
        # 按比例估算译文拆分点
        ratio = payload.split_offset / len(source_text)
        target_offset = round(ratio * len(target_text))
        first_target = target_text[:target_offset].rstrip()
        second_target = target_text[target_offset:].lstrip()

    # 生成新的 sentence_id：使用子编号方式
    new_sentence_id = _generate_split_sentence_id(sentence_id, db, file_record_id)

    # 更新原句段
    segment.source_text = first_source
    segment.display_text = first_source
    segment.source_html = None
    segment.target_text = first_target
    segment.target_html = None
    segment.status = "none" if not first_target.strip() else "confirmed"
    segment.score = 0.0
    segment.matched_source_text = None
    segment.matched_collection_name = None
    segment.matched_creator_name = None
    segment.matched_created_at = None
    segment.matched_updated_at = None

    # 创建新句段
    new_segment = Segment(
        file_record_id=file_record_id,
        sentence_id=new_sentence_id,
        source_text=second_source,
        display_text=second_source,
        source_html=None,
        target_text=second_target,
        target_html=None,
        status="none" if not second_target.strip() else "confirmed",
        score=0.0,
        source="manual",
        block_type=segment.block_type,
        block_index=segment.block_index,
        row_index=segment.row_index,
        cell_index=segment.cell_index,
    )
    db.add(new_segment)

    sync_file_record_status(db, file_record_id)
    db.commit()
    db.refresh(segment)
    db.refresh(new_segment)

    return {
        "first": _serialize_workbench_segment(segment),
        "second": _serialize_workbench_segment(new_segment),
    }


@router.post("/file-records/{file_record_id}/segments/{sentence_id}/merge")
def merge_segment(
    file_record_id: UUID,
    sentence_id: str,
    payload: SegmentMergeRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    """合并句段：将当前句段与指定的下一个句段合并为一个。"""
    _require_file_record_write_access(db, file_record_id, current_user, operation_token)
    first_seg = (
        db.query(Segment)
        .filter(Segment.file_record_id == file_record_id, Segment.sentence_id == sentence_id)
        .first()
    )
    if not first_seg:
        raise HTTPException(status_code=404, detail="片段不存在。")

    second_seg = (
        db.query(Segment)
        .filter(
            Segment.file_record_id == file_record_id,
            Segment.sentence_id == payload.target_sentence_id,
        )
        .first()
    )
    if not second_seg:
        raise HTTPException(status_code=404, detail="目标合并片段不存在。")

    # 校验两个句段必须在同一个 block 中
    if (
        first_seg.block_index != second_seg.block_index
        or first_seg.row_index != second_seg.row_index
        or first_seg.cell_index != second_seg.cell_index
    ):
        raise HTTPException(status_code=400, detail="只能合并同一区块内相邻的句段。")

    # 合并文本
    separator = "" if _is_cjk_text(first_seg.source_text) else " "
    merged_source = first_seg.source_text.rstrip() + separator + second_seg.source_text.lstrip()
    merged_target = ""
    if (first_seg.target_text or "").strip() or (second_seg.target_text or "").strip():
        merged_target = (first_seg.target_text or "").rstrip() + separator + (second_seg.target_text or "").lstrip()

    # 更新第一个句段
    first_seg.source_text = merged_source.strip()
    first_seg.display_text = merged_source.strip()
    first_seg.source_html = None
    first_seg.target_text = merged_target.strip()
    first_seg.target_html = None
    first_seg.status = "none" if not merged_target.strip() else "confirmed"
    first_seg.score = 0.0
    first_seg.source = "manual"
    first_seg.matched_source_text = None
    first_seg.matched_collection_name = None
    first_seg.matched_creator_name = None
    first_seg.matched_created_at = None
    first_seg.matched_updated_at = None

    # 迁移第二个句段的评论到第一个句段
    for comment in second_seg.comments:
        comment.segment_id = first_seg.id

    # 删除第二个句段的修订记录
    db.query(SegmentRevision).filter(SegmentRevision.segment_id == second_seg.id).delete()

    # 删除第二个句段
    db.delete(second_seg)

    sync_file_record_status(db, file_record_id)
    db.commit()
    db.refresh(first_seg)

    return {
        "merged": _serialize_workbench_segment(first_seg),
        "deleted_sentence_id": payload.target_sentence_id,
    }


@router.put("/file-records/{file_record_id}/segments")
@router.put("/documents/{file_record_id}/segments", include_in_schema=False)
def batch_update(
    file_record_id: UUID,
    batch: BatchSegmentUpdate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    """批量更新片段译文"""
    file_record = _require_file_record_write_access(db, file_record_id, current_user, operation_token)
    result = batch_update_segments(
        db=db,
        file_record_id=file_record_id,
        updates=[u.model_dump() for u in batch.updates],
        current_user=current_user,
        return_result=True,
    )
    auto_tm_summary = enqueue_confirmed_segments_for_auto_tm(
        db,
        file_record=file_record,
        segments=result.updated_segments,
        current_user=current_user,
    )
    if auto_tm_summary.queued_count > 0:
        db.commit()
        _schedule_auto_tm_processing(background_tasks, auto_tm_summary)
    return {
        "updated_count": result.updated_count,
        "conflicts": [_serialize_segment_update_conflict(conflict) for conflict in result.conflicts],
        "auto_tm": auto_tm_summary.to_dict(),
        "segments": [_serialize_workbench_segment(segment) for segment in result.updated_segments],
    }


@router.post("/file-records/{file_record_id}/segments/confirmation")
@router.post("/documents/{file_record_id}/segments/confirmation", include_in_schema=False)
def batch_update_segment_confirmation(
    file_record_id: UUID,
    payload: SegmentConfirmationBatchUpdate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    """批量确认或取消确认当前文件的全部句段。"""
    file_record = _require_file_record_write_access(db, file_record_id, current_user, operation_token)

    if payload.action == "confirm":
        segments = (
            db.query(Segment)
            .filter(
                Segment.file_record_id == file_record_id,
                or_(Segment.status.is_(None), Segment.status != "confirmed"),
            )
            .all()
        )
        next_status = "confirmed"
        updated_count = 0
        for segment in segments:
            if segment.status != next_status:
                segment.status = next_status
                segment.version = int(segment.version or 1) + 1
                updated_count += 1
    else:
        segments = (
            db.query(Segment)
            .filter(Segment.file_record_id == file_record_id, Segment.status == "confirmed")
            .all()
        )
        updated_count = 0
        for segment in segments:
            next_status = _resolve_unconfirmed_segment_status(segment)
            if segment.status != next_status:
                segment.status = next_status
                segment.version = int(segment.version or 1) + 1
                updated_count += 1

    auto_tm_summary = _empty_auto_tm_summary()
    if payload.action == "confirm" and updated_count:
        auto_tm_summary = enqueue_confirmed_segments_for_auto_tm(
            db,
            file_record=file_record,
            segments=segments,
            current_user=current_user,
        )

    if updated_count:
        db.commit()
        _schedule_auto_tm_processing(background_tasks, auto_tm_summary)

    return {"updated_count": updated_count, "auto_tm": auto_tm_summary.to_dict()}


@router.post("/file-records/{file_record_id}/segments/replace")
def replace_file_record_segment_targets(
    file_record_id: UUID,
    payload: SegmentReplaceRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    """在服务端按筛选条件替换译文，避免前端加载全文后再逐条修改。"""
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_write_access(db, file_record_id, current_user, operation_token)
    target_query = (payload.target_query or "").strip()
    if not target_query:
        raise HTTPException(status_code=400, detail="请先输入译文关键词用于替换。")

    query = db.query(Segment).filter(Segment.file_record_id == file_record_id)
    query = _apply_segment_scope_filter(query, payload.scope)
    query = _apply_segment_text_filters(
        query,
        source_query=payload.source_query,
        target_query=target_query,
    )
    segments = _order_segment_query(query).all()
    if not segments:
        return {"updated_count": 0, "occurrence_count": 0}

    flags = re.IGNORECASE
    pattern = re.compile(re.escape(target_query), flags)
    occurrence_count = 0
    updates: list[dict[str, str]] = []
    for segment in segments:
        target_text = segment.target_text or ""
        next_text, count = pattern.subn(payload.replace_text or "", target_text)
        if count <= 0 or next_text == target_text:
            continue
        occurrence_count += count
        updates.append({
            "sentence_id": segment.sentence_id,
            "target_text": next_text,
            "source": "manual",
        })
        if not payload.replace_all:
            break

    if not updates:
        return {"updated_count": 0, "occurrence_count": 0}

    updated_count = batch_update_segments(
        db=db,
        file_record_id=file_record_id,
        updates=updates,
        current_user=current_user,
    )
    return {"updated_count": updated_count, "occurrence_count": occurrence_count}


@router.get("/file-records/{file_record_id}/save-to-tm/stats")
def get_save_to_tm_stats(
    file_record_id: UUID,
    scope: Literal["confirmed", "translated", "all"] = "translated",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_read_access(file_record, current_user)
    total_segments = (
        db.query(func.count(Segment.id))
        .filter(Segment.file_record_id == file_record_id)
        .scalar()
        or 0
    )
    query = db.query(Segment).filter(Segment.file_record_id == file_record_id)
    if scope == "confirmed":
        query = query.filter(Segment.status == "confirmed")
    elif scope == "translated":
        query = query.filter(func.trim(Segment.target_text) != "")

    matched_count = query.count()
    valid_count = (
        query.filter(
            func.trim(Segment.source_text) != "",
            func.trim(Segment.target_text) != "",
        )
        .count()
    )
    return {
        "total_segments": int(total_segments),
        "matched_count": matched_count,
        "valid_count": valid_count,
        "skipped_count": max(int(total_segments) - valid_count, 0),
    }


@router.post("/file-records/{file_record_id}/save-to-tm")
def save_file_record_segments_to_tm(
    file_record_id: UUID,
    payload: SaveToTMRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    source_language, target_language = _resolve_file_record_language_pair(file_record)

    collection: TMCollection | None = None
    created_collection = False
    if payload.collection_mode == "existing":
        target_collection_id = payload.collection_id or file_record.collection_id
        if target_collection_id is None:
            raise HTTPException(status_code=400, detail="请先选择要追加的目标记忆库。")
        collection = _get_collection_or_404(db, target_collection_id)
        _resolve_collection_language_pair(collection, source_language, target_language)
    elif payload.collection_mode != "new":
        raise HTTPException(status_code=400, detail="不支持的记忆库保存方式。")

    segments = list_segments_for_file_record(db, file_record_id)
    skipped_count = 0
    save_rows: list[tuple[str, str]] = []

    for segment in segments:
        if payload.scope == "confirmed" and segment.status != "confirmed":
            skipped_count += 1
            continue
        if payload.scope == "translated" and not normalize_text(segment.target_text):
            skipped_count += 1
            continue

        source_text = normalize_text(segment.source_text)
        target_text = normalize_text(segment.target_text)
        if not source_text or not target_text:
            skipped_count += 1
            continue

        save_rows.append((source_text, target_text))

    if payload.collection_mode == "new" and save_rows:
        requested_name = payload.collection_name or _build_default_save_to_tm_collection_name(file_record)
        collection_name = _build_unique_memory_base_name(db, requested_name)
        collection = MemoryBase(
            name=collection_name,
            description=f"由任务「{file_record.filename}」保存生成",
            source_language=source_language,
            target_language=target_language,
        )
        db.add(collection)
        db.flush()
        created_collection = True

    if collection is None:
        skipped_count += len(save_rows)
        upsert_summary = None
    else:
        upsert_summary = batch_upsert_tm_entries(
            db,
            [
                TMUpsertEntry(
                    collection_id=collection.id,
                    source_text=source_text,
                    target_text=target_text,
                    source_language=source_language,
                    target_language=target_language,
                    creator_id=current_user.id,
                )
                for source_text, target_text in save_rows
            ],
        )
        skipped_count += upsert_summary.skipped_count

    if created_collection or (upsert_summary is not None and upsert_summary.total_written > 0):
        db.commit()
        sync_tm_embeddings(db, upsert_summary.sync_rows if upsert_summary is not None else [])

    created_count = upsert_summary.created_count if upsert_summary is not None else 0
    updated_count = upsert_summary.updated_count if upsert_summary is not None else 0

    return {
        "created_count": created_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "total_segments": len(segments),
        "collection_id": collection.id if collection else None,
        "collection_name": collection.name if collection else None,
        "created_collection": created_collection,
    }


@router.get("/file-records/{file_record_id}/revisions")
def get_file_record_revisions(
    file_record_id: UUID,
    sentence_id: str | None = None,
    sentence_ids: list[str] | None = Query(default=None),
    sentence_ids_bracket: list[str] | None = Query(default=None, alias="sentence_ids[]"),
    skip: int = 0,
    limit: int | None = None,
    scope: str = "all",
    source_query: str | None = None,
    target_query: str | None = None,
    search_fuzzy: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="File record not found.")

    _require_file_record_read_access(file_record, current_user)
    requested_sentence_ids = [item for item in ((sentence_ids or []) + (sentence_ids_bracket or [])) if item]
    if sentence_id:
        requested_sentence_ids.append(sentence_id)

    if requested_sentence_ids:
        revisions = list_revisions(
            db,
            file_record_id=file_record_id,
            sentence_ids=list(dict.fromkeys(requested_sentence_ids)),
        )
        return [serialize_segment_revision(revision) for revision in revisions]

    if limit is not None:
        page_sentence_ids = _get_segment_page_sentence_ids(
            db,
            file_record_id,
            skip=skip,
            limit=limit,
            scope=scope,
            source_query=source_query,
            target_query=target_query,
        )
        revisions = list_revisions(
            db,
            file_record_id=file_record_id,
            sentence_ids=page_sentence_ids,
        )
        return [serialize_segment_revision(revision) for revision in revisions]

    revisions = list_revisions(
        db,
        file_record_id=file_record_id,
    )
    return [serialize_segment_revision(revision) for revision in revisions]


@router.patch("/revisions/{revision_id}")
def resolve_revision(
    revision_id: UUID,
    payload: RevisionResolvePayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    existing_revision = get_revision_or_404(db, revision_id)
    _require_file_record_write_access(db, existing_revision.file_record_id, current_user)
    if payload.status == "accepted":
        revision = accept_revision(
            db,
            revision_id=revision_id,
            current_user=current_user,
        )
    else:
        revision = reject_revision(
            db,
            revision_id=revision_id,
            current_user=current_user,
        )
    return serialize_segment_revision(revision)


@router.post("/file-records/{file_record_id}/revisions/batch-accept")
def resolve_all_revisions_as_accepted(
    file_record_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_file_record_write_access(db, file_record_id, current_user)

    updated_count = batch_accept_revisions(
        db,
        file_record_id=file_record_id,
        current_user=current_user,
    )
    return {"updated_count": updated_count}


@router.post("/file-records/{file_record_id}/revisions/batch-reject")
def resolve_all_revisions_as_rejected(
    file_record_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _require_file_record_write_access(db, file_record_id, current_user)

    updated_count = batch_reject_revisions(
        db,
        file_record_id=file_record_id,
        current_user=current_user,
    )
    return {"updated_count": updated_count}


@router.get("/file-records/{file_record_id}/comments")
@router.get("/documents/{file_record_id}/comments", include_in_schema=False)
def get_file_record_comments(
    file_record_id: UUID,
    sentence_ids: list[str] | None = Query(default=None),
    sentence_ids_bracket: list[str] | None = Query(default=None, alias="sentence_ids[]"),
    skip: int = 0,
    limit: int | None = None,
    scope: str = "all",
    source_query: str | None = None,
    target_query: str | None = None,
    search_fuzzy: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    requested_sentence_ids = [item for item in ((sentence_ids or []) + (sentence_ids_bracket or [])) if item]
    _require_file_record_read_access(file_record, current_user)
    if limit is not None and not requested_sentence_ids:
        requested_sentence_ids = _get_segment_page_sentence_ids(
            db,
            file_record_id,
            skip=skip,
            limit=limit,
            scope=scope,
            source_query=source_query,
            target_query=target_query,
        )

    comments = list_segment_comments_for_file_record(
        db,
        file_record_id,
        sentence_ids=requested_sentence_ids if (limit is not None or requested_sentence_ids) else None,
    )
    return [serialize_segment_comment(comment) for comment in comments]


@router.post("/file-records/{file_record_id}/comments")
@router.post("/documents/{file_record_id}/comments", include_in_schema=False)
def create_file_record_comment(
    file_record_id: UUID,
    payload: CommentCreateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_read_access(file_record, current_user)
    comment = create_segment_comment(
        db,
        file_record_id=file_record_id,
        sentence_id=payload.sentence_id,
        segment_id=payload.segment_id,
        anchor_mode=payload.anchor_mode,
        range_start_offset=payload.range_start_offset,
        range_end_offset=payload.range_end_offset,
        anchor_text=payload.anchor_text,
        body=payload.body,
        author=current_user,
    )
    return serialize_segment_comment(comment)


@router.patch("/comments/{comment_id}")
def patch_comment(
    comment_id: UUID,
    payload: CommentUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    comment = update_segment_comment(
        db,
        comment_id=comment_id,
        body=payload.body,
        status=payload.status,
        current_user=current_user,
    )
    return serialize_segment_comment(comment)


@router.delete("/comments/{comment_id}")
def remove_comment(
    comment_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    delete_segment_comment(
        db,
        comment_id=comment_id,
        current_user=current_user,
    )
    return {"message": "批注已删除。"}


@router.post("/comments/{comment_id}/replies")
def create_comment_reply(
    comment_id: UUID,
    payload: CommentReplyRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    comment = create_segment_comment_reply(
        db,
        comment_id=comment_id,
        body=payload.body,
        author=current_user,
    )
    return serialize_segment_comment(comment)


def _normalize_term_extraction_models(models: list[str] | None) -> list[str]:
    requested = models or [TERM_EXTRACTION_MODEL]
    normalized: list[str] = []
    for model in requested:
        model_text = normalize_text(str(model or ""))
        if not model_text:
            continue
        if len(model_text) > 120:
            raise HTTPException(status_code=400, detail="模型 ID 过长。")
        if model_text not in normalized:
            normalized.append(model_text)
    if not normalized:
        raise HTTPException(status_code=400, detail="请至少选择一个提取模型。")
    if len(normalized) > 2:
        raise HTTPException(status_code=400, detail="最多只能选择两个模型进行比对。")
    return normalized


def _serialize_term_extraction_items(
    db: Session,
    term_base: TermBase | None,
    terms: list[ExtractedTerm],
) -> list[dict[str, Any]]:
    term_items = [
        {
            "source_text": item.source_text,
            "target_text": item.target_text,
            "source_normalized": item.source_normalized,
        }
        for item in terms
    ]
    if term_base is not None:
        return build_term_entry_conflict_items(
            db=db,
            term_base=term_base,
            entries=term_items,
        )
    return [
        {
            "index": index,
            "source_text": item["source_text"],
            "target_text": item["target_text"],
            "source_normalized": item["source_normalized"],
            "has_conflict": False,
            "conflict": None,
        }
        for index, item in enumerate(term_items)
    ]


@router.post("/file-records/{file_record_id}/term-extraction")
async def extract_file_record_terms(
    file_record_id: UUID,
    payload: TermExtractionRequest | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_read_access(file_record, current_user)
    source_language, target_language = _resolve_file_record_language_pair(file_record)
    body = payload or TermExtractionRequest()
    selected_models = _normalize_term_extraction_models(body.models)

    term_base = None
    target_term_base_id = body.term_base_id or file_record.term_base_id
    if target_term_base_id is not None:
        term_base = db.query(TermBase).filter(TermBase.id == target_term_base_id).first()
        if not term_base:
            raise HTTPException(status_code=404, detail="术语库不存在。")
        _ensure_resource_language_pair_matches(term_base, source_language, target_language, "术语库")

    segments = list_segments_for_file_record(db, file_record_id)
    if not segments:
        raise HTTPException(status_code=400, detail="当前文件没有可用于术语提取的已解析句段。")

    extractions = []
    extraction_errors: list[dict[str, str | int]] = []
    for model in selected_models:
        try:
            extractions.append(await extract_terms_from_segments(
                segments=segments,
                source_language=source_language,
                target_language=target_language,
                max_terms=body.max_terms,
                model=model,
                extraction_prompt=body.extraction_prompt,
            ))
        except LLMConfigurationError as exc:
            extraction_errors.append({"model": model, "message": str(exc), "status_code": 400})
        except LLMResponseValidationError as exc:
            extraction_errors.append({"model": model, "message": str(exc), "status_code": 502})
        except LLMRequestError as exc:
            extraction_errors.append({"model": model, "message": str(exc), "status_code": 502})
        except TermExtractionError as exc:
            extraction_errors.append({"model": model, "message": str(exc), "status_code": 400})

    if not extractions:
        status_code = 502 if any(error["status_code"] == 502 for error in extraction_errors) else 400
        detail = str(extraction_errors[0]["message"]) if extraction_errors else "术语提取失败。"
        raise HTTPException(status_code=status_code, detail=detail)

    results = []
    for extraction in extractions:
        terms = _serialize_term_extraction_items(db, term_base, extraction.terms)
        results.append({
            "provider": extraction.provider,
            "model": extraction.model,
            "terms": terms,
            "total": len(terms),
        })
    merged_terms = _serialize_term_extraction_items(
        db,
        term_base,
        merge_extracted_terms(extractions, max_terms=body.max_terms),
    )
    default_terms = merged_terms if len(results) > 1 else (results[0]["terms"] if results else [])
    primary_result = results[0] if results else {
        "provider": "openrouter",
        "model": selected_models[0],
        "terms": [],
        "total": 0,
    }

    return {
        "file_record": {
            "id": str(file_record.id),
            "filename": file_record.filename,
            "term_base_id": str(file_record.term_base_id) if file_record.term_base_id else None,
            "total_segments": len(segments),
        },
        "term_base_id": str(term_base.id) if term_base else None,
        "source_language": source_language,
        "target_language": target_language,
        "provider": primary_result["provider"],
        "model": primary_result["model"],
        "available_models": list(TERM_EXTRACTION_MODEL_OPTIONS),
        "results": results,
        "merged_terms": merged_terms,
        "terms": default_terms,
        "total": len(default_terms),
        "errors": [
            {"model": str(error["model"]), "message": str(error["message"])}
            for error in extraction_errors
        ],
    }


@router.post("/file-records/{file_record_id}/llm-translate")
@router.post("/documents/{file_record_id}/llm-translate", include_in_schema=False)
async def llm_translate_file_record(
    file_record_id: UUID,
    request: Request,
    payload: LLMTranslateRequest | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    operation_token: str | None = Header(default=None, alias=FILE_OPERATION_TOKEN_HEADER),
):
    """对指定范围的片段触发 LLM 译文修正，并通过 SSE 逐条返回结果。"""
    file_record = get_file_record_model(db, file_record_id)
    if not file_record:
        raise HTTPException(status_code=404, detail="文档不存在。")

    _require_file_record_work_access(file_record, current_user)
    ensure_file_record_write_allowed(db, file_record, operation_token=operation_token)
    source_language, target_language = _resolve_file_record_language_pair(file_record)
    body = payload or LLMTranslateRequest()
    requested_model = normalize_text(body.model or "") or None
    try:
        validate_provider_choice(body.provider, model_override=requested_model)
    except LLMConfigurationError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    guidelines = _resolve_llm_guidelines(db, file_record, body)

    translation_tasks = _build_llm_translation_tasks(
        db=db,
        file_record_id=file_record_id,
        scope=body.scope,
        source_language=source_language,
        target_language=target_language,
        collection_id=file_record.collection_id,
        include_context=body.translation_unit == "paragraph",
    )

    async def event_stream():
        updated_count = 0
        error_count = 0
        total_count = sum(1 for task in translation_tasks if task.should_translate)

        yield _sse_event(
            "start",
            {
                "file_record_id": str(file_record_id),
                "scope": body.scope,
                "provider": body.provider,
                "model": requested_model,
                "translation_unit": body.translation_unit,
                "source_language": source_language,
                "target_language": target_language,
                "total": total_count,
            },
        )

        if total_count == 0:
            yield _sse_event(
                "complete",
                {
                    "file_record_id": str(file_record_id),
                    "updated_count": 0,
                    "error_count": 0,
                    "total": 0,
                },
            )
            return

        # Pre-load all segments into memory to avoid per-item SELECT
        all_segments = (
            db.query(Segment)
            .filter(Segment.file_record_id == file_record_id)
            .all()
        )
        seg_map = {s.sentence_id: s for s in all_segments}

        COMMIT_INTERVAL = 50
        uncommitted_count = 0

        async for result in iter_batch_translate(
            translation_tasks,
            provider=body.provider,
            translation_guidelines=guidelines,
            translation_unit=body.translation_unit,
            model_override=requested_model,
        ):
            if await request.is_disconnected():
                break

            if isinstance(result, LLMTranslationFailure):
                error_count += 1
                yield _sse_event(
                    "error",
                    {
                        "sentence_id": result.sentence_id,
                        "status": result.status,
                        "message": result.error_message,
                    },
                )
                continue

            try:
                ensure_file_record_write_allowed(db, file_record, operation_token=operation_token)
            except Exception as exc:  # noqa: BLE001
                db.rollback()
                error_count += 1
                yield _sse_event(
                    "error",
                    {
                        "sentence_id": result.sentence_id,
                        "status": result.status,
                        "message": f"数据库更新失败：{exc}",
                    },
                )
                continue

            segment = seg_map.get(result.sentence_id)
            if not segment:
                error_count += 1
                yield _sse_event(
                    "error",
                    {
                        "sentence_id": result.sentence_id,
                        "status": result.status,
                        "message": "片段不存在，无法写回 LLM 译文。",
                    },
                )
                continue

            try:
                before_text = segment.target_text
                segment.target_text = result.translated_text
                segment.target_html = None
                segment.source = "llm"
                segment.version = int(segment.version or 1) + 1
                segment.source_word_count = segment.source_word_count or count_source_words(segment.source_text)
                segment.llm_provider = result.provider
                segment.llm_model = result.model

                # Inline revision creation — skip pending query for LLM bulk writes
                if (before_text or "") != (result.translated_text or ""):
                    db.add(SegmentRevision(
                        file_record_id=file_record_id,
                        segment_id=segment.id,
                        sentence_id=segment.sentence_id,
                        before_text=before_text or "",
                        after_text=result.translated_text or "",
                        source="llm",
                        status="pending",
                        author_id=current_user.id if current_user else None,
                    ))
                record_translation_metric_event(
                    db,
                    segment=segment,
                    before_text=before_text,
                    after_text=result.translated_text,
                    source="llm",
                    current_user=current_user,
                )

                uncommitted_count += 1
                if uncommitted_count >= COMMIT_INTERVAL:
                    db.commit()
                    uncommitted_count = 0

            except Exception as exc:  # noqa: BLE001
                db.rollback()
                uncommitted_count = 0
                error_count += 1
                yield _sse_event(
                    "error",
                    {
                        "sentence_id": result.sentence_id,
                        "status": result.status,
                        "message": f"数据库更新失败：{exc}",
                    },
                )
                continue

            updated_count += 1
            yield _sse_event(
                "segment",
                {
                    "sentence_id": segment.sentence_id,
                    "target_text": segment.target_text,
                    "status": segment.status,
                    "source": segment.source,
                    "provider": result.provider,
                    "model": result.model,
                },
            )

        # Final commit for remaining + sync status once
        try:
            if uncommitted_count > 0:
                db.commit()
            if updated_count > 0:
                sync_file_record_status(db, file_record_id)
                db.commit()
        except Exception:  # noqa: BLE001
            db.rollback()

        if not await request.is_disconnected():
            yield _sse_event(
                "complete",
                {
                    "file_record_id": str(file_record_id),
                    "updated_count": updated_count,
                    "error_count": error_count,
                    "total": total_count,
                },
            )

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.delete("/file-records/{file_record_id}")
@router.delete("/documents/{file_record_id}", include_in_schema=False)
def remove_file_record(
    file_record_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    """删除文档及其所有片段"""
    success = delete_file_record(db, file_record_id)
    if not success:
        raise HTTPException(status_code=404, detail="文档不存在。")
    return {"message": "文档已删除。"}


# ========== TM 管理 API ==========

class TMEntry(BaseModel):
    source_text: str
    target_text: str
    collection_id: UUID | None = None
    source_language: str | None = None
    target_language: str | None = None


class BatchTMEntry(BaseModel):
    collection_id: UUID | None = None
    source_language: str | None = None
    target_language: str | None = None
    entries: list[TMEntry]


class TMEntryUpdatePayload(BaseModel):
    source_text: str
    target_text: str


def _serialize_tm_entry(entry: TranslationMemory) -> dict:
    return {
        "id": entry.id,
        "collection_id": entry.collection_id,
        "source_text": entry.source_text,
        "target_text": entry.target_text,
        "source_language": entry.source_language,
        "target_language": entry.target_language,
        "created_at": entry.created_at.isoformat(),
        "updated_at": entry.updated_at.isoformat(),
    }


@router.get("/translation-memory/collections")
@router.get("/tm/collections", include_in_schema=False)
def list_tm_collections(
    db: Session = Depends(get_db),
):
    rows = (
        db.query(MemoryBase, func.count(MemoryEntry.id).label("entry_count"))
        .outerjoin(MemoryEntry, MemoryEntry.collection_id == MemoryBase.id)
        .group_by(MemoryBase.id)
        .order_by(MemoryBase.created_at.desc())
        .all()
    )
    return [
        _serialize_tm_collection(collection, int(entry_count))
        for collection, entry_count in rows
    ]


@router.get("/translation-memory/collections/{collection_id}")
@router.get("/tm/collections/{collection_id}", include_in_schema=False)
def get_tm_collection(
    collection_id: UUID,
    db: Session = Depends(get_db),
):
    collection = _get_collection_or_404(db, collection_id)
    if collection is None:
        raise HTTPException(status_code=404, detail="TM 记忆库不存在。")

    entry_count = (
        db.query(TranslationMemory)
        .filter(TranslationMemory.collection_id == collection.id)
        .count()
    )
    return _serialize_tm_collection(collection, entry_count)


@router.post("/translation-memory/collections")
@router.post("/tm/collections", include_in_schema=False)
def create_tm_collection(
    payload: MemoryBasePayload,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    name = _normalize_collection_name(payload.name)
    source_language, target_language = _require_tm_language_pair(
        payload.source_language,
        payload.target_language,
    )
    if not name:
        raise HTTPException(status_code=400, detail="记忆库名称不能为空。")

    collection = MemoryBase(
        name=name,
        description=normalize_text(payload.description or "") or None,
        source_language=source_language,
        target_language=target_language,
    )
    db.add(collection)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="同名记忆库已存在。") from exc

    db.refresh(collection)
    return _serialize_tm_collection(collection)


@router.put("/translation-memory/collections/{collection_id}")
@router.put("/tm/collections/{collection_id}", include_in_schema=False)
def update_tm_collection(
    collection_id: UUID,
    payload: MemoryBasePayload,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    collection = _get_collection_or_404(db, collection_id)
    if collection is None:
        raise HTTPException(status_code=404, detail="TM 记忆库不存在。")

    name = _normalize_collection_name(payload.name)
    source_language, target_language = _require_tm_language_pair(
        payload.source_language,
        payload.target_language,
    )
    if not name:
        raise HTTPException(status_code=400, detail="记忆库名称不能为空。")

    collection.name = name
    collection.description = normalize_text(payload.description or "") or None
    collection.source_language = source_language
    collection.target_language = target_language
    (
        db.query(TranslationMemory)
        .filter(TranslationMemory.collection_id == collection.id)
        .update(
            {
                TranslationMemory.source_language: source_language,
                TranslationMemory.target_language: target_language,
            },
            synchronize_session=False,
        )
    )
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="同名记忆库已存在。") from exc

    db.refresh(collection)
    entry_count = (
        db.query(MemoryEntry)
        .filter(MemoryEntry.collection_id == collection.id)
        .count()
    )
    return _serialize_tm_collection(collection, entry_count)


@router.post("/translation-memory/collections/merge")
@router.post("/tm/collections/merge", include_in_schema=False)
def merge_tm_collections(
    payload: TMCollectionMergePayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    source_collection_ids = list(dict.fromkeys(payload.source_collection_ids))
    if len(source_collection_ids) < 2:
        raise HTTPException(status_code=400, detail="请至少选择两个记忆库进行合并。")

    collections = (
        db.query(MemoryBase)
        .filter(MemoryBase.id.in_(source_collection_ids))
        .all()
    )
    collection_by_id = {collection.id: collection for collection in collections}
    missing_ids = [
        collection_id
        for collection_id in source_collection_ids
        if collection_id not in collection_by_id
    ]
    if missing_ids:
        raise HTTPException(status_code=404, detail="选择的记忆库不存在。")

    ordered_collections = [
        collection_by_id[collection_id]
        for collection_id in source_collection_ids
    ]
    source_language, target_language = _require_same_tm_collection_language_pair(
        ordered_collections,
    )
    name = _normalize_collection_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="合并后的记忆库名称不能为空。")

    target_collection = MemoryBase(
        name=name,
        description=normalize_text(payload.description or "") or None,
        source_language=source_language,
        target_language=target_language,
    )
    db.add(target_collection)
    try:
        db.flush()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="同名记忆库已存在。") from exc

    created_rows = 0
    updated_rows = 0
    skipped_rows = 0
    merged_entries: list[MemoryEntry] = []
    merged_by_hash: dict[str, MemoryEntry] = {}
    merged_by_source_text: dict[str, MemoryEntry] = {}

    for source_collection_id in source_collection_ids:
        source_entries = (
            db.query(MemoryEntry)
            .filter(MemoryEntry.collection_id == source_collection_id)
            .order_by(MemoryEntry.created_at.asc(), MemoryEntry.updated_at.asc())
            .all()
        )
        for entry in source_entries:
            source_text = normalize_text(entry.source_text)
            target_text = normalize_text(entry.target_text)
            if not source_text or not target_text:
                skipped_rows += 1
                continue

            source_hash = entry.source_hash or build_source_hash(source_text)
            source_normalized = entry.source_normalized or normalize_match_text(source_text) or source_text
            existing = merged_by_hash.get(source_hash) or merged_by_source_text.get(source_text)
            if existing is not None:
                existing.source_text = source_text
                existing.target_text = target_text
                existing.source_hash = source_hash
                existing.source_normalized = source_normalized
                existing.source_language = source_language
                existing.target_language = target_language
                existing.creator_id = entry.creator_id or current_user.id
                merged_by_hash[source_hash] = existing
                merged_by_source_text[source_text] = existing
                merged_entries.append(existing)
                updated_rows += 1
                continue

            merged_entry = MemoryEntry(
                collection_id=target_collection.id,
                source_text=source_text,
                target_text=target_text,
                source_hash=source_hash,
                source_normalized=source_normalized,
                source_language=source_language,
                target_language=target_language,
                creator_id=entry.creator_id or current_user.id,
            )
            db.add(merged_entry)
            merged_by_hash[source_hash] = merged_entry
            merged_by_source_text[source_text] = merged_entry
            merged_entries.append(merged_entry)
            created_rows += 1

    db.flush()
    sync_rows = list(
        {
            row.id: row.source_text
            for row in merged_entries
            if row.id is not None and row.source_text
        }.items()
    )
    db.commit()
    sync_tm_embeddings(db, sync_rows)

    entry_count = (
        db.query(MemoryEntry)
        .filter(MemoryEntry.collection_id == target_collection.id)
        .count()
    )
    db.refresh(target_collection)
    return {
        "collection": _serialize_tm_collection(target_collection, entry_count),
        "source_count": len(source_collection_ids),
        "created_rows": created_rows,
        "updated_rows": updated_rows,
        "skipped_rows": skipped_rows,
        "merged_rows": created_rows + updated_rows,
    }


@router.delete("/translation-memory/collections/{collection_id}")
@router.delete("/tm/collections/{collection_id}", include_in_schema=False)
def delete_tm_collection(
    collection_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    collection = _get_collection_or_404(db, collection_id)
    if collection is None:
        raise HTTPException(status_code=404, detail="TM 记忆库不存在。")

    entry_count = (
        db.query(MemoryEntry)
        .filter(MemoryEntry.collection_id == collection.id)
        .count()
    )
    (
        db.query(FileRecord)
        .filter(FileRecord.collection_id == collection.id)
        .update({FileRecord.collection_id: None}, synchronize_session=False)
    )
    (
        db.query(MemoryEntry)
        .filter(MemoryEntry.collection_id == collection.id)
        .delete(synchronize_session=False)
    )

    db.delete(collection)
    db.commit()
    return {"message": "记忆库已删除。", "deleted_entries": entry_count}


@router.post("/translation-memory/preview-sdltm")
@router.post("/tm/preview-sdltm", include_in_schema=False)
async def preview_sdltm(
    file: UploadFile = File(...),
):
    """Preview SDLTM file metadata without importing."""
    extension = f".{(file.filename or '').split('.')[-1].lower()}" if file.filename else ""
    if extension not in SDLTM_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持 .sdltm 文件。")

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="上传的文件为空。")

    try:
        metadata = preview_sdltm_metadata(raw_bytes)
        return {
            "name": metadata.name,
            "source_language": metadata.source_language,
            "target_language": metadata.target_language,
            "entry_count": metadata.entry_count,
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"读取 SDLTM 元数据失败：{exc}") from exc


@router.post("/translation-memory/import-xlsx")
@router.post("/tm/import-xlsx", include_in_schema=False)
@router.post("/translation-memory/import", include_in_schema=False)
@router.post("/tm/import", include_in_schema=False)
async def import_tm_xlsx(
    file: UploadFile = File(...),
    collection_id: UUID | None = Form(default=None),
    source_language: str = Form(...),
    target_language: str = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    extension = f".{(file.filename or '').split('.')[-1].lower()}" if file.filename else ""
    if extension not in TM_IMPORT_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持上传 .xlsx 或 .sdltm 文件。")

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="上传的文件为空。")

    collection = _get_collection_or_404(db, collection_id)
    resolved_source_language, resolved_target_language = _resolve_collection_language_pair(
        collection,
        source_language,
        target_language,
    )
    try:
        if extension in SDLTM_EXTENSIONS:
            import_summary = import_tm_from_sdltm_upload(
                db=db,
                raw_bytes=raw_bytes,
                filename=file.filename or "uploaded.sdltm",
                collection_id=collection_id,
                source_language=resolved_source_language,
                target_language=resolved_target_language,
                creator_id=current_user.id,
            )
        else:
            import_summary = import_tm_from_xlsx_upload(
                db=db,
                raw_bytes=raw_bytes,
                filename=file.filename or "uploaded.xlsx",
                collection_id=collection_id,
                source_language=resolved_source_language,
                target_language=resolved_target_language,
                creator_id=current_user.id,
            )
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"TM 导入失败：{exc}") from exc

    return {
        "filename": import_summary.filename,
        "created_rows": import_summary.created_rows,
        "updated_rows": import_summary.updated_rows,
        "skipped_empty_rows": import_summary.skipped_empty_rows,
        "skipped_header_rows": import_summary.skipped_header_rows,
        "imported_rows": import_summary.imported_rows,
        "collection_id": collection.id if collection else None,
        "collection_name": collection.name if collection else None,
        "source_language": resolved_source_language,
        "target_language": resolved_target_language,
    }


@router.get("/translation-memory/collections/{collection_id}/entries")
@router.get("/tm/collections/{collection_id}/entries", include_in_schema=False)
def list_tm_collection_entries(
    collection_id: UUID,
    skip: int = 0,
    limit: int = 50,
    search: str | None = None,
    case_sensitive: bool = False,
    db: Session = Depends(get_db),
):
    collection = _get_collection_or_404(db, collection_id)
    if collection is None:
        raise HTTPException(status_code=404, detail="TM 记忆库不存在。")

    safe_skip = max(skip, 0)
    safe_limit = min(max(limit, 1), 200)
    query = (
        db.query(TranslationMemory)
        .filter(TranslationMemory.collection_id == collection.id)
    )
    normalized_search = normalize_text(search or "")
    if normalized_search:
        like_pattern = f"%{normalized_search}%"
        if case_sensitive:
            query = query.filter(
                or_(
                    TranslationMemory.source_text.like(like_pattern),
                    TranslationMemory.target_text.like(like_pattern),
                )
            )
        else:
            query = query.filter(
                or_(
                    TranslationMemory.source_text.ilike(like_pattern),
                    TranslationMemory.target_text.ilike(like_pattern),
                )
            )

    total = query.count()
    rows = (
        query
        .order_by(TranslationMemory.updated_at.desc(), TranslationMemory.created_at.desc())
        .offset(safe_skip)
        .limit(safe_limit)
        .all()
    )
    return {
        "items": [_serialize_tm_entry(row) for row in rows],
        "total": total,
        "skip": safe_skip,
        "limit": safe_limit,
    }


@router.get("/translation-memory/collections/{collection_id}/export-xlsx")
def export_tm_collection_entries_xlsx(
    collection_id: UUID,
    db: Session = Depends(get_db),
):
    collection = _get_collection_or_404(db, collection_id)
    if collection is None:
        raise HTTPException(status_code=404, detail="TM 记忆库不存在。")

    entries = (
        db.query(TranslationMemory)
        .filter(TranslationMemory.collection_id == collection.id)
        .order_by(TranslationMemory.updated_at.desc(), TranslationMemory.created_at.desc())
        .all()
    )
    rows = [
        [
            entry.source_text,
            entry.target_text,
            entry.source_language or "",
            entry.target_language or "",
            entry.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            entry.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
        ]
        for entry in entries
    ]
    xlsx_bytes = build_tabular_xlsx(
        sheet_title=collection.name,
        headers=["原文", "译文", "源语言", "目标语言", "创建时间", "更新时间"],
        rows=rows,
    )
    return build_xlsx_download_response(f"{collection.name}-tm.xlsx", xlsx_bytes)


@router.post("/translation-memory/collections/{collection_id}/exports")
def queue_tm_collection_export(
    collection_id: UUID,
    format: ResourceExportFormat = Query(default="xlsx"),
    db: Session = Depends(get_db),
):
    collection = _get_collection_or_404(db, collection_id)
    if collection is None:
        raise HTTPException(status_code=404, detail="TM 记忆库不存在。")

    return JSONResponse(
        status_code=202,
        content=queue_resource_export(
            resource_type="tm",
            resource_id=collection.id,
            export_format=format,
        ),
    )


@router.get("/translation-memory/export-tasks/{task_id}")
def get_tm_export_task(task_id: str):
    return ensure_export_task_status(task_id, expected_resource_type="tm")


@router.get("/translation-memory/export-tasks/{task_id}/download")
def download_tm_export_task(task_id: str):
    return build_resource_export_download_response(task_id, expected_resource_type="tm")


@router.post("/translation-memory/collections/{collection_id}/entries")
def add_tm_collection_entry(
    collection_id: UUID,
    payload: TMEntryUpdatePayload,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    collection = _get_collection_or_404(db, collection_id)
    if collection is None:
        raise HTTPException(status_code=404, detail="TM 记忆库不存在。")

    result = add_tm_entry(
        TMEntry(
            collection_id=collection.id,
            source_text=payload.source_text,
            target_text=payload.target_text,
            source_language=collection.source_language,
            target_language=collection.target_language,
        ),
        db,
    )
    entry = db.query(TranslationMemory).filter(TranslationMemory.id == result["id"]).first()
    if entry is None:
        raise HTTPException(status_code=500, detail="TM 条目保存成功，但读取结果失败。")
    return _serialize_tm_entry(entry)


@router.post("/translation-memory/entries")
@router.post("/tm/add", include_in_schema=False)
def add_tm_entry(
    entry: TMEntry,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    """添加单条 TM 记录（去重：相同原文不重复添加）"""
    source_text = normalize_text(entry.source_text)
    target_text = normalize_text(entry.target_text)

    if not source_text or not target_text:
        raise HTTPException(status_code=400, detail="原文和译文不能为空。")

    collection = _get_collection_or_404(db, entry.collection_id)
    source_language, target_language = _resolve_collection_language_pair(
        collection,
        entry.source_language,
        entry.target_language,
    )
    source_hash = build_source_hash(source_text)

    # 检查是否已存在
    existing_query = db.query(MemoryEntry).filter(
        MemoryEntry.source_hash == source_hash
    )
    existing = _filter_tm_collection(
        existing_query,
        entry.collection_id,
        source_language=source_language,
        target_language=target_language,
    ).first()

    if existing:
        # 已存在，更新译文
        existing.source_text = source_text
        existing.target_text = target_text
        existing.source_hash = source_hash
        existing.source_normalized = normalize_match_text(source_text) or source_text
        existing.source_language = source_language
        existing.target_language = target_language
        db.commit()
        sync_tm_embeddings(db, [(existing.id, existing.source_text)])
        return {"status": "updated", "id": existing.id, "message": "已更新现有记录。"}

    # 不存在，新增
    tm = MemoryEntry(
        collection_id=entry.collection_id,
        source_text=source_text,
        target_text=target_text,
        source_hash=source_hash,
        source_normalized=normalize_match_text(source_text) or source_text,
        source_language=source_language,
        target_language=target_language,
    )
    db.add(tm)
    db.commit()
    db.refresh(tm)
    sync_tm_embeddings(db, [(tm.id, tm.source_text)])

    return {"status": "created", "id": tm.id, "message": "已添加新记录。"}


@router.put("/translation-memory/entries/{entry_id}")
@router.put("/tm/entries/{entry_id}", include_in_schema=False)
def update_tm_entry(
    entry_id: UUID,
    payload: TMEntryUpdatePayload,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    entry = db.query(TranslationMemory).filter(TranslationMemory.id == entry_id).first()
    if entry is None:
        raise HTTPException(status_code=404, detail="TM 条目不存在。")

    source_text = normalize_text(payload.source_text)
    target_text = normalize_text(payload.target_text)
    if not source_text or not target_text:
        raise HTTPException(status_code=400, detail="原文和译文不能为空。")

    collection = _get_collection_or_404(db, entry.collection_id)
    source_language = collection.source_language if collection else entry.source_language
    target_language = collection.target_language if collection else entry.target_language
    if not source_language or not target_language:
        raise HTTPException(status_code=400, detail="当前 TM 条目缺少语言对，请先更新记忆库信息。")

    source_hash = build_source_hash(source_text)
    duplicate_query = db.query(TranslationMemory).filter(
        TranslationMemory.id != entry.id,
        or_(
            TranslationMemory.source_hash == source_hash,
            TranslationMemory.source_text == source_text,
        ),
    )
    duplicate = _filter_tm_collection(
        duplicate_query,
        entry.collection_id,
        source_language=source_language,
        target_language=target_language,
    ).first()
    if duplicate is not None:
        raise HTTPException(status_code=409, detail="当前记忆库中已存在相同原文的 TM 条目。")

    entry.source_text = source_text
    entry.target_text = target_text
    entry.source_hash = source_hash
    entry.source_normalized = normalize_match_text(source_text) or source_text
    entry.source_language = source_language
    entry.target_language = target_language
    db.commit()
    db.refresh(entry)
    sync_tm_embeddings(db, [(entry.id, entry.source_text)])
    return _serialize_tm_entry(entry)


@router.delete("/translation-memory/entries/{entry_id}")
@router.delete("/tm/entries/{entry_id}", include_in_schema=False)
def delete_tm_entry(
    entry_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    entry = db.query(TranslationMemory).filter(TranslationMemory.id == entry_id).first()
    if entry is None:
        raise HTTPException(status_code=404, detail="TM 条目不存在。")

    db.delete(entry)
    db.commit()
    return {"message": "TM 条目已删除。"}


def _upsert_tm_entry(
    db: Session,
    source_text: str,
    target_text: str,
    collection_id: UUID,
    source_language: str,
    target_language: str,
) -> tuple[Literal["created", "updated"], MemoryEntry]:
    source_hash = build_source_hash(source_text)
    existing_query = db.query(MemoryEntry).filter(
        MemoryEntry.source_hash == source_hash
    )
    existing = _filter_tm_collection(
        existing_query,
        collection_id,
        source_language=source_language,
        target_language=target_language,
    ).first()

    if existing:
        existing.source_text = source_text
        existing.target_text = target_text
        existing.source_hash = source_hash
        existing.source_normalized = normalize_match_text(source_text) or source_text
        existing.collection_id = collection_id
        existing.source_language = source_language
        existing.target_language = target_language
        return "updated", existing

    tm = MemoryEntry(
        collection_id=collection_id,
        source_text=source_text,
        target_text=target_text,
        source_hash=source_hash,
        source_normalized=normalize_match_text(source_text) or source_text,
        source_language=source_language,
        target_language=target_language,
    )
    db.add(tm)
    return "created", tm


@router.post("/translation-memory/entries/batch")
@router.post("/tm/batch-add", include_in_schema=False)
def batch_add_tm_entries(
    batch: BatchTMEntry,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """批量添加 TM 记录（去重）"""
    skipped_count = 0
    upsert_entries: list[TMUpsertEntry] = []
    collection_ids = [
        collection_id
        for collection_id in (
            [batch.collection_id]
            + [entry.collection_id for entry in batch.entries]
        )
        if collection_id is not None
    ]
    _validate_collection_ids(db, collection_ids)

    for entry in batch.entries:
        source_text = normalize_text(entry.source_text)
        target_text = normalize_text(entry.target_text)
        collection_id = entry.collection_id or batch.collection_id

        if not source_text or not target_text:
            skipped_count += 1
            continue

        collection = _get_collection_or_404(db, collection_id)
        source_language, target_language = _resolve_collection_language_pair(
            collection,
            entry.source_language or batch.source_language,
            entry.target_language or batch.target_language,
        )
        upsert_entries.append(
            TMUpsertEntry(
                collection_id=collection_id,
                source_text=source_text,
                target_text=target_text,
                source_language=source_language,
                target_language=target_language,
                creator_id=current_user.id,
            )
        )

    upsert_summary = batch_upsert_tm_entries(db, upsert_entries)
    skipped_count += upsert_summary.skipped_count
    if upsert_summary.total_written > 0:
        db.commit()
        sync_tm_embeddings(db, upsert_summary.sync_rows or [])

    return {
        "created": upsert_summary.created_count,
        "updated": upsert_summary.updated_count,
        "skipped": skipped_count,
    }


# ========== 术语库管理 API ==========

def _serialize_termbase_collection(collection: TermBase, entry_count: int = 0) -> dict:
    return {
        "id": collection.id,
        "name": collection.name,
        "description": collection.description,
        "source_language": collection.source_language,
        "target_language": collection.target_language,
        "created_at": collection.created_at.isoformat(),
        "updated_at": collection.updated_at.isoformat(),
        "entry_count": entry_count,
    }


def _get_termbase_collection_or_404(db: Session, collection_id: UUID | None) -> TermBase | None:
    if collection_id is None:
        return None

    collection = db.query(TermBase).filter(TermBase.id == collection_id).first()
    if collection is None:
        raise HTTPException(status_code=404, detail="术语库不存在。")
    return collection


@router.get("/termbase/collections")
def list_termbase_collections(
    db: Session = Depends(get_db),
):
    rows = (
        db.query(TermBase, func.count(TermEntry.id).label("entry_count"))
        .outerjoin(TermEntry, TermEntry.term_base_id == TermBase.id)
        .group_by(TermBase.id)
        .order_by(TermBase.created_at.desc())
        .all()
    )
    return [
        _serialize_termbase_collection(collection, int(entry_count))
        for collection, entry_count in rows
    ]


@router.post("/termbase/collections")
def create_termbase_collection(
    payload: TermBasePayload,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    name = _normalize_collection_name(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="术语库名称不能为空。")

    collection = TermBase(
        name=name,
        description=normalize_text(payload.description or "") or None,
        source_language=payload.source_language,
        target_language=payload.target_language,
    )
    db.add(collection)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="同名术语库已存在。") from exc

    db.refresh(collection)
    return _serialize_termbase_collection(collection)


@router.delete("/termbase/collections/{collection_id}")
def delete_termbase_collection(
    collection_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    collection = _get_termbase_collection_or_404(db, collection_id)
    if collection is None:
        raise HTTPException(status_code=404, detail="术语库不存在。")

    entry_count = (
        db.query(TermEntry)
        .filter(TermEntry.term_base_id == collection.id)
        .count()
    )
    if entry_count:
        raise HTTPException(status_code=409, detail="请先清空该术语库中的术语记录。")

    db.delete(collection)
    db.commit()
    return {"message": "术语库已删除。"}


@router.get("/termbase/terms")
def list_terms(
    collection_id: UUID | None = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
):
    query = db.query(TermEntry)
    if collection_id:
        query = query.filter(TermEntry.term_base_id == collection_id)

    total = query.count()
    terms = query.order_by(TermEntry.source_text).offset(skip).limit(limit).all()

    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "terms": [
            {
                "id": term.id,
                "source_text": term.source_text,
                "target_text": term.target_text,
                "term_base_id": term.term_base_id,
                "created_at": term.created_at.isoformat(),
            }
            for term in terms
        ],
    }


@router.post("/termbase/terms")
def add_term(
    payload: TermPayload,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    source_text = normalize_text(payload.source_text)
    target_text = normalize_text(payload.target_text)

    if not source_text or not target_text:
        raise HTTPException(status_code=400, detail="原文和译文不能为空。")

    _get_termbase_collection_or_404(db, payload.collection_id)

    # 检查是否已存在相同原文的术语
    existing = (
        db.query(TermEntry)
        .filter(TermEntry.source_text == source_text, TermEntry.term_base_id == payload.collection_id)
        .first()
    )

    if existing:
        existing.target_text = target_text
        db.commit()
        return {"status": "updated", "id": existing.id, "message": "已更新现有术语。"}

    # 获取术语库的语言设置
    term_base = _get_termbase_collection_or_404(db, payload.collection_id)
    source_lang = term_base.source_language if term_base else "zh"
    target_lang = term_base.target_language if term_base else "en"

    term = TermEntry(
        term_base_id=payload.collection_id,
        source_text=source_text,
        target_text=target_text,
        source_language=source_lang,
        target_language=target_lang,
    )
    db.add(term)
    db.commit()
    db.refresh(term)

    return {"status": "created", "id": term.id, "message": "已添加新术语。"}


@router.delete("/termbase/terms/{term_id}")
def delete_term(
    term_id: UUID,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    term = db.query(TermEntry).filter(TermEntry.id == term_id).first()
    if not term:
        raise HTTPException(status_code=404, detail="术语不存在。")

    db.delete(term)
    db.commit()
    return {"message": "术语已删除。"}


@router.post("/termbase/import-xlsx")
async def import_termbase_xlsx(
    file: UploadFile = File(...),
    collection_id: UUID | None = Form(default=None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    extension = f".{(file.filename or '').split('.')[-1].lower()}" if file.filename else ""
    if extension not in XLSX_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持上传 .xlsx 文件。")

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="上传的 XLSX 文件为空。")

    collection = _get_termbase_collection_or_404(db, collection_id)

    try:
        from openpyxl import load_workbook
        from io import BytesIO

        wb = load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb.active

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not row or len(row) < 2:
                skipped_count += 1
                continue

            source_text = normalize_text(str(row[0] or ""))
            target_text = normalize_text(str(row[1] or ""))

            if not source_text or not target_text:
                skipped_count += 1
                continue

            # 跳过表头
            if row_idx == 1 and (source_text.lower() in ("source", "原文", "术语") or target_text.lower() in ("target", "译文", "翻译")):
                skipped_count += 1
                continue

            existing = (
                db.query(TermEntry)
                .filter(TermEntry.source_text == source_text, TermEntry.term_base_id == collection_id)
                .first()
            )

            if existing:
                existing.target_text = target_text
                updated_count += 1
            else:
                # 获取术语库的语言设置
                source_lang = collection.source_language if collection else "zh"
                target_lang = collection.target_language if collection else "en"
                term = TermEntry(
                    term_base_id=collection_id,
                    source_text=source_text,
                    target_text=target_text,
                    source_language=source_lang,
                    target_language=target_lang,
                )
                db.add(term)
                created_count += 1

        db.commit()
        wb.close()

    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"术语库导入失败：{exc}") from exc

    return {
        "filename": file.filename,
        "created_rows": created_count,
        "updated_rows": updated_count,
        "skipped_rows": skipped_count,
        "imported_rows": created_count + updated_count,
        "collection_id": collection.id if collection else None,
        "collection_name": collection.name if collection else None,
    }


@router.get("/termbase/match")
def match_terms(
    text: str,
    collection_ids: list[UUID] | None = None,
    db: Session = Depends(get_db),
):
    """匹配文本中的术语，返回匹配到的术语列表（长术语优先）"""
    if not text:
        return {"matches": []}

    query = db.query(TermEntry)
    if collection_ids:
        query = query.filter(TermEntry.term_base_id.in_(collection_ids))

    all_terms = query.all()

    # 按原文长度降序排序（长术语优先）
    sorted_terms = sorted(all_terms, key=lambda t: len(t.source_text), reverse=True)

    matches = []
    matched_positions = set()
    text_lower = text.lower()

    for term in sorted_terms:
        term_lower = term.source_text.lower()
        start = 0
        while True:
            pos = text_lower.find(term_lower, start)
            if pos == -1:
                break

            end_pos = pos + len(term.source_text)
            # 检查是否与已匹配的位置重叠
            overlap = False
            for matched_start, matched_end in matched_positions:
                if not (end_pos <= matched_start or pos >= matched_end):
                    overlap = True
                    break

            if not overlap:
                matched_positions.add((pos, end_pos))
                matches.append({
                    "term_id": str(term.id),
                    "source_text": term.source_text,
                    "target_text": term.target_text,
                    "start": pos,
                    "end": end_pos,
                })

            start = pos + 1

    # 按位置排序
    matches.sort(key=lambda m: m["start"])

    return {"matches": matches}
