"""Excel 文件解析"""

import os
from .base import BaseParser, Document, Paragraph, Table


class XlsxParser(BaseParser):

    @staticmethod
    def supported_extensions():
        return [".xlsx", ".xls"]

    def parse(self, file_path: str) -> Document:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True)
        tables = []
        paragraphs = []

        for sheet in wb.worksheets:
            rows_data = []
            for row in sheet.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else "" for cell in row]
                # 跳过完全空行
                if any(cell.strip() for cell in row_text):
                    rows_data.append(row_text)

            if rows_data:
                headers = rows_data[0]
                data_rows = rows_data[1:]
                tables.append(Table(headers=headers, rows=data_rows))

                # 也把内容存为段落（方便全文分析）
                for row in rows_data:
                    text = " | ".join(cell for cell in row if cell.strip())
                    if text:
                        paragraphs.append(Paragraph(text=text))

        wb.close()
        raw_text = "\n".join(p.text for p in paragraphs)

        return Document(
            paragraphs=paragraphs,
            tables=tables,
            filename=os.path.basename(file_path),
            raw_text=raw_text,
        )
