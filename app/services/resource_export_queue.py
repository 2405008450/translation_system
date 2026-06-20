from __future__ import annotations

import logging
import re
import time
from concurrent.futures import CancelledError, Future, ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Literal
from urllib.parse import quote
from uuid import UUID, uuid4
from xml.sax.saxutils import escape, quoteattr

from fastapi import HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, aliased

from app.config import get_settings
from app.services.cache import get_json as cache_get_json
from app.services.cache import set_json as cache_set_json
from app.services.xlsx_exporter import XLSX_MEDIA_TYPE


logger = logging.getLogger(__name__)

ResourceExportKind = Literal["tm", "term", "glossary"]
ResourceExportFormat = Literal["xlsx", "tmx"]
ResourceExportStatus = Literal["queued", "running", "completed", "failed", "canceling", "canceled"]
ProgressCallback = Callable[[int, str], None]
CancelCheck = Callable[[], bool]

EXPORT_TASK_TTL_SECONDS = 24 * 60 * 60
EXPORT_QUERY_BATCH_SIZE = 1000
EXPORT_MEDIA_TYPES: dict[ResourceExportFormat, str] = {
    "xlsx": XLSX_MEDIA_TYPE,
    "tmx": "application/x-tmx+xml",
}

# 本地模式下串行处理大文件导出，避免多个重任务同时占用数据库连接和文件句柄。
_EXPORT_EXECUTOR = ThreadPoolExecutor(max_workers=1, thread_name_prefix="resource-export")
_EXPORT_FUTURES: dict[str, Future] = {}


class ResourceExportCanceled(Exception):
    """导出任务已被用户取消。"""


def queue_resource_export(
    *,
    resource_type: ResourceExportKind,
    resource_id: UUID,
    export_format: ResourceExportFormat,
) -> dict[str, Any]:
    task_id = str(uuid4())
    payload = {
        "task_id": task_id,
        "resource_type": resource_type,
        "resource_id": str(resource_id),
        "format": export_format,
    }
    _set_export_task_status(
        task_id,
        "queued",
        progress=0,
        message="导出任务已进入队列。",
        payload=payload,
    )
    future = _EXPORT_EXECUTOR.submit(_run_resource_export_task, task_id, payload)
    _EXPORT_FUTURES[task_id] = future
    future.add_done_callback(_log_export_task_failure)
    status = get_resource_export_task_status(task_id)
    return status or payload


def get_resource_export_task_status(task_id: str) -> dict[str, Any] | None:
    payload = cache_get_json(_export_task_cache_key(task_id))
    return payload if isinstance(payload, dict) else None


def ensure_export_task_status(
    task_id: str,
    *,
    expected_resource_type: ResourceExportKind,
) -> dict[str, Any]:
    status = get_resource_export_task_status(task_id)
    if not status or status.get("resource_type") != expected_resource_type:
        raise HTTPException(status_code=404, detail="导出任务不存在。")
    return status


def _export_task_cancel_requested(task_id: str) -> bool:
    status = get_resource_export_task_status(task_id)
    return bool(status and status.get("status") in {"canceling", "canceled"})


def _raise_if_export_canceled(task_id: str) -> None:
    if _export_task_cancel_requested(task_id):
        raise ResourceExportCanceled("导出已取消。")


def cancel_resource_export_task(
    task_id: str,
    *,
    expected_resource_type: ResourceExportKind,
) -> dict[str, Any]:
    status = ensure_export_task_status(task_id, expected_resource_type=expected_resource_type)
    current_status = str(status.get("status") or "")
    if current_status in {"completed", "failed", "canceled"}:
        return status

    payload = {
        "resource_type": status.get("resource_type"),
        "resource_id": status.get("resource_id"),
        "format": status.get("format"),
    }
    future = _EXPORT_FUTURES.get(task_id)
    if current_status == "queued" and future is not None and future.cancel():
        return _set_export_task_status(
            task_id,
            "canceled",
            progress=100,
            message="导出已取消。",
            payload=payload,
        )

    return _set_export_task_status(
        task_id,
        "canceling",
        progress=int(status.get("progress") or 0),
        message="正在取消导出，请稍候。",
        payload=payload,
    )


def build_resource_export_download_response(
    task_id: str,
    *,
    expected_resource_type: ResourceExportKind,
) -> FileResponse:
    status = ensure_export_task_status(task_id, expected_resource_type=expected_resource_type)
    if status.get("status") != "completed":
        raise HTTPException(status_code=409, detail="导出任务尚未完成。")

    result = status.get("result")
    if not isinstance(result, dict):
        raise HTTPException(status_code=404, detail="导出文件不存在。")

    file_path = Path(str(result.get("file_path") or ""))
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="导出文件不存在或已过期。")

    filename = str(result.get("filename") or file_path.name)
    media_type = str(result.get("media_type") or "application/octet-stream")
    ascii_filename = filename.encode("ascii", "ignore").decode("ascii").strip() or file_path.name
    ascii_filename = ascii_filename.replace('"', "")
    quoted_filename = quote(filename)

    return FileResponse(
        file_path,
        media_type=media_type,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{ascii_filename}"; '
                f"filename*=UTF-8''{quoted_filename}"
            )
        },
    )


def export_tm_collection_now(
    *,
    db: Session,
    collection_id: UUID,
    export_format: ResourceExportFormat,
    output_path: Path,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    from app.models import MemoryBase, TranslationMemory

    collection = db.query(MemoryBase).filter(MemoryBase.id == collection_id).first()
    if collection is None:
        raise ValueError("TM 记忆库不存在。")

    total_entries = (
        db.query(TranslationMemory)
        .filter(TranslationMemory.collection_id == collection.id)
        .count()
    )
    progress = progress_callback or _noop_progress_callback
    progress(10, "正在读取记忆库条目。")
    if export_format == "xlsx":
        _write_xlsx_file(
            output_path=output_path,
            sheet_title=collection.name,
            headers=["原文", "译文", "源语言", "目标语言", "创建人", "创建时间", "最后修改人", "更新时间"],
            row_iter=_iter_tm_xlsx_rows(db, collection.id),
            total_entries=total_entries,
            progress_callback=progress,
        )
    elif export_format == "tmx":
        _write_tmx_file(
            output_path=output_path,
            source_language=collection.source_language or "und",
            target_language=collection.target_language or "und",
            rows=_iter_tm_tmx_rows(db, collection.id),
            total_entries=total_entries,
            progress_callback=progress,
            filename=collection.name,
            tuid_prefix="tm",
        )
    else:
        raise ValueError("不支持的导出格式。")

    return _build_export_result(
        resource_name=collection.name,
        filename_suffix="tm",
        export_format=export_format,
        output_path=output_path,
        total_entries=total_entries,
    )


def export_term_base_now(
    *,
    db: Session,
    term_base_id: UUID,
    export_format: ResourceExportFormat,
    output_path: Path,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    from app.models import TermBase, TermEntry

    term_base = db.query(TermBase).filter(TermBase.id == term_base_id).first()
    if term_base is None:
        raise ValueError("术语库不存在。")

    total_entries = (
        db.query(TermEntry)
        .filter(TermEntry.term_base_id == term_base.id)
        .count()
    )
    progress = progress_callback or _noop_progress_callback
    progress(10, "正在读取术语库条目。")
    if export_format == "xlsx":
        _write_xlsx_file(
            output_path=output_path,
            sheet_title=term_base.name,
            headers=["术语原文", "术语译文", "源语言", "目标语言", "创建人", "创建时间", "最后修改人", "更新时间"],
            row_iter=_iter_term_xlsx_rows(db, term_base.id),
            total_entries=total_entries,
            progress_callback=progress,
        )
    elif export_format == "tmx":
        _write_tmx_file(
            output_path=output_path,
            source_language=term_base.source_language or "und",
            target_language=term_base.target_language or "und",
            rows=_iter_term_tmx_rows(db, term_base.id),
            total_entries=total_entries,
            progress_callback=progress,
            filename=term_base.name,
            tuid_prefix="term",
        )
    else:
        raise ValueError("不支持的导出格式。")

    return _build_export_result(
        resource_name=term_base.name,
        filename_suffix="term-base",
        export_format=export_format,
        output_path=output_path,
        total_entries=total_entries,
    )


def export_glossary_base_now(
    *,
    db: Session,
    glossary_base_id: UUID,
    export_format: ResourceExportFormat,
    output_path: Path,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    from app.models import GlossaryBase, GlossaryEntry

    glossary_base = db.query(GlossaryBase).filter(GlossaryBase.id == glossary_base_id).first()
    if glossary_base is None:
        raise ValueError("词汇表不存在。")

    total_entries = (
        db.query(GlossaryEntry)
        .filter(GlossaryEntry.glossary_base_id == glossary_base.id)
        .count()
    )
    progress = progress_callback or _noop_progress_callback
    progress(10, "正在读取词汇表条目。")
    if export_format == "xlsx":
        _write_xlsx_file(
            output_path=output_path,
            sheet_title=glossary_base.name,
            headers=["原文", "译文", "备注", "源语言", "目标语言", "创建人", "创建时间", "最后修改人", "更新时间"],
            row_iter=_iter_glossary_xlsx_rows(db, glossary_base.id),
            total_entries=total_entries,
            progress_callback=progress,
        )
    elif export_format == "tmx":
        _write_tmx_file(
            output_path=output_path,
            source_language=glossary_base.source_language or "und",
            target_language=glossary_base.target_language or "und",
            rows=_iter_glossary_tmx_rows(db, glossary_base.id),
            total_entries=total_entries,
            progress_callback=progress,
            filename=glossary_base.name,
            tuid_prefix="glossary",
        )
    else:
        raise ValueError("不支持的导出格式。")

    return _build_export_result(
        resource_name=glossary_base.name,
        filename_suffix="glossary",
        export_format=export_format,
        output_path=output_path,
        total_entries=total_entries,
    )


def _run_resource_export_task(task_id: str, payload: dict[str, Any]) -> None:
    resource_type = payload.get("resource_type")
    export_format = payload.get("format")
    output_path: Path | None = None
    try:
        resource_id = UUID(str(payload["resource_id"]))
        if resource_type not in ("tm", "term", "glossary") or export_format not in ("xlsx", "tmx"):
            raise ValueError("导出任务参数不正确。")
        _raise_if_export_canceled(task_id)

        output_dir = _ensure_export_dir()
        _cleanup_expired_export_files(output_dir)
        output_path = output_dir / f"{task_id}.{export_format}"

        def update_progress(progress: int, message: str) -> None:
            _raise_if_export_canceled(task_id)
            _set_export_task_status(
                task_id,
                "running",
                progress=progress,
                message=message,
                payload=payload,
            )

        update_progress(5, "导出任务开始处理。")
        from app.database import SessionLocal

        with SessionLocal() as db:
            if resource_type == "tm":
                result = export_tm_collection_now(
                    db=db,
                    collection_id=resource_id,
                    export_format=export_format,
                    output_path=output_path,
                    progress_callback=update_progress,
                )
            elif resource_type == "term":
                result = export_term_base_now(
                    db=db,
                    term_base_id=resource_id,
                    export_format=export_format,
                    output_path=output_path,
                    progress_callback=update_progress,
                )
            else:
                result = export_glossary_base_now(
                    db=db,
                    glossary_base_id=resource_id,
                    export_format=export_format,
                    output_path=output_path,
                    progress_callback=update_progress,
                )

        _raise_if_export_canceled(task_id)
        _set_export_task_status(
            task_id,
            "completed",
            progress=100,
            message="导出完成。",
            payload=payload,
            result=result,
        )
    except ResourceExportCanceled:
        if output_path is not None:
            try:
                output_path.unlink(missing_ok=True)
            except OSError:
                logger.debug("skip cleanup for canceled export path=%s", output_path, exc_info=True)
        _set_export_task_status(
            task_id,
            "canceled",
            progress=100,
            message="导出已取消。",
            payload=payload,
        )
    except Exception as exc:
        logger.exception("resource export task failed task_id=%s", task_id)
        _set_export_task_status(
            task_id,
            "failed",
            progress=100,
            message="导出失败。",
            payload=payload,
            error=str(exc),
        )


def _write_xlsx_file(
    *,
    output_path: Path,
    sheet_title: str,
    headers: list[str],
    row_iter: Any,
    total_entries: int,
    progress_callback: ProgressCallback,
) -> None:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=_normalize_sheet_title(sheet_title))
    for index, width in enumerate([48, 48, 16, 16, 22, 22], start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    worksheet.append(headers)
    processed = 0
    for processed, row in enumerate(row_iter, start=1):
        worksheet.append(row)
        _report_export_progress(
            processed=processed,
            total_entries=total_entries,
            progress_callback=progress_callback,
            message="正在写入 Excel 文件。",
        )

    if processed == 0:
        progress_callback(90, "正在写入空 Excel 文件。")

    progress_callback(92, "正在保存 Excel 文件。")
    workbook.save(output_path)
    workbook.close()


def _write_tmx_file(
    *,
    output_path: Path,
    source_language: str,
    target_language: str,
    rows: Any,
    total_entries: int,
    progress_callback: ProgressCallback,
    filename: str,
    tuid_prefix: str,
) -> None:
    creation_date = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    with output_path.open("w", encoding="utf-8", newline="\n") as file:
        file.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        file.write('<!DOCTYPE tmx SYSTEM "tmx14.dtd">\n')
        file.write('<tmx version="1.4">\n')
        file.write("  <header\n")
        file.write('    creationtool="Translation Memory Demo"\n')
        file.write('    creationtoolversion="1.0"\n')
        file.write('    datatype="plaintext"\n')
        file.write('    segtype="sentence"\n')
        file.write('    adminlang="en-US"\n')
        file.write(f"    srclang={quoteattr(source_language)}\n")
        file.write(f"    creationdate={quoteattr(creation_date)}\n")
        file.write(f"    o-tmf={quoteattr(filename)}\n")
        file.write("  />\n")
        file.write("  <body>\n")

        processed = 0
        for processed, row in enumerate(rows, start=1):
            row_id, source_text, target_text = row
            if not source_text:
                continue

            tuid = f"{tuid_prefix}_{row_id}"
            file.write(f"    <tu tuid={quoteattr(tuid)}>\n")
            file.write(f"      <tuv xml:lang={quoteattr(source_language)}>\n")
            file.write(f"        <seg>{escape(str(source_text))}</seg>\n")
            file.write("      </tuv>\n")
            file.write(f"      <tuv xml:lang={quoteattr(target_language)}>\n")
            file.write(f"        <seg>{escape(str(target_text or ''))}</seg>\n")
            file.write("      </tuv>\n")
            file.write("    </tu>\n")
            _report_export_progress(
                processed=processed,
                total_entries=total_entries,
                progress_callback=progress_callback,
                message="正在写入 TMX 文件。",
            )

        if processed == 0:
            progress_callback(90, "正在写入空 TMX 文件。")

        progress_callback(92, "正在保存 TMX 文件。")
        file.write("  </body>\n")
        file.write("</tmx>\n")


def _iter_tm_xlsx_rows(db: Session, collection_id: UUID):
    from app.models import TranslationMemory, User

    Creator = aliased(User)
    Modifier = aliased(User)

    query = (
        db.query(
            TranslationMemory.source_text,
            TranslationMemory.target_text,
            TranslationMemory.source_language,
            TranslationMemory.target_language,
            Creator.nickname.label("creator_nickname"),
            Creator.username.label("creator_username"),
            TranslationMemory.created_at,
            Modifier.nickname.label("modifier_nickname"),
            Modifier.username.label("modifier_username"),
            TranslationMemory.updated_at,
        )
        .outerjoin(Creator, TranslationMemory.creator_id == Creator.id)
        .outerjoin(Modifier, TranslationMemory.last_modified_by_id == Modifier.id)
        .filter(TranslationMemory.collection_id == collection_id)
        .order_by(TranslationMemory.updated_at.desc(), TranslationMemory.created_at.desc())
        .execution_options(stream_results=True)
    )
    for row in query.yield_per(EXPORT_QUERY_BATCH_SIZE):
        yield [
            row.source_text,
            row.target_text,
            row.source_language or "",
            row.target_language or "",
            _format_user_name(row.creator_nickname, row.creator_username),
            _format_datetime(row.created_at),
            _format_user_name(row.modifier_nickname, row.modifier_username),
            _format_datetime(row.updated_at),
        ]


def _iter_term_xlsx_rows(db: Session, term_base_id: UUID):
    from app.models import TermEntry, User

    Creator = aliased(User)
    Modifier = aliased(User)

    query = (
        db.query(
            TermEntry.source_text,
            TermEntry.target_text,
            TermEntry.source_language,
            TermEntry.target_language,
            Creator.nickname.label("creator_nickname"),
            Creator.username.label("creator_username"),
            TermEntry.created_at,
            Modifier.nickname.label("modifier_nickname"),
            Modifier.username.label("modifier_username"),
            TermEntry.updated_at,
        )
        .outerjoin(Creator, TermEntry.creator_id == Creator.id)
        .outerjoin(Modifier, TermEntry.last_modified_by_id == Modifier.id)
        .filter(TermEntry.term_base_id == term_base_id)
        .order_by(TermEntry.updated_at.desc(), TermEntry.created_at.desc())
        .execution_options(stream_results=True)
    )
    for row in query.yield_per(EXPORT_QUERY_BATCH_SIZE):
        yield [
            row.source_text,
            row.target_text,
            row.source_language or "",
            row.target_language or "",
            _format_user_name(row.creator_nickname, row.creator_username),
            _format_datetime(row.created_at),
            _format_user_name(row.modifier_nickname, row.modifier_username),
            _format_datetime(row.updated_at),
        ]


def _iter_glossary_xlsx_rows(db: Session, glossary_base_id: UUID):
    from app.models import GlossaryEntry, User

    Creator = aliased(User)
    Modifier = aliased(User)

    query = (
        db.query(
            GlossaryEntry.source_text,
            GlossaryEntry.target_text,
            GlossaryEntry.note,
            GlossaryEntry.source_language,
            GlossaryEntry.target_language,
            Creator.nickname.label("creator_nickname"),
            Creator.username.label("creator_username"),
            GlossaryEntry.created_at,
            Modifier.nickname.label("modifier_nickname"),
            Modifier.username.label("modifier_username"),
            GlossaryEntry.updated_at,
        )
        .outerjoin(Creator, GlossaryEntry.creator_id == Creator.id)
        .outerjoin(Modifier, GlossaryEntry.last_modified_by_id == Modifier.id)
        .filter(GlossaryEntry.glossary_base_id == glossary_base_id)
        .order_by(GlossaryEntry.updated_at.desc(), GlossaryEntry.created_at.desc())
        .execution_options(stream_results=True)
    )
    for row in query.yield_per(EXPORT_QUERY_BATCH_SIZE):
        yield [
            row.source_text,
            row.target_text,
            row.note or "",
            row.source_language or "",
            row.target_language or "",
            _format_user_name(row.creator_nickname, row.creator_username),
            _format_datetime(row.created_at),
            _format_user_name(row.modifier_nickname, row.modifier_username),
            _format_datetime(row.updated_at),
        ]


def _iter_tm_tmx_rows(db: Session, collection_id: UUID):
    from app.models import TranslationMemory

    query = (
        db.query(
            TranslationMemory.id,
            TranslationMemory.source_text,
            TranslationMemory.target_text,
        )
        .filter(TranslationMemory.collection_id == collection_id)
        .order_by(TranslationMemory.updated_at.desc(), TranslationMemory.created_at.desc())
        .execution_options(stream_results=True)
    )
    for row in query.yield_per(EXPORT_QUERY_BATCH_SIZE):
        yield row.id, row.source_text, row.target_text


def _iter_term_tmx_rows(db: Session, term_base_id: UUID):
    from app.models import TermEntry

    query = (
        db.query(
            TermEntry.id,
            TermEntry.source_text,
            TermEntry.target_text,
        )
        .filter(TermEntry.term_base_id == term_base_id)
        .order_by(TermEntry.updated_at.desc(), TermEntry.created_at.desc())
        .execution_options(stream_results=True)
    )
    for row in query.yield_per(EXPORT_QUERY_BATCH_SIZE):
        yield row.id, row.source_text, row.target_text


def _iter_glossary_tmx_rows(db: Session, glossary_base_id: UUID):
    from app.models import GlossaryEntry

    query = (
        db.query(
            GlossaryEntry.id,
            GlossaryEntry.source_text,
            GlossaryEntry.target_text,
        )
        .filter(GlossaryEntry.glossary_base_id == glossary_base_id)
        .order_by(GlossaryEntry.updated_at.desc(), GlossaryEntry.created_at.desc())
        .execution_options(stream_results=True)
    )
    for row in query.yield_per(EXPORT_QUERY_BATCH_SIZE):
        yield row.id, row.source_text, row.target_text


def _set_export_task_status(
    task_id: str,
    status: ResourceExportStatus,
    *,
    progress: int,
    message: str,
    payload: dict[str, Any],
    result: dict[str, Any] | None = None,
    error: str | None = None,
) -> dict[str, Any]:
    existing = get_resource_export_task_status(task_id) or {}
    now = datetime.now(timezone.utc).isoformat()
    task_payload: dict[str, Any] = {
        "task_id": task_id,
        "resource_type": payload.get("resource_type"),
        "resource_id": payload.get("resource_id"),
        "format": payload.get("format"),
        "status": status,
        "progress": max(0, min(100, int(progress))),
        "message": message,
        "result": result,
        "error": error,
        "created_at": existing.get("created_at") or now,
        "updated_at": now,
    }
    cache_set_json(_export_task_cache_key(task_id), task_payload, ttl_seconds=EXPORT_TASK_TTL_SECONDS)
    return task_payload


def _build_export_result(
    *,
    resource_name: str,
    filename_suffix: str,
    export_format: ResourceExportFormat,
    output_path: Path,
    total_entries: int,
) -> dict[str, Any]:
    filename = f"{_sanitize_download_name(resource_name)}-{filename_suffix}.{export_format}"
    return {
        "filename": filename,
        "file_path": str(output_path),
        "media_type": EXPORT_MEDIA_TYPES[export_format],
        "size_bytes": output_path.stat().st_size,
        "total_entries": total_entries,
    }


def _report_export_progress(
    *,
    processed: int,
    total_entries: int,
    progress_callback: ProgressCallback,
    message: str,
) -> None:
    if processed % EXPORT_QUERY_BATCH_SIZE != 0 and processed != total_entries:
        return
    if total_entries <= 0:
        progress_callback(90, message)
        return
    progress = 10 + int((processed / total_entries) * 80)
    progress_callback(min(progress, 90), message)


def _ensure_export_dir() -> Path:
    output_dir = Path(get_settings().export_task_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def _cleanup_expired_export_files(output_dir: Path) -> None:
    cutoff = time.time() - EXPORT_TASK_TTL_SECONDS
    for path in output_dir.glob("*"):
        try:
            if path.is_file() and path.stat().st_mtime < cutoff:
                path.unlink()
        except OSError:
            logger.debug("skip cleanup for export file path=%s", path, exc_info=True)


def _export_task_cache_key(task_id: str) -> str:
    return f"resource-export-task:{task_id}"


def _normalize_sheet_title(value: str) -> str:
    cleaned = "".join(char for char in value if char not in '\\/*?:[]')
    cleaned = cleaned.strip() or "Sheet1"
    return cleaned[:31]


def _sanitize_download_name(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip()
    return cleaned[:80] or "export"


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _format_user_name(nickname: str | None, username: str | None) -> str:
    return nickname or username or ""


def _noop_progress_callback(progress: int, message: str) -> None:
    return None


def _log_export_task_failure(future: Future) -> None:
    for task_id, candidate in list(_EXPORT_FUTURES.items()):
        if candidate is future:
            _EXPORT_FUTURES.pop(task_id, None)
            break
    try:
        future.result()
    except CancelledError:
        return
    except Exception:
        logger.exception("resource export worker crashed")
