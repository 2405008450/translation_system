from __future__ import annotations

import csv
import hashlib
import json
import re
import sqlite3
import tempfile
import time
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Literal
from uuid import UUID, uuid4

from openpyxl import load_workbook
from sqlalchemy import or_, text
from sqlalchemy.exc import DBAPIError
from sqlalchemy.orm import Session

from app.models import MemoryEntry
from app.services.import_task_state import ImportTaskCanceled
from app.services.language_pairs import require_language_pair
from app.services.normalizer import build_source_hash, normalize_match_text, normalize_text
from app.services.tmx_stream import TMXRow, iter_tmx_rows as iter_stream_tmx_rows
from app.services.tm_vector import sync_tm_embeddings


CSV_EXTENSIONS = {".csv"}
TMX_EXTENSIONS = {".tmx"}
XLS_EXTENSIONS = {".xls"}
XLSX_EXTENSIONS = {".xlsx"}
SDLTM_EXTENSIONS = {".sdltm"}
WORKBOOK_EXTENSIONS = XLSX_EXTENSIONS | XLS_EXTENSIONS
TM_IMPORT_EXTENSIONS = CSV_EXTENSIONS | TMX_EXTENSIONS | WORKBOOK_EXTENSIONS | SDLTM_EXTENSIONS
TM_STATUS_LOOKUP_CHUNK_SIZE = 10000
TM_PREVIEW_MAX_SCAN_ROWS = 1000
TM_IMPORT_WRITE_RETRY_ATTEMPTS = 5
TM_IMPORT_WRITE_RETRY_BASE_DELAY_SECONDS = 0.2
TM_IMPORT_RETRYABLE_SQLSTATES = {"40P01", "40001", "55P03"}
HEADER_ALIASES = {
    ("zh-cn", "en-us"),
    ("source", "target"),
    ("source_text", "target_text"),
    ("源文", "译文"),
    ("原文", "译文"),
    ("中文", "英文"),
}

CancelCheck = Callable[[], bool]


def _raise_if_canceled(cancel_check: CancelCheck | None) -> None:
    if cancel_check is not None and cancel_check():
        raise ImportTaskCanceled("记忆库导入已取消。")


@dataclass
class TMImportSummary:
    filename: str
    created_rows: int
    updated_rows: int
    skipped_empty_rows: int
    skipped_header_rows: int
    skipped_duplicate_rows: int = 0

    @property
    def imported_rows(self) -> int:
        return self.created_rows + self.updated_rows


@dataclass
class TMImportPreviewRow:
    row_index: int
    source_text: str
    target_text: str
    status: str
    message: str


@dataclass
class TMImportPreview:
    filename: str
    rows: list[TMImportPreviewRow]
    total_rows: int
    valid_rows: int
    create_rows: int
    update_rows: int
    keep_rows: int
    duplicate_rows: int
    skipped_empty_rows: int
    skipped_header_rows: int
    preview_limit: int
    duplicate_policy: str
    scanned_rows: int = 0
    truncated: bool = False


DuplicatePolicy = Literal["overwrite", "keep"]


@dataclass
class SDLTMMetadata:
    """Metadata extracted from SDLTM file."""
    name: str
    source_language: str
    target_language: str
    entry_count: int


def preview_sdltm_metadata(raw_bytes: bytes) -> SDLTMMetadata:
    """Extract metadata from SDLTM file without importing."""
    with tempfile.NamedTemporaryFile(suffix=".sdltm", delete=False) as tmp:
        tmp.write(raw_bytes)
        tmp_path = tmp.name

    try:
        return preview_sdltm_metadata_from_path(tmp_path)
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def preview_sdltm_metadata_from_path(sdltm_path: str | Path) -> SDLTMMetadata:
    """Extract metadata from an SDLTM file path without importing."""
    conn = None
    try:
        conn = sqlite3.connect(str(sdltm_path))
        cursor = conn.cursor()

        # Get TM metadata
        cursor.execute("""
            SELECT name, source_language, target_language
            FROM translation_memories
            LIMIT 1
        """)
        row = cursor.fetchone()

        if row:
            name, source_lang, target_lang = row
        else:
            name, source_lang, target_lang = "", "", ""

        # Get entry count
        cursor.execute("SELECT COUNT(*) FROM translation_units")
        entry_count = cursor.fetchone()[0]

        return SDLTMMetadata(
            name=name or "",
            source_language=source_lang or "",
            target_language=target_lang or "",
            entry_count=entry_count,
        )
    finally:
        if conn is not None:
            conn.close()


def import_tm_from_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    batch_size: int = 5000,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    extension = f".{filename.rsplit('.', 1)[-1].lower()}" if "." in filename else ""
    if extension in CSV_EXTENSIONS:
        return import_tm_from_csv_upload(
            db=db,
            raw_bytes=raw_bytes,
            filename=filename,
            source_language=source_language,
            target_language=target_language,
            batch_size=batch_size,
            collection_id=collection_id,
            creator_id=creator_id,
            duplicate_policy=duplicate_policy,
            skip_duplicate_row_indexes=skip_duplicate_row_indexes,
            skip_header=skip_header,
            cancel_check=cancel_check,
            import_batch_id=import_batch_id,
        )
    if extension in TMX_EXTENSIONS:
        return import_tm_from_tmx_upload(
            db=db,
            raw_bytes=raw_bytes,
            filename=filename,
            source_language=source_language,
            target_language=target_language,
            batch_size=batch_size,
            collection_id=collection_id,
            creator_id=creator_id,
            duplicate_policy=duplicate_policy,
            skip_duplicate_row_indexes=skip_duplicate_row_indexes,
            cancel_check=cancel_check,
            import_batch_id=import_batch_id,
        )
    return import_tm_from_xlsx_upload(
        db=db,
        raw_bytes=raw_bytes,
        filename=filename,
        source_language=source_language,
        target_language=target_language,
        batch_size=batch_size,
        collection_id=collection_id,
        creator_id=creator_id,
        duplicate_policy=duplicate_policy,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        skip_header=skip_header,
        cancel_check=cancel_check,
        import_batch_id=import_batch_id,
    )


def import_tm_from_xlsx_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    batch_size: int = 5000,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    extension = f".{filename.rsplit('.', 1)[-1].lower()}" if "." in filename else ""
    if extension in XLS_EXTENSIONS:
        rows = _iter_xls_rows(raw_bytes)
        return _import_text_rows(
            db=db,
            rows=rows,
            filename=filename,
            batch_size=batch_size,
            collection_id=collection_id,
            creator_id=creator_id,
            duplicate_policy=duplicate_policy,
            skip_duplicate_row_indexes=skip_duplicate_row_indexes,
            skip_header=skip_header,
            source_language=source_language,
            target_language=target_language,
            cancel_check=cancel_check,
            import_batch_id=import_batch_id,
        )

    workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    return _import_workbook(
        db=db,
        workbook=workbook,
        filename=filename,
        batch_size=batch_size,
        collection_id=collection_id,
        creator_id=creator_id,
        duplicate_policy=duplicate_policy,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        skip_header=skip_header,
        source_language=source_language,
        target_language=target_language,
        cancel_check=cancel_check,
        import_batch_id=import_batch_id,
    )


def import_tm_from_xlsx_path(
    db: Session,
    xlsx_path: str | Path,
    source_language: str,
    target_language: str,
    filename: str | None = None,
    batch_size: int = 5000,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    workbook = load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    return _import_workbook(
        db=db,
        workbook=workbook,
        filename=filename or Path(xlsx_path).name,
        batch_size=batch_size,
        collection_id=collection_id,
        creator_id=creator_id,
        duplicate_policy=duplicate_policy,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        skip_header=skip_header,
        source_language=source_language,
        target_language=target_language,
        cancel_check=cancel_check,
        import_batch_id=import_batch_id,
    )


def import_tm_from_csv_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    batch_size: int = 5000,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    rows = _iter_csv_rows(raw_bytes)
    return _import_text_rows(
        db=db,
        rows=rows,
        filename=filename,
        batch_size=batch_size,
        collection_id=collection_id,
        creator_id=creator_id,
        duplicate_policy=duplicate_policy,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        skip_header=skip_header,
        source_language=source_language,
        target_language=target_language,
        cancel_check=cancel_check,
        import_batch_id=import_batch_id,
    )


def import_tm_from_tmx_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    batch_size: int = 5000,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    rows = _iter_tmx_rows(raw_bytes, normalized_source_language, normalized_target_language)
    return _import_text_rows(
        db=db,
        rows=rows,
        filename=filename,
        batch_size=batch_size,
        collection_id=collection_id,
        creator_id=creator_id,
        duplicate_policy=duplicate_policy,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        source_language=normalized_source_language,
        target_language=normalized_target_language,
        cancel_check=cancel_check,
        import_batch_id=import_batch_id,
    )


def import_tm_from_tmx_path(
    db: Session,
    tmx_path: str | Path,
    filename: str,
    source_language: str,
    target_language: str,
    batch_size: int = 5000,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    return _import_text_rows(
        db=db,
        rows=_iter_tmx_path_rows(
            tmx_path,
            normalized_source_language,
            normalized_target_language,
        ),
        filename=filename,
        batch_size=batch_size,
        collection_id=collection_id,
        creator_id=creator_id,
        duplicate_policy=duplicate_policy,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        source_language=normalized_source_language,
        target_language=normalized_target_language,
        cancel_check=cancel_check,
        import_batch_id=import_batch_id,
    )


def preview_tm_from_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    collection_id: UUID | None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = TM_PREVIEW_MAX_SCAN_ROWS,
) -> TMImportPreview:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    rows = _iter_upload_rows(
        raw_bytes,
        filename,
        normalized_source_language,
        normalized_target_language,
    )
    preview_rows: list[TMImportPreviewRow] = []
    source_hashes: set[str] = set()
    source_texts: set[str] = set()
    final_candidates: dict[str, dict] = {}
    skipped_empty_rows = 0
    skipped_header_rows = 0
    total_rows = 0
    duplicate_rows = 0
    safe_max_scan_rows = max(1, int(max_scan_rows or TM_PREVIEW_MAX_SCAN_ROWS))
    truncated = False

    for row_index, row in enumerate(rows, start=1):
        if row_index > safe_max_scan_rows:
            truncated = True
            break
        total_rows += 1
        source_text = normalize_text(_cell_to_text(row, 0))
        target_text = normalize_text(_cell_to_text(row, 1))

        if _should_skip_header_row(filename, row_index, source_text, target_text, skip_header):
            skipped_header_rows += 1
            _append_preview_row(
                preview_rows,
                preview_limit,
                TMImportPreviewRow(row_index, source_text, target_text, "header", "表头行，导入时会跳过。"),
            )
            continue

        if not source_text or not target_text:
            skipped_empty_rows += 1
            _append_preview_row(
                preview_rows,
                preview_limit,
                TMImportPreviewRow(row_index, source_text, target_text, "empty", "源文或译文为空，导入时会跳过。"),
            )
            continue

        tm_row = _build_tm_row(
            source_text=source_text,
            target_text=target_text,
            source_language=normalized_source_language,
            target_language=normalized_target_language,
            collection_id=collection_id,
        )
        source_hash = tm_row["source_hash"]
        status = "pending"
        message = "待导入。"
        if source_hash in source_hashes or source_text in source_texts:
            duplicate_rows += 1
            status = "duplicate"
            message = "文件内重复，导入时会保留首次出现的数据。" if duplicate_policy == "keep" else "文件内重复，导入时以后出现的这一条为准。"
            if duplicate_policy == "keep":
                _append_preview_row(
                    preview_rows,
                    preview_limit,
                    TMImportPreviewRow(row_index, source_text, target_text, status, message),
                )
                continue

        source_hashes.add(source_hash)
        source_texts.add(source_text)
        final_candidates[source_hash] = tm_row
        _append_preview_row(
            preview_rows,
            preview_limit,
            TMImportPreviewRow(row_index, source_text, target_text, status, message),
        )

    existing_status = _load_existing_tm_status(
        db=db,
        collection_id=collection_id,
        source_language=normalized_source_language,
        target_language=normalized_target_language,
        source_hashes=list(source_hashes),
        source_texts=list(source_texts),
    )

    create_rows = 0
    update_rows = 0
    keep_rows = 0
    for candidate in final_candidates.values():
        exists = candidate["source_hash"] in existing_status or candidate["source_text"] in existing_status
        if not exists:
            create_rows += 1
        elif duplicate_policy == "keep":
            keep_rows += 1
        else:
            update_rows += 1

    for preview_row in preview_rows:
        if preview_row.status != "pending":
            continue
        source_hash = build_source_hash(preview_row.source_text)
        exists = source_hash in existing_status or preview_row.source_text in existing_status
        if not exists:
            preview_row.status = "create"
            preview_row.message = "导入时会新增。"
        elif duplicate_policy == "keep":
            preview_row.status = "keep"
            preview_row.message = "记忆库中已有相同源文，导入时会保留旧数据并跳过这一条。"
        else:
            preview_row.status = "update"
            preview_row.message = "记忆库中已有相同源文，导入时会用新译文覆盖。"

    return TMImportPreview(
        filename=filename,
        rows=preview_rows,
        total_rows=total_rows,
        valid_rows=len(final_candidates),
        create_rows=create_rows,
        update_rows=update_rows,
        keep_rows=keep_rows,
        duplicate_rows=duplicate_rows,
        skipped_empty_rows=skipped_empty_rows,
        skipped_header_rows=skipped_header_rows,
        preview_limit=preview_limit,
        duplicate_policy=duplicate_policy,
        scanned_rows=total_rows,
        truncated=truncated,
    )


def preview_tm_from_xlsx_path(
    db: Session,
    xlsx_path: str | Path,
    filename: str,
    source_language: str,
    target_language: str,
    collection_id: UUID | None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = TM_PREVIEW_MAX_SCAN_ROWS,
) -> TMImportPreview:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    workbook = load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    try:
        return _preview_tm_from_rows(
            db=db,
            rows=workbook.active.iter_rows(values_only=True),
            filename=filename,
            collection_id=collection_id,
            source_language=normalized_source_language,
            target_language=normalized_target_language,
            duplicate_policy=duplicate_policy,
            preview_limit=preview_limit,
            skip_header=skip_header,
            max_scan_rows=max_scan_rows,
        )
    finally:
        workbook.close()


def preview_tm_from_sdltm_path(
    db: Session,
    sdltm_path: str | Path,
    filename: str,
    source_language: str,
    target_language: str,
    collection_id: UUID | None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = TM_PREVIEW_MAX_SCAN_ROWS,
) -> TMImportPreview:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    return _preview_tm_from_rows(
        db=db,
        rows=_iter_sdltm_path_rows(sdltm_path),
        filename=filename,
        collection_id=collection_id,
        source_language=normalized_source_language,
        target_language=normalized_target_language,
        duplicate_policy=duplicate_policy,
        preview_limit=preview_limit,
        skip_header=skip_header,
        max_scan_rows=max_scan_rows,
    )


def preview_tm_from_tmx_path(
    db: Session,
    tmx_path: str | Path,
    filename: str,
    source_language: str,
    target_language: str,
    collection_id: UUID | None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = TM_PREVIEW_MAX_SCAN_ROWS,
) -> TMImportPreview:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    return _preview_tm_from_rows(
        db=db,
        rows=_iter_tmx_path_rows(
            tmx_path,
            normalized_source_language,
            normalized_target_language,
        ),
        filename=filename,
        collection_id=collection_id,
        source_language=normalized_source_language,
        target_language=normalized_target_language,
        duplicate_policy=duplicate_policy,
        preview_limit=preview_limit,
        skip_header=skip_header,
        max_scan_rows=max_scan_rows,
    )


def _preview_tm_from_rows(
    db: Session,
    rows,
    filename: str,
    collection_id: UUID | None,
    source_language: str,
    target_language: str,
    duplicate_policy: DuplicatePolicy = "overwrite",
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = TM_PREVIEW_MAX_SCAN_ROWS,
) -> TMImportPreview:
    preview_rows: list[TMImportPreviewRow] = []
    source_hashes: set[str] = set()
    source_texts: set[str] = set()
    final_candidates: dict[str, dict] = {}
    skipped_empty_rows = 0
    skipped_header_rows = 0
    total_rows = 0
    duplicate_rows = 0
    safe_max_scan_rows = max(1, int(max_scan_rows or TM_PREVIEW_MAX_SCAN_ROWS))
    truncated = False

    for row_index, row in enumerate(rows, start=1):
        if row_index > safe_max_scan_rows:
            truncated = True
            break
        total_rows += 1
        source_text = normalize_text(_cell_to_text(row, 0))
        target_text = normalize_text(_cell_to_text(row, 1))

        if _should_skip_header_row(filename, row_index, source_text, target_text, skip_header):
            skipped_header_rows += 1
            _append_preview_row(
                preview_rows,
                preview_limit,
                TMImportPreviewRow(row_index, source_text, target_text, "header", "表头行，导入时会跳过。"),
            )
            continue

        if not source_text or not target_text:
            skipped_empty_rows += 1
            _append_preview_row(
                preview_rows,
                preview_limit,
                TMImportPreviewRow(row_index, source_text, target_text, "empty", "源文或译文为空，导入时会跳过。"),
            )
            continue

        tm_row = _build_tm_row(
            source_text=source_text,
            target_text=target_text,
            source_language=source_language,
            target_language=target_language,
            collection_id=collection_id,
        )
        source_hash = tm_row["source_hash"]
        status = "pending"
        message = "待导入。"
        if source_hash in source_hashes or source_text in source_texts:
            duplicate_rows += 1
            status = "duplicate"
            message = (
                "文件内重复，导入时会保留首次出现的数据。"
                if duplicate_policy == "keep"
                else "文件内重复，导入时以后出现的这一条为准。"
            )
            if duplicate_policy == "keep":
                _append_preview_row(
                    preview_rows,
                    preview_limit,
                    TMImportPreviewRow(row_index, source_text, target_text, status, message),
                )
                continue

        source_hashes.add(source_hash)
        source_texts.add(source_text)
        final_candidates[source_hash] = tm_row
        _append_preview_row(
            preview_rows,
            preview_limit,
            TMImportPreviewRow(row_index, source_text, target_text, status, message),
        )

    existing_status = _load_existing_tm_status(
        db=db,
        collection_id=collection_id,
        source_language=source_language,
        target_language=target_language,
        source_hashes=list(source_hashes),
        source_texts=list(source_texts),
    )

    create_rows = 0
    update_rows = 0
    keep_rows = 0
    for candidate in final_candidates.values():
        exists = candidate["source_hash"] in existing_status or candidate["source_text"] in existing_status
        if not exists:
            create_rows += 1
        elif duplicate_policy == "keep":
            keep_rows += 1
        else:
            update_rows += 1

    for preview_row in preview_rows:
        if preview_row.status != "pending":
            continue
        source_hash = build_source_hash(preview_row.source_text)
        exists = source_hash in existing_status or preview_row.source_text in existing_status
        if not exists:
            preview_row.status = "create"
            preview_row.message = "导入时会新增。"
        elif duplicate_policy == "keep":
            preview_row.status = "keep"
            preview_row.message = "记忆库中已有相同源文，导入时会保留旧数据并跳过这一条。"
        else:
            preview_row.status = "update"
            preview_row.message = "记忆库中已有相同源文，导入时会用新译文覆盖。"

    return TMImportPreview(
        filename=filename,
        rows=preview_rows,
        total_rows=total_rows,
        valid_rows=len(final_candidates),
        create_rows=create_rows,
        update_rows=update_rows,
        keep_rows=keep_rows,
        duplicate_rows=duplicate_rows,
        skipped_empty_rows=skipped_empty_rows,
        skipped_header_rows=skipped_header_rows,
        preview_limit=preview_limit,
        duplicate_policy=duplicate_policy,
        scanned_rows=total_rows,
        truncated=truncated,
    )


def _iter_upload_rows(raw_bytes: bytes, filename: str, source_language: str, target_language: str):
    extension = f".{filename.rsplit('.', 1)[-1].lower()}" if "." in filename else ""
    if extension in CSV_EXTENSIONS:
        return _iter_csv_rows(raw_bytes)
    if extension in TMX_EXTENSIONS:
        return _iter_tmx_rows(raw_bytes, source_language, target_language)
    if extension in SDLTM_EXTENSIONS:
        return _iter_sdltm_rows(raw_bytes)
    if extension in XLS_EXTENSIONS:
        return _iter_xls_rows(raw_bytes)
    if extension not in XLSX_EXTENSIONS:
        raise RuntimeError("仅支持上传 .tmx、.sdltm、.xls、.xlsx 或 .csv 文件。")

    workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)

    def iter_workbook_rows():
        try:
            yield from workbook.active.iter_rows(values_only=True)
        finally:
            workbook.close()

    return iter_workbook_rows()


def _append_preview_row(
    rows: list[TMImportPreviewRow],
    preview_limit: int,
    row: TMImportPreviewRow,
) -> None:
    if len(rows) < preview_limit:
        rows.append(row)


def _row_to_metadata(row: Any) -> dict[str, Any]:
    if isinstance(row, TMXRow):
        return row.metadata
    if isinstance(row, dict):
        metadata = row.get("metadata")
        return metadata if isinstance(metadata, dict) else {}
    if isinstance(row, (tuple, list)) and len(row) >= 3 and isinstance(row[2], dict):
        return row[2]
    return {}


def _load_existing_tm_status(
    db: Session,
    collection_id: UUID | None,
    source_language: str,
    target_language: str,
    source_hashes: list[str],
    source_texts: list[str],
) -> set[str]:
    if not source_hashes and not source_texts:
        return set()

    keys: set[str] = set()
    source_hashes = list(dict.fromkeys(source_hashes))
    source_texts = list(dict.fromkeys(source_texts))
    max_lookup_count = max(len(source_hashes), len(source_texts))

    for offset in range(0, max_lookup_count, TM_STATUS_LOOKUP_CHUNK_SIZE):
        source_hash_chunk = source_hashes[offset : offset + TM_STATUS_LOOKUP_CHUNK_SIZE]
        source_text_chunk = source_texts[offset : offset + TM_STATUS_LOOKUP_CHUNK_SIZE]
        lookup_conditions = []
        if source_hash_chunk:
            lookup_conditions.append(MemoryEntry.source_hash.in_(source_hash_chunk))
        if source_text_chunk:
            lookup_conditions.append(MemoryEntry.source_text.in_(source_text_chunk))

        existing_query = (
            db.query(MemoryEntry.source_hash, MemoryEntry.source_text)
            .filter(
                or_(*lookup_conditions),
                MemoryEntry.source_language == source_language,
                MemoryEntry.target_language == target_language,
            )
        )
        if collection_id is None:
            existing_query = existing_query.filter(MemoryEntry.collection_id.is_(None))
        else:
            existing_query = existing_query.filter(MemoryEntry.collection_id == collection_id)

        for source_hash, source_text in existing_query.all():
            if source_hash:
                keys.add(source_hash)
            if source_text:
                keys.add(source_text)
    return keys


def _iter_csv_rows(raw_bytes: bytes):
    text = _decode_csv_bytes(raw_bytes)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(text.splitlines(), dialect)
    for row in reader:
        yield tuple(row)


def _iter_xls_rows(raw_bytes: bytes):
    try:
        import xlrd
    except ModuleNotFoundError as exc:
        raise RuntimeError("导入 .xls 文件需要安装 xlrd 依赖。") from exc

    workbook = xlrd.open_workbook(file_contents=raw_bytes)
    sheet = workbook.sheet_by_index(0)
    for row_index in range(sheet.nrows):
        yield tuple(sheet.cell_value(row_index, column_index) for column_index in range(sheet.ncols))


def _iter_tmx_rows(raw_bytes: bytes, source_language: str, target_language: str):
    yield from _iter_tmx_source_rows(raw_bytes, source_language, target_language)


def _iter_tmx_path_rows(tmx_path: str | Path, source_language: str, target_language: str):
    yield from _iter_tmx_source_rows(Path(tmx_path), source_language, target_language)


def _iter_tmx_source_rows(source: Any, source_language: str, target_language: str):
    for row in iter_stream_tmx_rows(source, source_language, target_language):
        yield (row.source_text, row.target_text, row.metadata)


def _iter_sdltm_rows(raw_bytes: bytes):
    with tempfile.NamedTemporaryFile(suffix=".sdltm", delete=False) as tmp:
        tmp.write(raw_bytes)
        tmp_path = tmp.name

    try:
        yield from _iter_sdltm_path_rows(tmp_path)
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def _iter_sdltm_path_rows(sdltm_path: str | Path):
    conn = None
    try:
        conn = sqlite3.connect(str(sdltm_path))
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT source_segment, target_segment
            FROM translation_units
            WHERE source_segment IS NOT NULL AND target_segment IS NOT NULL
            """
        )
        for raw_source, raw_target in cursor:
            yield (_extract_sdltm_text(raw_source), _extract_sdltm_text(raw_target))
    finally:
        if conn is not None:
            conn.close()


def _decode_csv_bytes(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8", errors="replace")


def _find_tmx_language_text(language_segments: list[tuple[str, str]], language: str) -> str:
    normalized_language = _normalize_language_tag(language)
    for candidate_language, text in language_segments:
        if _normalize_language_tag(candidate_language) == normalized_language:
            return text
    primary_language = normalized_language.split("-", 1)[0]
    for candidate_language, text in language_segments:
        if _normalize_language_tag(candidate_language).split("-", 1)[0] == primary_language:
            return text
    return ""


def _import_workbook(
    db: Session,
    workbook,
    filename: str,
    batch_size: int,
    source_language: str,
    target_language: str,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    worksheet = workbook.active
    try:
        return _import_text_rows(
            db=db,
            rows=worksheet.iter_rows(values_only=True),
            filename=filename,
            batch_size=batch_size,
            collection_id=collection_id,
            creator_id=creator_id,
            duplicate_policy=duplicate_policy,
            skip_duplicate_row_indexes=skip_duplicate_row_indexes,
            skip_header=skip_header,
            source_language=source_language,
            target_language=target_language,
            cancel_check=cancel_check,
            import_batch_id=import_batch_id,
        )
    finally:
        workbook.close()


def _import_text_rows(
    db: Session,
    rows,
    filename: str,
    batch_size: int,
    source_language: str,
    target_language: str,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    batch_rows: dict[str, dict] = {}
    created_rows = 0
    updated_rows = 0
    skipped_empty_rows = 0
    skipped_header_rows = 0
    skipped_duplicate_rows = 0
    skipped_row_indexes = skip_duplicate_row_indexes or set()

    for row_index, row in enumerate(rows, start=1):
        if row_index == 1 or row_index % batch_size == 0:
            _raise_if_canceled(cancel_check)
        row_metadata = _row_to_metadata(row)
        source_text = normalize_text(_cell_to_text(row, 0))
        target_text = normalize_text(_cell_to_text(row, 1))

        if _should_skip_header_row(filename, row_index, source_text, target_text, skip_header):
            skipped_header_rows += 1
            continue

        if not source_text or not target_text:
            skipped_empty_rows += 1
            continue

        if row_index in skipped_row_indexes:
            skipped_duplicate_rows += 1
            continue

        tm_row = _build_tm_row(
            source_text=source_text,
            target_text=target_text,
            source_language=normalized_source_language,
            target_language=normalized_target_language,
            collection_id=collection_id,
            creator_id=creator_id,
            external_tuid=str(row_metadata.get("tuid") or "") or None,
            tmx_metadata=row_metadata,
            import_batch_id=import_batch_id,
        )
        if duplicate_policy == "keep" and tm_row["source_hash"] in batch_rows:
            skipped_duplicate_rows += 1
            continue
        batch_rows[tm_row["source_hash"]] = tm_row

        if len(batch_rows) >= batch_size:
            created_in_batch, updated_in_batch, skipped_duplicate_in_batch = _flush_tm_batch(
                db=db,
                batch_rows=list(batch_rows.values()),
                collection_id=collection_id,
                creator_id=creator_id,
                duplicate_policy=duplicate_policy,
                source_language=normalized_source_language,
                target_language=normalized_target_language,
            )
            created_rows += created_in_batch
            updated_rows += updated_in_batch
            skipped_duplicate_rows += skipped_duplicate_in_batch
            batch_rows.clear()
            _raise_if_canceled(cancel_check)

    if batch_rows:
        _raise_if_canceled(cancel_check)
        created_in_batch, updated_in_batch, skipped_duplicate_in_batch = _flush_tm_batch(
            db=db,
            batch_rows=list(batch_rows.values()),
            collection_id=collection_id,
            creator_id=creator_id,
            duplicate_policy=duplicate_policy,
            source_language=normalized_source_language,
            target_language=normalized_target_language,
        )
        created_rows += created_in_batch
        updated_rows += updated_in_batch
        skipped_duplicate_rows += skipped_duplicate_in_batch

    return TMImportSummary(
        filename=filename,
        created_rows=created_rows,
        updated_rows=updated_rows,
        skipped_empty_rows=skipped_empty_rows,
        skipped_header_rows=skipped_header_rows,
        skipped_duplicate_rows=skipped_duplicate_rows,
    )


def _build_tm_row(
    source_text: str,
    target_text: str,
    source_language: str,
    target_language: str,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    external_tuid: str | None = None,
    tmx_metadata: dict[str, Any] | None = None,
    import_batch_id: UUID | None = None,
) -> dict:
    return {
        "collection_id": collection_id,
        "source_text": source_text,
        "target_text": target_text,
        "source_hash": build_source_hash(source_text),
        "source_normalized": normalize_match_text(source_text) or normalize_text(source_text),
        "source_language": source_language,
        "target_language": target_language,
        "creator_id": creator_id,
        "last_modified_by_id": creator_id,
        "external_tuid": external_tuid,
        "tmx_metadata": tmx_metadata or None,
        "import_batch_id": import_batch_id,
    }


def _flush_tm_batch(
    db: Session,
    batch_rows: list[dict],
    source_language: str,
    target_language: str,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
) -> tuple[int, int, int]:
    if not batch_rows:
        return 0, 0, 0
    ordered_batch_rows = sorted(batch_rows, key=_tm_import_row_sort_key)
    attempts = max(1, TM_IMPORT_WRITE_RETRY_ATTEMPTS)
    for attempt_index in range(attempts):
        try:
            return _flush_tm_batch_once(
                db=db,
                batch_rows=ordered_batch_rows,
                source_language=source_language,
                target_language=target_language,
                collection_id=collection_id,
                creator_id=creator_id,
                duplicate_policy=duplicate_policy,
            )
        except DBAPIError as exc:
            db.rollback()
            if not _is_retryable_tm_write_error(exc) or attempt_index + 1 >= attempts:
                raise
            delay_seconds = min(
                TM_IMPORT_WRITE_RETRY_BASE_DELAY_SECONDS * (2 ** attempt_index),
                2.0,
            )
            time.sleep(delay_seconds)

    return 0, 0, 0


def _flush_tm_batch_once(
    db: Session,
    batch_rows: list[dict],
    source_language: str,
    target_language: str,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
) -> tuple[int, int, int]:
    if not batch_rows:
        return 0, 0, 0

    if db.get_bind().dialect.name == "postgresql" and collection_id is not None:
        try:
            return _flush_tm_batch_postgres(
                db=db,
                batch_rows=batch_rows,
                source_language=source_language,
                target_language=target_language,
                collection_id=collection_id,
                duplicate_policy=duplicate_policy,
            )
        except DBAPIError as exc:
            db.rollback()
            if _is_retryable_tm_write_error(exc):
                raise
            return _flush_tm_batch_orm(
                db=db,
                batch_rows=batch_rows,
                source_language=source_language,
                target_language=target_language,
                collection_id=collection_id,
                creator_id=creator_id,
                duplicate_policy=duplicate_policy,
            )

    return _flush_tm_batch_orm(
        db=db,
        batch_rows=batch_rows,
        source_language=source_language,
        target_language=target_language,
        collection_id=collection_id,
        creator_id=creator_id,
        duplicate_policy=duplicate_policy,
    )


def _flush_tm_batch_orm(
    db: Session,
    batch_rows: list[dict],
    source_language: str,
    target_language: str,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
) -> tuple[int, int, int]:
    if not batch_rows:
        return 0, 0, 0

    _acquire_tm_import_transaction_lock(
        db=db,
        collection_id=collection_id,
        source_language=source_language,
        target_language=target_language,
    )

    source_hashes = [row["source_hash"] for row in batch_rows]
    source_texts = [row["source_text"] for row in batch_rows]
    existing_query = (
        db.query(MemoryEntry)
        .filter(
            or_(
                MemoryEntry.source_hash.in_(source_hashes),
                MemoryEntry.source_text.in_(source_texts),
            ),
            MemoryEntry.source_language == source_language,
            MemoryEntry.target_language == target_language,
        )
    )
    if collection_id is None:
        existing_query = existing_query.filter(MemoryEntry.collection_id.is_(None))
    else:
        existing_query = existing_query.filter(MemoryEntry.collection_id == collection_id)
    existing_rows = sorted(existing_query.all(), key=_memory_entry_sort_key)

    existing_by_hash: dict[str, MemoryEntry] = {}
    existing_by_source_text: dict[str, MemoryEntry] = {}
    for existing in existing_rows:
        if existing.source_hash:
            existing_by_hash.setdefault(existing.source_hash, existing)
        existing_by_source_text.setdefault(existing.source_text, existing)

    created_rows = 0
    updated_rows = 0
    skipped_duplicate_rows = 0
    sync_candidates: list[MemoryEntry] = []
    for row in batch_rows:
        existing = existing_by_hash.get(row["source_hash"]) or existing_by_source_text.get(
            row["source_text"]
        )
        if existing is None:
            tm_row = MemoryEntry(**row)
            db.add(tm_row)
            sync_candidates.append(tm_row)
            created_rows += 1
            continue

        if duplicate_policy == "keep":
            existing_by_hash[row["source_hash"]] = existing
            existing_by_source_text[row["source_text"]] = existing
            skipped_duplicate_rows += 1
            continue

        if _tm_entry_matches_import_row(existing, row):
            existing_by_hash[row["source_hash"]] = existing
            existing_by_source_text[row["source_text"]] = existing
            skipped_duplicate_rows += 1
            continue

        existing.source_text = row["source_text"]
        existing.target_text = row["target_text"]
        existing.source_hash = row["source_hash"]
        existing.source_normalized = row["source_normalized"]
        existing.collection_id = row["collection_id"]
        existing.source_language = row["source_language"]
        existing.target_language = row["target_language"]
        existing.external_tuid = row.get("external_tuid")
        existing.tmx_metadata = row.get("tmx_metadata")
        existing.import_batch_id = row.get("import_batch_id")
        if row.get("last_modified_by_id"):
            existing.last_modified_by_id = row["last_modified_by_id"]
        existing_by_hash[row["source_hash"]] = existing
        existing_by_source_text[row["source_text"]] = existing
        sync_candidates.append(existing)
        updated_rows += 1

    db.flush()
    sync_rows = sorted(
        {
            row.id: row.source_text
            for row in sync_candidates
            if row.id is not None and row.source_text
        }.items(),
        key=lambda item: str(item[0]),
    )
    db.commit()
    if sync_rows:
        sync_tm_embeddings(db, sync_rows)
    db.expunge_all()
    return created_rows, updated_rows, skipped_duplicate_rows


def _tm_entry_matches_import_row(existing: MemoryEntry, row: dict) -> bool:
    return (
        normalize_text(existing.source_text or "") == row["source_text"]
        and normalize_text(existing.target_text or "") == row["target_text"]
        and (existing.source_hash or "") == row["source_hash"]
        and existing.collection_id == row["collection_id"]
        and (existing.source_language or "") == row["source_language"]
        and (existing.target_language or "") == row["target_language"]
        and (getattr(existing, "external_tuid", None) or None) == row.get("external_tuid")
        and (getattr(existing, "tmx_metadata", None) or None) == row.get("tmx_metadata")
        and getattr(existing, "import_batch_id", None) == row.get("import_batch_id")
    )


def _flush_tm_batch_postgres(
    db: Session,
    batch_rows: list[dict],
    source_language: str,
    target_language: str,
    collection_id: UUID,
    duplicate_policy: DuplicatePolicy = "overwrite",
) -> tuple[int, int, int]:
    _acquire_tm_import_transaction_lock(
        db=db,
        collection_id=collection_id,
        source_language=source_language,
        target_language=target_language,
    )

    payload = [
        {
            "id": str(row.get("id") or uuid4()),
            "collection_id": str(collection_id),
            "source_text": row["source_text"],
            "target_text": row["target_text"],
            "source_hash": row["source_hash"],
            "source_normalized": row["source_normalized"],
            "source_language": row["source_language"],
            "target_language": row["target_language"],
            "creator_id": str(row["creator_id"]) if row.get("creator_id") else None,
            "last_modified_by_id": str(row["last_modified_by_id"]) if row.get("last_modified_by_id") else None,
            "external_tuid": row.get("external_tuid"),
            "tmx_metadata": row.get("tmx_metadata"),
            "import_batch_id": str(row["import_batch_id"]) if row.get("import_batch_id") else None,
        }
        for row in batch_rows
    ]
    if duplicate_policy == "keep":
        statement = text(
            """
            WITH incoming AS (
                SELECT *
                FROM jsonb_to_recordset(CAST(:payload AS jsonb)) AS item(
                    id uuid,
                    collection_id uuid,
                    source_text text,
                    target_text text,
                    source_hash text,
                    source_normalized text,
                    source_language text,
                    target_language text,
                    creator_id uuid,
                    last_modified_by_id uuid,
                    external_tuid text,
                    tmx_metadata jsonb,
                    import_batch_id uuid
                )
            )
            INSERT INTO memory_entries (
                id,
                collection_id,
                source_text,
                target_text,
                source_hash,
                source_normalized,
                source_language,
                target_language,
                creator_id,
                last_modified_by_id,
                external_tuid,
                tmx_metadata,
                import_batch_id
            )
            SELECT
                id,
                collection_id,
                source_text,
                target_text,
                source_hash,
                source_normalized,
                source_language,
                target_language,
                creator_id,
                last_modified_by_id,
                external_tuid,
                tmx_metadata,
                import_batch_id
            FROM incoming
            ORDER BY source_hash, source_text
            ON CONFLICT (collection_id, source_hash, source_language, target_language)
            DO NOTHING
            RETURNING id, source_text, TRUE AS inserted
            """
        )
    else:
        statement = text(
            """
            WITH incoming AS (
                SELECT *
                FROM jsonb_to_recordset(CAST(:payload AS jsonb)) AS item(
                    id uuid,
                    collection_id uuid,
                    source_text text,
                    target_text text,
                    source_hash text,
                    source_normalized text,
                    source_language text,
                    target_language text,
                    creator_id uuid,
                    last_modified_by_id uuid,
                    external_tuid text,
                    tmx_metadata jsonb,
                    import_batch_id uuid
                )
            )
            INSERT INTO memory_entries (
                id,
                collection_id,
                source_text,
                target_text,
                source_hash,
                source_normalized,
                source_language,
                target_language,
                creator_id,
                last_modified_by_id,
                external_tuid,
                tmx_metadata,
                import_batch_id
            )
            SELECT
                id,
                collection_id,
                source_text,
                target_text,
                source_hash,
                source_normalized,
                source_language,
                target_language,
                creator_id,
                last_modified_by_id,
                external_tuid,
                tmx_metadata,
                import_batch_id
            FROM incoming
            ORDER BY source_hash, source_text
            ON CONFLICT (collection_id, source_hash, source_language, target_language)
            DO UPDATE SET
                source_text = EXCLUDED.source_text,
                target_text = EXCLUDED.target_text,
                source_normalized = EXCLUDED.source_normalized,
                creator_id = COALESCE(memory_entries.creator_id, EXCLUDED.creator_id),
                last_modified_by_id = COALESCE(EXCLUDED.last_modified_by_id, memory_entries.last_modified_by_id),
                external_tuid = EXCLUDED.external_tuid,
                tmx_metadata = EXCLUDED.tmx_metadata,
                import_batch_id = EXCLUDED.import_batch_id,
                updated_at = NOW()
            WHERE
                memory_entries.source_text IS DISTINCT FROM EXCLUDED.source_text
                OR memory_entries.target_text IS DISTINCT FROM EXCLUDED.target_text
                OR memory_entries.external_tuid IS DISTINCT FROM EXCLUDED.external_tuid
                OR memory_entries.tmx_metadata IS DISTINCT FROM EXCLUDED.tmx_metadata
                OR memory_entries.import_batch_id IS DISTINCT FROM EXCLUDED.import_batch_id
            RETURNING id, source_text, (xmax = 0) AS inserted
            """
        )

    returned_rows = list(db.execute(statement, {"payload": json.dumps(payload, ensure_ascii=False)}))
    created_rows = 0
    updated_rows = 0
    sync_rows: list[tuple[UUID | str, str]] = []
    for returned in returned_rows:
        row_id, source_text, inserted = _read_tm_upsert_returned_row(returned)
        if inserted:
            created_rows += 1
        else:
            updated_rows += 1
        if row_id is not None and source_text:
            sync_rows.append((row_id, source_text))

    db.commit()
    if sync_rows:
        sync_tm_embeddings(db, sorted(sync_rows, key=lambda item: str(item[0])))
    db.expunge_all()
    skipped_duplicate_rows = len(batch_rows) - created_rows - updated_rows
    return created_rows, updated_rows, skipped_duplicate_rows


def _read_tm_upsert_returned_row(row) -> tuple[UUID | str | None, str, bool]:
    mapping = getattr(row, "_mapping", None)
    if mapping is not None:
        return mapping.get("id"), mapping.get("source_text") or "", bool(mapping.get("inserted"))
    return getattr(row, "id", None), getattr(row, "source_text", "") or "", bool(getattr(row, "inserted", False))


def _tm_import_row_sort_key(row: dict) -> tuple[str, str]:
    return (str(row.get("source_hash") or ""), normalize_text(str(row.get("source_text") or "")))


def _memory_entry_sort_key(entry: MemoryEntry) -> tuple[str, str, str]:
    return (str(entry.id or ""), str(entry.source_hash or ""), normalize_text(entry.source_text or ""))


def _is_retryable_tm_write_error(exc: DBAPIError) -> bool:
    original = getattr(exc, "orig", None)
    sqlstate = getattr(original, "sqlstate", None) or getattr(original, "pgcode", None)
    if sqlstate in TM_IMPORT_RETRYABLE_SQLSTATES:
        return True

    message = str(exc).lower()
    return (
        "deadlock detected" in message
        or "could not serialize access" in message
        or "lock not available" in message
    )


def _acquire_tm_import_transaction_lock(
    db: Session,
    collection_id: UUID | None,
    source_language: str,
    target_language: str,
) -> None:
    if db.get_bind().dialect.name != "postgresql":
        return

    lock_key = _build_tm_import_lock_key(collection_id, source_language, target_language)
    db.execute(text("SELECT pg_advisory_xact_lock(:lock_key)"), {"lock_key": lock_key})


def _build_tm_import_lock_key(
    collection_id: UUID | None,
    source_language: str,
    target_language: str,
) -> int:
    lock_scope = f"tm-import:{collection_id or 'global'}:{source_language}:{target_language}"
    digest = hashlib.blake2b(lock_scope.encode("utf-8"), digest_size=8).digest()
    return int.from_bytes(digest, byteorder="big", signed=True)


def _cell_to_text(row: tuple, index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value)


def _normalize_language_tag(language: str) -> str:
    return (language or "").strip().lower().replace("_", "-")


def _looks_like_header(source_text: str, target_text: str) -> bool:
    if not source_text or not target_text:
        return False

    header_key = (source_text.lower(), target_text.lower())
    return header_key in HEADER_ALIASES


def _is_table_import(filename: str) -> bool:
    extension = f".{filename.rsplit('.', 1)[-1].lower()}" if "." in filename else ""
    return extension in CSV_EXTENSIONS or extension in XLS_EXTENSIONS or extension in XLSX_EXTENSIONS


def _should_skip_header_row(
    filename: str,
    row_index: int,
    source_text: str,
    target_text: str,
    skip_header: bool,
) -> bool:
    if row_index != 1:
        return False
    return _looks_like_header(source_text, target_text) or (skip_header and _is_table_import(filename))


# ============================================================================
# SDLTM Import Functions
# ============================================================================

def import_tm_from_sdltm_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    batch_size: int = 5000,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    """Import translation memory from an uploaded SDLTM file."""
    with tempfile.NamedTemporaryFile(suffix=".sdltm", delete=False) as tmp:
        tmp.write(raw_bytes)
        tmp_path = tmp.name

    try:
        return _import_sdltm(
            db=db,
            sdltm_path=tmp_path,
            filename=filename,
            batch_size=batch_size,
            collection_id=collection_id,
            creator_id=creator_id,
            duplicate_policy=duplicate_policy,
            skip_duplicate_row_indexes=skip_duplicate_row_indexes,
            source_language=source_language,
            target_language=target_language,
            cancel_check=cancel_check,
            import_batch_id=import_batch_id,
        )
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def import_tm_from_sdltm_path(
    db: Session,
    sdltm_path: str | Path,
    source_language: str,
    target_language: str,
    filename: str | None = None,
    batch_size: int = 5000,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    """Import translation memory from an SDLTM file path."""
    return _import_sdltm(
        db=db,
        sdltm_path=str(sdltm_path),
        filename=filename or Path(sdltm_path).name,
        batch_size=batch_size,
        collection_id=collection_id,
        creator_id=creator_id,
        duplicate_policy=duplicate_policy,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        source_language=source_language,
        target_language=target_language,
        cancel_check=cancel_check,
        import_batch_id=import_batch_id,
    )


def _import_sdltm(
    db: Session,
    sdltm_path: str,
    filename: str,
    batch_size: int,
    source_language: str,
    target_language: str,
    collection_id: UUID | None = None,
    creator_id: UUID | None = None,
    duplicate_policy: DuplicatePolicy = "overwrite",
    skip_duplicate_row_indexes: set[int] | None = None,
    cancel_check: CancelCheck | None = None,
    import_batch_id: UUID | None = None,
) -> TMImportSummary:
    """Core SDLTM import logic."""
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )

    batch_rows: dict[str, dict] = {}
    created_rows = 0
    updated_rows = 0
    skipped_empty_rows = 0
    skipped_duplicate_rows = 0
    skipped_row_indexes = skip_duplicate_row_indexes or set()

    conn = sqlite3.connect(sdltm_path)
    try:
        cursor = conn.cursor()

        # SDLTM stores source/target as XML in translation_units table
        query = """
            SELECT
                id,
                source_segment,
                target_segment
            FROM translation_units
            WHERE source_segment IS NOT NULL AND target_segment IS NOT NULL
        """

        cursor.execute(query)

        for row_index, row in enumerate(cursor, start=1):
            if row_index == 1 or row_index % batch_size == 0:
                _raise_if_canceled(cancel_check)
            _, raw_source, raw_target = row

            # Extract text from SDLTM XML format
            source_text = normalize_text(_extract_sdltm_text(raw_source))
            target_text = normalize_text(_extract_sdltm_text(raw_target))

            if not source_text or not target_text:
                skipped_empty_rows += 1
                continue

            if row_index in skipped_row_indexes:
                skipped_duplicate_rows += 1
                continue

            tm_row = _build_tm_row(
                source_text=source_text,
                target_text=target_text,
                source_language=normalized_source_language,
                target_language=normalized_target_language,
                collection_id=collection_id,
                creator_id=creator_id,
                import_batch_id=import_batch_id,
            )
            if duplicate_policy == "keep" and tm_row["source_hash"] in batch_rows:
                skipped_duplicate_rows += 1
                continue
            batch_rows[tm_row["source_hash"]] = tm_row

            if len(batch_rows) >= batch_size:
                created_in_batch, updated_in_batch, skipped_duplicate_in_batch = _flush_tm_batch(
                    db=db,
                    batch_rows=list(batch_rows.values()),
                    collection_id=collection_id,
                    creator_id=creator_id,
                    duplicate_policy=duplicate_policy,
                    source_language=normalized_source_language,
                    target_language=normalized_target_language,
                )
                created_rows += created_in_batch
                updated_rows += updated_in_batch
                skipped_duplicate_rows += skipped_duplicate_in_batch
                batch_rows.clear()
                _raise_if_canceled(cancel_check)

        if batch_rows:
            _raise_if_canceled(cancel_check)
            created_in_batch, updated_in_batch, skipped_duplicate_in_batch = _flush_tm_batch(
                db=db,
                batch_rows=list(batch_rows.values()),
                collection_id=collection_id,
                creator_id=creator_id,
                duplicate_policy=duplicate_policy,
                source_language=normalized_source_language,
                target_language=normalized_target_language,
            )
            created_rows += created_in_batch
            updated_rows += updated_in_batch
            skipped_duplicate_rows += skipped_duplicate_in_batch

    finally:
        conn.close()

    return TMImportSummary(
        filename=filename,
        created_rows=created_rows,
        updated_rows=updated_rows,
        skipped_empty_rows=skipped_empty_rows,
        skipped_header_rows=0,
        skipped_duplicate_rows=skipped_duplicate_rows,
    )


def _extract_sdltm_text(xml_content: str) -> str:
    """Extract plain text from SDLTM XML segment format.
    
    SDLTM stores segments as XML like:
    <Segment><Elements><Text><Value>actual text</Value></Text></Elements>...</Segment>
    """
    if not xml_content:
        return ""
    
    # Extract all text values from <Value> tags
    values = re.findall(r"<Value>(.*?)</Value>", xml_content, re.DOTALL)
    text = "".join(values)
    
    # Clean up any remaining XML entities
    text = text.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
    text = text.replace("&quot;", '"').replace("&apos;", "'")
    
    return text.strip()
