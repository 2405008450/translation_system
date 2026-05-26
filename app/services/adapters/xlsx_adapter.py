from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any, Iterable, List, Optional
from zipfile import BadZipFile, ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.adapters.base import FormatAdapter
from app.services.adapters.exceptions import ParseError
from app.services.adapters.models import (
    BlockNode,
    DocumentAST,
    NodeType,
    ParseResult,
    TMEntry,
    TMImportResult,
)
from app.services.adapters.segment_extractor import extract_segments
from app.services.document_workspace import normalize_document_parse_options


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
CORE_NS = "http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
DC_NS = "http://purl.org/dc/elements/1.1/"
DCTERMS_NS = "http://purl.org/dc/terms/"
EXTENDED_PROPS_NS = "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
NS = {
    "main": MAIN_NS,
    "a": DRAWING_NS,
    "cp": CORE_NS,
    "dc": DC_NS,
    "dcterms": DCTERMS_NS,
    "ep": EXTENDED_PROPS_NS,
}


@dataclass
class ColumnMapping:
    source_column: int = 0
    target_column: int = 1
    skip_header: bool = True


class XlsxAdapter(FormatAdapter):
    def __init__(self, column_mapping: Optional[ColumnMapping] = None):
        super().__init__()
        self.column_mapping = column_mapping or ColumnMapping()

    def supported_extensions(self) -> List[str]:
        return [".xlsx"]

    def parse(self, raw_bytes: bytes) -> ParseResult:
        return self._parse_task(raw_bytes, options=None)

    def parse_with_options(
        self,
        raw_bytes: bytes,
        filename: str = "<unknown>",
        options: Optional[dict] = None,
    ) -> ParseResult:
        self.validate_file_size(raw_bytes, filename)
        return self._parse_task(raw_bytes, options=options)

    def parse_for_tm(self, raw_bytes: bytes) -> TMImportResult:
        result = self._parse_tm(raw_bytes)
        return result.metadata.get("tm_import_result", TMImportResult([], 0, 0))

    def _parse_task(self, raw_bytes: bytes, options: Optional[dict]) -> ParseResult:
        if not raw_bytes:
            return ParseResult(
                ast=DocumentAST(nodes=[], source_format=".xlsx"),
                segments=[],
                metadata={},
            )

        parse_options = normalize_document_parse_options(options)
        nodes: list[BlockNode] = []
        nodes.extend(self._parse_cell_nodes(raw_bytes, parse_options))

        if parse_options["xlsx_translate_comments"]:
            nodes.extend(self._parse_comment_nodes(raw_bytes))
        if parse_options["xlsx_translate_drawing_text"]:
            nodes.extend(self._parse_drawing_text_nodes(raw_bytes))
        if parse_options["xlsx_translate_sheet_names"]:
            nodes.extend(self._parse_sheet_name_nodes(raw_bytes, parse_options))
        if parse_options["xlsx_translate_document_properties"]:
            nodes.extend(self._parse_document_property_nodes(raw_bytes))

        ast = DocumentAST(nodes=nodes, source_format=".xlsx")
        return ParseResult(
            ast=ast,
            segments=extract_segments(ast),
            metadata={"node_count": len(nodes)},
        )

    def _parse_cell_nodes(self, raw_bytes: bytes, options: dict[str, Any]) -> list[BlockNode]:
        try:
            workbook = load_workbook(BytesIO(raw_bytes), read_only=False, data_only=False)
        except Exception as exc:
            raise ParseError(filename="<unknown>", reason=f"无法解析 XLSX 文件: {exc}") from exc

        skipped_colors = {
            str(color).strip().upper().lstrip("#")
            for color in options.get("xlsx_skip_fill_colors", [])
            if str(color).strip()
        }
        include_hidden = bool(options["xlsx_translate_hidden_content"])
        nodes: list[BlockNode] = []

        try:
            for sheet_index, worksheet in enumerate(workbook.worksheets):
                if not include_hidden and worksheet.sheet_state != "visible":
                    continue

                for row in worksheet.iter_rows():
                    for cell in row:
                        if not include_hidden and self._is_hidden_cell(worksheet, cell.row, cell.column):
                            continue
                        if skipped_colors and self._cell_fill_color(cell) in skipped_colors:
                            continue

                        text = self._cell_text(cell, options)
                        if not text:
                            continue

                        nodes.append(
                            BlockNode(
                                node_type=NodeType.TABLE_CELL,
                                text_content=text,
                                metadata={
                                    "item_type": "cell",
                                    "sheet_index": sheet_index,
                                    "sheet_name": worksheet.title,
                                    "table_index": sheet_index,
                                    "row": cell.row - 1,
                                    "col": cell.column - 1,
                                    "cell_ref": cell.coordinate,
                                },
                            )
                        )
        finally:
            workbook.close()

        return nodes

    def _is_hidden_cell(self, worksheet: Any, row_index: int, column_index: int) -> bool:
        row_dimension = worksheet.row_dimensions.get(row_index)
        if row_dimension is not None and row_dimension.hidden:
            return True
        column_letter = get_column_letter(column_index)
        column_dimension = worksheet.column_dimensions.get(column_letter)
        return bool(column_dimension is not None and column_dimension.hidden)

    def _cell_text(self, cell: Any, options: dict[str, Any]) -> str:
        value = cell.value
        if value is None:
            return ""
        if cell.data_type == "f":
            return str(value).strip() if options["xlsx_translate_formula_cells"] else ""
        if isinstance(value, bool):
            return str(value).strip() if options["xlsx_translate_boolean_cells"] else ""
        if getattr(cell, "is_date", False) or isinstance(value, (datetime, date, time)):
            return str(value).strip() if options["xlsx_translate_date_cells"] else ""
        if isinstance(value, (int, float, Decimal)):
            return str(value).strip() if options["xlsx_translate_numeric_cells"] else ""

        text = str(value).strip()
        if not text:
            return ""
        if self._is_numeric(text) and not options["xlsx_translate_numeric_cells"]:
            return ""
        return text

    def _cell_fill_color(self, cell: Any) -> str | None:
        fill = getattr(cell, "fill", None)
        if fill is None or not getattr(fill, "fill_type", None):
            return None
        color = getattr(fill, "fgColor", None)
        if color is None or getattr(color, "type", None) != "rgb":
            return None
        rgb = getattr(color, "rgb", None)
        if not rgb:
            return None
        rgb = str(rgb).upper()
        if len(rgb) >= 6:
            return rgb[-6:]
        return None

    def _parse_comment_nodes(self, raw_bytes: bytes) -> list[BlockNode]:
        nodes: list[BlockNode] = []
        for part_name, root in self._iter_xml_parts(raw_bytes, "xl/comments"):
            for comment_index, comment in enumerate(root.findall(".//main:comment", NS)):
                text = self._collect_text(comment.findall(".//main:t", NS))
                if not text:
                    continue
                nodes.append(
                    BlockNode(
                        node_type=NodeType.NOTE,
                        text_content=text,
                        metadata={
                            "item_type": "comment",
                            "part_name": part_name,
                            "comment_index": comment_index,
                            "cell_ref": comment.get("ref", ""),
                        },
                    )
                )
        return nodes

    def _parse_drawing_text_nodes(self, raw_bytes: bytes) -> list[BlockNode]:
        nodes: list[BlockNode] = []
        for part_name, root in self._iter_xml_parts(raw_bytes, "xl/drawings/drawing"):
            for paragraph_index, paragraph in enumerate(root.findall(".//a:p", NS)):
                text = self._collect_text(paragraph.findall(".//a:t", NS))
                if not text:
                    continue
                nodes.append(
                    BlockNode(
                        node_type=NodeType.PARAGRAPH,
                        text_content=text,
                        metadata={
                            "item_type": "drawing_text",
                            "part_name": part_name,
                            "paragraph_index": paragraph_index,
                        },
                    )
                )
        return nodes

    def _parse_sheet_name_nodes(self, raw_bytes: bytes, options: dict[str, Any]) -> list[BlockNode]:
        workbook_root = self._read_xml(raw_bytes, "xl/workbook.xml")
        if workbook_root is None:
            return []

        include_hidden = bool(options["xlsx_translate_hidden_content"])
        nodes: list[BlockNode] = []
        for sheet_index, sheet in enumerate(workbook_root.findall(".//main:sheets/main:sheet", NS)):
            if not include_hidden and sheet.get("state", "visible") != "visible":
                continue
            name = (sheet.get("name") or "").strip()
            if not name:
                continue
            nodes.append(
                BlockNode(
                    node_type=NodeType.HEADING,
                    text_content=name,
                    metadata={
                        "item_type": "sheet_name",
                        "part_name": "xl/workbook.xml",
                        "sheet_index": sheet_index,
                    },
                )
            )
        return nodes

    def _parse_document_property_nodes(self, raw_bytes: bytes) -> list[BlockNode]:
        nodes: list[BlockNode] = []
        for part_name in ("docProps/core.xml", "docProps/app.xml"):
            root = self._read_xml(raw_bytes, part_name)
            if root is None:
                continue
            for prop_index, element in enumerate(list(root)):
                text = (element.text or "").strip()
                if not text:
                    continue
                nodes.append(
                    BlockNode(
                        node_type=NodeType.PARAGRAPH,
                        text_content=text,
                        metadata={
                            "item_type": "document_property",
                            "part_name": part_name,
                            "property_index": prop_index,
                            "property_name": _local_name(element.tag),
                        },
                    )
                )
        return nodes

    def _parse_tm(self, raw_bytes: bytes) -> ParseResult:
        if not raw_bytes:
            return self._empty_tm_result()

        try:
            workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise ParseError(filename="<unknown>", reason=f"无法解析 XLSX 文件: {exc}") from exc

        entries: List[TMEntry] = []
        total_rows = 0
        skipped_rows = 0

        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_entries, sheet_total, sheet_skipped = self._parse_sheet_for_tm(sheet, sheet_name)
                entries.extend(sheet_entries)
                total_rows += sheet_total
                skipped_rows += sheet_skipped
        finally:
            workbook.close()

        tm_result = TMImportResult(entries=entries, skipped_rows=skipped_rows, total_rows=total_rows)
        return ParseResult(
            ast=DocumentAST(nodes=[], source_format=".xlsx"),
            segments=[],
            metadata={"tm_import_result": tm_result},
        )

    def _parse_sheet_for_tm(self, sheet: Any, sheet_name: str) -> tuple[List[TMEntry], int, int]:
        entries: List[TMEntry] = []
        total_rows = 0
        skipped_rows = 0
        rows = list(sheet.iter_rows(values_only=True))
        start_row = 1 if self.column_mapping.skip_header else 0

        for row_idx, row in enumerate(rows[start_row:], start=start_row):
            total_rows += 1
            source_text = self._get_cell_value(row, self.column_mapping.source_column)
            target_text = self._get_cell_value(row, self.column_mapping.target_column)
            if not source_text or not source_text.strip():
                skipped_rows += 1
                continue
            entries.append(
                TMEntry(
                    source_text=source_text.strip(),
                    target_text=(target_text or "").strip(),
                    metadata={"sheet": sheet_name, "row": row_idx + 1},
                )
            )

        return entries, total_rows, skipped_rows

    def _get_cell_value(self, row: tuple, column_index: int) -> Optional[str]:
        if column_index >= len(row):
            return None
        value = row[column_index]
        return None if value is None else str(value)

    def _empty_tm_result(self) -> ParseResult:
        return ParseResult(
            ast=DocumentAST(nodes=[], source_format=".xlsx"),
            segments=[],
            metadata={"tm_import_result": TMImportResult(entries=[], skipped_rows=0, total_rows=0)},
        )

    def _iter_xml_parts(self, raw_bytes: bytes, prefix: str) -> Iterable[tuple[str, ET.Element]]:
        try:
            with ZipFile(BytesIO(raw_bytes)) as archive:
                for name in sorted(archive.namelist()):
                    if not name.startswith(prefix) or not name.endswith(".xml"):
                        continue
                    yield name, ET.fromstring(archive.read(name))
        except (BadZipFile, ET.ParseError) as exc:
            raise ParseError(filename="<unknown>", reason=f"无法解析 XLSX XML 内容: {exc}") from exc

    def _read_xml(self, raw_bytes: bytes, part_name: str) -> ET.Element | None:
        try:
            with ZipFile(BytesIO(raw_bytes)) as archive:
                try:
                    return ET.fromstring(archive.read(part_name))
                except KeyError:
                    return None
        except (BadZipFile, ET.ParseError) as exc:
            raise ParseError(filename="<unknown>", reason=f"无法解析 XLSX XML 内容: {exc}") from exc

    def _collect_text(self, elements: Iterable[ET.Element]) -> str:
        return "".join((element.text or "") for element in elements).strip()

    def _is_numeric(self, text: str) -> bool:
        try:
            float(text.replace(",", ""))
            return True
        except ValueError:
            return False


def _local_name(tag: str) -> str:
    if tag.startswith("{") and "}" in tag:
        return tag.split("}", 1)[1]
    return tag
