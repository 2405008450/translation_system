from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from uuid import UUID
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.models import TermEntry
from app.services.language_pairs import require_language_pair
from app.services.normalizer import normalize_match_text, normalize_text


CSV_EXTENSIONS = {".csv"}
TMX_EXTENSIONS = {".tmx"}
XLS_EXTENSIONS = {".xls"}
XLSX_EXTENSIONS = {".xlsx"}
TERM_IMPORT_EXTENSIONS = CSV_EXTENSIONS | TMX_EXTENSIONS | XLS_EXTENSIONS | XLSX_EXTENSIONS
TERM_STATUS_LOOKUP_CHUNK_SIZE = 10000
TERM_PREVIEW_MAX_SCAN_ROWS = 5000
HEADER_ALIASES = {
    ("source", "target"),
    ("source_term", "target_term"),
    ("source_text", "target_text"),
    ("term", "translation"),
    ("术语", "译文"),
    ("原文", "译文"),
    ("中文", "英文"),
}


@dataclass
class TermImportSummary:
    filename: str
    created_rows: int
    updated_rows: int
    skipped_empty_rows: int
    skipped_header_rows: int
    skipped_duplicate_rows: int = 0

    @property
    def imported_rows(self) -> int:
        return self.created_rows + self.updated_rows


@dataclass
class TermImportPreviewRow:
    row_index: int
    source_text: str
    target_text: str
    status: str
    message: str


@dataclass
class TermImportPreview:
    filename: str
    rows: list[TermImportPreviewRow]
    total_rows: int
    valid_rows: int
    create_rows: int
    update_rows: int
    duplicate_rows: int
    skipped_empty_rows: int
    skipped_header_rows: int
    preview_limit: int
    scanned_rows: int
    truncated: bool


def import_terms_from_xlsx_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    batch_size: int = 5000,
    term_base_id: UUID | None = None,
    creator_id: UUID | None = None,
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
) -> TermImportSummary:
    extension = f".{filename.rsplit('.', 1)[-1].lower()}" if "." in filename else ""
    if extension in CSV_EXTENSIONS:
        return import_terms_from_csv_upload(
            db=db,
            raw_bytes=raw_bytes,
            filename=filename,
            source_language=source_language,
            target_language=target_language,
            batch_size=batch_size,
            term_base_id=term_base_id,
            creator_id=creator_id,
            skip_duplicate_row_indexes=skip_duplicate_row_indexes,
            skip_header=skip_header,
        )
    if extension in TMX_EXTENSIONS:
        return import_terms_from_tmx_upload(
            db=db,
            raw_bytes=raw_bytes,
            filename=filename,
            source_language=source_language,
            target_language=target_language,
            batch_size=batch_size,
            term_base_id=term_base_id,
            creator_id=creator_id,
            skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        )
    if extension in XLS_EXTENSIONS:
        return import_terms_from_xls_upload(
            db=db,
            raw_bytes=raw_bytes,
            filename=filename,
            source_language=source_language,
            target_language=target_language,
            batch_size=batch_size,
            term_base_id=term_base_id,
            creator_id=creator_id,
            skip_duplicate_row_indexes=skip_duplicate_row_indexes,
            skip_header=skip_header,
        )

    workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    return _import_workbook(
        db=db,
        workbook=workbook,
        filename=filename,
        batch_size=batch_size,
        term_base_id=term_base_id,
        creator_id=creator_id,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        skip_header=skip_header,
        source_language=source_language,
        target_language=target_language,
    )


def import_terms_from_csv_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    batch_size: int = 5000,
    term_base_id: UUID | None = None,
    creator_id: UUID | None = None,
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
) -> TermImportSummary:
    return _import_text_rows(
        db=db,
        rows=_iter_csv_rows(raw_bytes),
        filename=filename,
        batch_size=batch_size,
        term_base_id=term_base_id,
        creator_id=creator_id,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        skip_header=skip_header,
        source_language=source_language,
        target_language=target_language,
    )


def import_terms_from_tmx_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    batch_size: int = 5000,
    term_base_id: UUID | None = None,
    creator_id: UUID | None = None,
    skip_duplicate_row_indexes: set[int] | None = None,
) -> TermImportSummary:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    return _import_text_rows(
        db=db,
        rows=_iter_tmx_rows(
            raw_bytes,
            normalized_source_language,
            normalized_target_language,
        ),
        filename=filename,
        batch_size=batch_size,
        term_base_id=term_base_id,
        creator_id=creator_id,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        source_language=normalized_source_language,
        target_language=normalized_target_language,
    )


def import_terms_from_xls_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    batch_size: int = 5000,
    term_base_id: UUID | None = None,
    creator_id: UUID | None = None,
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
) -> TermImportSummary:
    return _import_text_rows(
        db=db,
        rows=_iter_xls_rows(raw_bytes),
        filename=filename,
        batch_size=batch_size,
        term_base_id=term_base_id,
        creator_id=creator_id,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        skip_header=skip_header,
        source_language=source_language,
        target_language=target_language,
    )


def preview_terms_from_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    term_base_id: UUID | None,
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = TERM_PREVIEW_MAX_SCAN_ROWS,
) -> TermImportPreview:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    rows = _iter_upload_rows(
        raw_bytes,
        filename,
        normalized_source_language,
        normalized_target_language,
    )
    preview_rows: list[TermImportPreviewRow] = []
    source_normalized_values: set[str] = set()
    source_texts: set[str] = set()
    final_candidates: dict[str, dict] = {}
    skipped_empty_rows = 0
    skipped_header_rows = 0
    total_rows = 0
    duplicate_rows = 0
    safe_max_scan_rows = max(1, int(max_scan_rows or TERM_PREVIEW_MAX_SCAN_ROWS))
    truncated = False

    for row_index, row in enumerate(rows, start=1):
        if row_index > safe_max_scan_rows:
            truncated = True
            break
        total_rows += 1
        source_text = normalize_text(_cell_to_text(row, 0))
        target_text = normalize_text(_cell_to_text(row, 1))

        if _should_skip_header_row(filename, row_index, source_text, target_text, skip_header):
            skipped_header_rows += 1
            _append_preview_row(
                preview_rows,
                preview_limit,
                TermImportPreviewRow(
                    row_index=row_index,
                    source_text=source_text,
                    target_text=target_text,
                    status="header",
                    message="表头行，导入时会跳过。",
                ),
            )
            continue

        if not source_text or not target_text:
            skipped_empty_rows += 1
            _append_preview_row(
                preview_rows,
                preview_limit,
                TermImportPreviewRow(
                    row_index=row_index,
                    source_text=source_text,
                    target_text=target_text,
                    status="empty",
                    message="源术语或目标术语为空，导入时会跳过。",
                ),
            )
            continue

        term_row = _build_term_row(
            source_text=source_text,
            target_text=target_text,
            term_base_id=term_base_id,
            source_language=normalized_source_language,
            target_language=normalized_target_language,
        )
        source_normalized = term_row["source_normalized"]
        status = "pending"
        message = "待导入。"
        if source_normalized in source_normalized_values or source_text in source_texts:
            duplicate_rows += 1
            status = "duplicate"
            message = "文件内重复，实际导入时以后出现的这一条为准。"
        source_normalized_values.add(source_normalized)
        source_texts.add(source_text)
        final_candidates[source_normalized] = term_row
        _append_preview_row(
            preview_rows,
            preview_limit,
            TermImportPreviewRow(
                row_index=row_index,
                source_text=source_text,
                target_text=target_text,
                status=status,
                message=message,
            ),
        )

    existing_status = _load_existing_term_status(
        db=db,
        term_base_id=term_base_id,
        source_language=normalized_source_language,
        target_language=normalized_target_language,
        source_normalized_values=list(source_normalized_values),
        source_texts=list(source_texts),
    )

    create_rows = 0
    update_rows = 0
    for candidate in final_candidates.values():
        key = candidate["source_normalized"]
        source_text = candidate["source_text"]
        if key in existing_status or source_text in existing_status:
            update_rows += 1
        else:
            create_rows += 1

    for preview_row in preview_rows:
        if preview_row.status != "pending":
            continue
        if preview_row.source_text in existing_status:
            preview_row.status = "update"
            preview_row.message = "术语库中已有相同源术语，导入时会覆盖目标术语。"
        else:
            normalized = normalize_match_text(preview_row.source_text) or normalize_text(preview_row.source_text)
            if normalized in existing_status:
                preview_row.status = "update"
                preview_row.message = "术语库中已有相同源术语，导入时会覆盖目标术语。"
            else:
                preview_row.status = "create"
                preview_row.message = "导入时会新增。"

    return TermImportPreview(
        filename=filename,
        rows=preview_rows,
        total_rows=total_rows,
        valid_rows=len(final_candidates),
        create_rows=create_rows,
        update_rows=update_rows,
        duplicate_rows=duplicate_rows,
        skipped_empty_rows=skipped_empty_rows,
        skipped_header_rows=skipped_header_rows,
        preview_limit=preview_limit,
        scanned_rows=total_rows,
        truncated=truncated,
    )


def preview_terms_from_xlsx_path(
    db: Session,
    xlsx_path: str | Path,
    filename: str,
    source_language: str,
    target_language: str,
    term_base_id: UUID | None,
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = TERM_PREVIEW_MAX_SCAN_ROWS,
) -> TermImportPreview:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    workbook = load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    try:
        return _preview_terms_from_rows(
            db=db,
            rows=workbook.active.iter_rows(values_only=True),
            filename=filename,
            term_base_id=term_base_id,
            source_language=normalized_source_language,
            target_language=normalized_target_language,
            preview_limit=preview_limit,
            skip_header=skip_header,
            max_scan_rows=max_scan_rows,
        )
    finally:
        workbook.close()


def _preview_terms_from_rows(
    db: Session,
    rows,
    filename: str,
    term_base_id: UUID | None,
    source_language: str,
    target_language: str,
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = TERM_PREVIEW_MAX_SCAN_ROWS,
) -> TermImportPreview:
    preview_rows: list[TermImportPreviewRow] = []
    source_normalized_values: set[str] = set()
    source_texts: set[str] = set()
    final_candidates: dict[str, dict] = {}
    skipped_empty_rows = 0
    skipped_header_rows = 0
    total_rows = 0
    duplicate_rows = 0
    safe_max_scan_rows = max(1, int(max_scan_rows or TERM_PREVIEW_MAX_SCAN_ROWS))
    truncated = False

    for row_index, row in enumerate(rows, start=1):
        if row_index > safe_max_scan_rows:
            truncated = True
            break
        total_rows += 1
        source_text = normalize_text(_cell_to_text(row, 0))
        target_text = normalize_text(_cell_to_text(row, 1))

        if _should_skip_header_row(filename, row_index, source_text, target_text, skip_header):
            skipped_header_rows += 1
            _append_preview_row(
                preview_rows,
                preview_limit,
                TermImportPreviewRow(row_index, source_text, target_text, "header", "表头行，导入时会跳过。"),
            )
            continue

        if not source_text or not target_text:
            skipped_empty_rows += 1
            _append_preview_row(
                preview_rows,
                preview_limit,
                TermImportPreviewRow(row_index, source_text, target_text, "empty", "源术语或目标术语为空，导入时会跳过。"),
            )
            continue

        term_row = _build_term_row(
            source_text=source_text,
            target_text=target_text,
            term_base_id=term_base_id,
            source_language=source_language,
            target_language=target_language,
        )
        source_normalized = term_row["source_normalized"]
        status = "pending"
        message = "待导入。"
        if source_normalized in source_normalized_values or source_text in source_texts:
            duplicate_rows += 1
            status = "duplicate"
            message = "文件内重复，实际导入时以后出现的这一条为准。"
        source_normalized_values.add(source_normalized)
        source_texts.add(source_text)
        final_candidates[source_normalized] = term_row
        _append_preview_row(
            preview_rows,
            preview_limit,
            TermImportPreviewRow(row_index, source_text, target_text, status, message),
        )

    existing_status = _load_existing_term_status(
        db=db,
        term_base_id=term_base_id,
        source_language=source_language,
        target_language=target_language,
        source_normalized_values=list(source_normalized_values),
        source_texts=list(source_texts),
    )

    create_rows = 0
    update_rows = 0
    for candidate in final_candidates.values():
        key = candidate["source_normalized"]
        source_text = candidate["source_text"]
        if key in existing_status or source_text in existing_status:
            update_rows += 1
        else:
            create_rows += 1

    for preview_row in preview_rows:
        if preview_row.status != "pending":
            continue
        normalized = normalize_match_text(preview_row.source_text) or normalize_text(preview_row.source_text)
        if normalized in existing_status or preview_row.source_text in existing_status:
            preview_row.status = "update"
            preview_row.message = "术语库中已有相同源术语，导入时会覆盖目标术语。"
        else:
            preview_row.status = "create"
            preview_row.message = "导入时会新增。"

    return TermImportPreview(
        filename=filename,
        rows=preview_rows,
        total_rows=total_rows,
        valid_rows=len(final_candidates),
        create_rows=create_rows,
        update_rows=update_rows,
        duplicate_rows=duplicate_rows,
        skipped_empty_rows=skipped_empty_rows,
        skipped_header_rows=skipped_header_rows,
        preview_limit=preview_limit,
        scanned_rows=total_rows,
        truncated=truncated,
    )


def import_terms_from_xlsx_path(
    db: Session,
    xlsx_path: str | Path,
    source_language: str,
    target_language: str,
    filename: str | None = None,
    batch_size: int = 5000,
    term_base_id: UUID | None = None,
    creator_id: UUID | None = None,
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
) -> TermImportSummary:
    workbook = load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    return _import_workbook(
        db=db,
        workbook=workbook,
        filename=filename or Path(xlsx_path).name,
        batch_size=batch_size,
        term_base_id=term_base_id,
        creator_id=creator_id,
        skip_duplicate_row_indexes=skip_duplicate_row_indexes,
        skip_header=skip_header,
        source_language=source_language,
        target_language=target_language,
    )


def _iter_upload_rows(raw_bytes: bytes, filename: str, source_language: str, target_language: str):
    extension = f".{filename.rsplit('.', 1)[-1].lower()}" if "." in filename else ""
    if extension in CSV_EXTENSIONS:
        return _iter_csv_rows(raw_bytes)
    if extension in TMX_EXTENSIONS:
        return _iter_tmx_rows(raw_bytes, source_language, target_language)
    if extension in XLS_EXTENSIONS:
        return _iter_xls_rows(raw_bytes)
    if extension not in XLSX_EXTENSIONS:
        raise RuntimeError("仅支持上传 .tmx、.xls、.xlsx 或 .csv 文件。")

    workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)

    def iter_workbook_rows():
        try:
            yield from workbook.active.iter_rows(values_only=True)
        finally:
            workbook.close()

    return iter_workbook_rows()


def _import_workbook(
    db: Session,
    workbook,
    filename: str,
    batch_size: int,
    source_language: str,
    target_language: str,
    term_base_id: UUID | None = None,
    creator_id: UUID | None = None,
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
) -> TermImportSummary:
    worksheet = workbook.active
    try:
        return _import_text_rows(
            db=db,
            rows=worksheet.iter_rows(values_only=True),
            filename=filename,
            batch_size=batch_size,
            term_base_id=term_base_id,
            creator_id=creator_id,
            skip_duplicate_row_indexes=skip_duplicate_row_indexes,
            skip_header=skip_header,
            source_language=source_language,
            target_language=target_language,
        )
    finally:
        workbook.close()


def _import_text_rows(
    db: Session,
    rows,
    filename: str,
    batch_size: int,
    source_language: str,
    target_language: str,
    term_base_id: UUID | None = None,
    creator_id: UUID | None = None,
    skip_duplicate_row_indexes: set[int] | None = None,
    skip_header: bool = False,
) -> TermImportSummary:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    batch_rows: dict[str, dict] = {}
    created_rows = 0
    updated_rows = 0
    skipped_empty_rows = 0
    skipped_header_rows = 0
    skipped_duplicate_rows = 0
    skipped_row_indexes = skip_duplicate_row_indexes or set()

    for row_index, row in enumerate(rows, start=1):
        source_text = normalize_text(_cell_to_text(row, 0))
        target_text = normalize_text(_cell_to_text(row, 1))

        if _should_skip_header_row(filename, row_index, source_text, target_text, skip_header):
            skipped_header_rows += 1
            continue

        if not source_text or not target_text:
            skipped_empty_rows += 1
            continue

        if row_index in skipped_row_indexes:
            skipped_duplicate_rows += 1
            continue

        term_row = _build_term_row(
            source_text=source_text,
            target_text=target_text,
            term_base_id=term_base_id,
            source_language=normalized_source_language,
            target_language=normalized_target_language,
        )
        batch_rows[term_row["source_normalized"]] = term_row

        if len(batch_rows) >= batch_size:
            created_in_batch, updated_in_batch = _flush_term_batch(
                db=db,
                batch_rows=list(batch_rows.values()),
                term_base_id=term_base_id,
                source_language=normalized_source_language,
                target_language=normalized_target_language,
            )
            created_rows += created_in_batch
            updated_rows += updated_in_batch
            batch_rows.clear()

    if batch_rows:
        created_in_batch, updated_in_batch = _flush_term_batch(
            db=db,
            batch_rows=list(batch_rows.values()),
            term_base_id=term_base_id,
            source_language=normalized_source_language,
            target_language=normalized_target_language,
        )
        created_rows += created_in_batch
        updated_rows += updated_in_batch

    return TermImportSummary(
        filename=filename,
        created_rows=created_rows,
        updated_rows=updated_rows,
        skipped_empty_rows=skipped_empty_rows,
        skipped_header_rows=skipped_header_rows,
        skipped_duplicate_rows=skipped_duplicate_rows,
    )


def _append_preview_row(
    rows: list[TermImportPreviewRow],
    preview_limit: int,
    row: TermImportPreviewRow,
) -> None:
    if len(rows) < preview_limit:
        rows.append(row)


def _load_existing_term_status(
    db: Session,
    term_base_id: UUID | None,
    source_language: str,
    target_language: str,
    source_normalized_values: list[str],
    source_texts: list[str],
) -> set[str]:
    if term_base_id is None or (not source_normalized_values and not source_texts):
        return set()

    keys: set[str] = set()
    source_normalized_values = list(dict.fromkeys(source_normalized_values))
    source_texts = list(dict.fromkeys(source_texts))
    max_lookup_count = max(len(source_normalized_values), len(source_texts))

    for offset in range(0, max_lookup_count, TERM_STATUS_LOOKUP_CHUNK_SIZE):
        normalized_chunk = source_normalized_values[offset : offset + TERM_STATUS_LOOKUP_CHUNK_SIZE]
        source_text_chunk = source_texts[offset : offset + TERM_STATUS_LOOKUP_CHUNK_SIZE]
        lookup_conditions = []
        if normalized_chunk:
            lookup_conditions.append(TermEntry.source_normalized.in_(normalized_chunk))
        if source_text_chunk:
            lookup_conditions.append(TermEntry.source_text.in_(source_text_chunk))

        existing_rows = (
            db.query(TermEntry.source_normalized, TermEntry.source_text)
            .filter(
                TermEntry.term_base_id == term_base_id,
                TermEntry.source_language == source_language,
                TermEntry.target_language == target_language,
                or_(*lookup_conditions),
            )
            .all()
        )
        for source_normalized, source_text in existing_rows:
            if source_normalized:
                keys.add(source_normalized)
            if source_text:
                keys.add(source_text)
    return keys


def _iter_csv_rows(raw_bytes: bytes):
    text = _decode_csv_bytes(raw_bytes)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(text.splitlines(), dialect)
    for row in reader:
        yield tuple(row)


def _iter_xls_rows(raw_bytes: bytes):
    try:
        import xlrd
    except ModuleNotFoundError as exc:
        raise RuntimeError("导入 .xls 文件需要安装 xlrd 依赖。") from exc

    workbook = xlrd.open_workbook(file_contents=raw_bytes)
    sheet = workbook.sheet_by_index(0)
    for row_index in range(sheet.nrows):
        yield tuple(sheet.cell_value(row_index, column_index) for column_index in range(sheet.ncols))


def _iter_tmx_rows(raw_bytes: bytes, source_language: str, target_language: str):
    root = ET.fromstring(raw_bytes)
    for translation_unit in root.findall(".//{*}tu"):
        language_segments: list[tuple[str, str]] = []
        for tuv in translation_unit.findall("{*}tuv"):
            language = (
                tuv.attrib.get("{http://www.w3.org/XML/1998/namespace}lang")
                or tuv.attrib.get("lang")
                or tuv.attrib.get("xml:lang")
                or ""
            )
            segment = tuv.find("{*}seg")
            if segment is None:
                continue
            text = normalize_text("".join(segment.itertext()))
            if text:
                language_segments.append((language, text))

        source_text = _find_tmx_language_text(language_segments, source_language)
        target_text = _find_tmx_language_text(language_segments, target_language)
        if (not source_text or not target_text) and len(language_segments) >= 2:
            source_text = source_text or language_segments[0][1]
            target_text = target_text or language_segments[1][1]
        yield (source_text, target_text)


def _decode_csv_bytes(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8", errors="replace")


def _find_tmx_language_text(language_segments: list[tuple[str, str]], language: str) -> str:
    normalized_language = _normalize_language_tag(language)
    for candidate_language, text in language_segments:
        if _normalize_language_tag(candidate_language) == normalized_language:
            return text
    primary_language = normalized_language.split("-", 1)[0]
    for candidate_language, text in language_segments:
        if _normalize_language_tag(candidate_language).split("-", 1)[0] == primary_language:
            return text
    return ""


def _build_term_row(
    source_text: str,
    target_text: str,
    source_language: str,
    target_language: str,
    term_base_id: UUID | None = None,
    creator_id: UUID | None = None,
) -> dict:
    return {
        "term_base_id": term_base_id,
        "source_text": source_text,
        "target_text": target_text,
        "source_normalized": normalize_match_text(source_text) or normalize_text(source_text),
        "source_language": source_language,
        "target_language": target_language,
        "creator_id": creator_id,
        "last_modified_by_id": creator_id,
    }


def _normalize_language_tag(language: str) -> str:
    return (language or "").strip().lower().replace("_", "-")


def _flush_term_batch(
    db: Session,
    batch_rows: list[dict],
    source_language: str,
    target_language: str,
    term_base_id: UUID | None = None,
) -> tuple[int, int]:
    if not batch_rows:
        return 0, 0

    source_normalized_values = [row["source_normalized"] for row in batch_rows]
    source_texts = [row["source_text"] for row in batch_rows]
    existing_query = (
        db.query(TermEntry)
        .filter(
            or_(
                TermEntry.source_normalized.in_(source_normalized_values),
                TermEntry.source_text.in_(source_texts),
            ),
            TermEntry.source_language == source_language,
            TermEntry.target_language == target_language,
        )
    )
    if term_base_id is not None:
        existing_query = existing_query.filter(TermEntry.term_base_id == term_base_id)
    existing_rows = existing_query.all()

    existing_by_normalized: dict[str, TermEntry] = {}
    existing_by_source_text: dict[str, TermEntry] = {}
    for existing in existing_rows:
        if existing.source_normalized:
            existing_by_normalized.setdefault(existing.source_normalized, existing)
        existing_by_source_text.setdefault(existing.source_text, existing)

    created_rows = 0
    updated_rows = 0
    for row in batch_rows:
        existing = existing_by_normalized.get(row["source_normalized"]) or existing_by_source_text.get(
            row["source_text"]
        )
        if existing is None:
            db.add(TermEntry(**row))
            created_rows += 1
            continue

        if _term_entry_matches_import_row(existing, row):
            continue

        existing.source_text = row["source_text"]
        existing.target_text = row["target_text"]
        existing.source_normalized = row["source_normalized"]
        existing.source_language = row["source_language"]
        existing.target_language = row["target_language"]
        if existing.creator_id is None and row.get("creator_id"):
            existing.creator_id = row["creator_id"]
        if row.get("last_modified_by_id"):
            existing.last_modified_by_id = row["last_modified_by_id"]
        updated_rows += 1

    db.commit()
    db.expunge_all()
    return created_rows, updated_rows


def _term_entry_matches_import_row(existing: TermEntry, row: dict) -> bool:
    return (
        normalize_text(existing.source_text or "") == row["source_text"]
        and normalize_text(existing.target_text or "") == row["target_text"]
        and existing.term_base_id == row["term_base_id"]
        and (existing.source_language or "") == row["source_language"]
        and (existing.target_language or "") == row["target_language"]
    )


def _cell_to_text(row: tuple, index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value)


def _looks_like_header(source_text: str, target_text: str) -> bool:
    if not source_text or not target_text:
        return False

    header_key = (source_text.lower(), target_text.lower())
    return header_key in HEADER_ALIASES


def _is_table_import(filename: str) -> bool:
    extension = f".{filename.rsplit('.', 1)[-1].lower()}" if "." in filename else ""
    return extension in CSV_EXTENSIONS or extension in XLS_EXTENSIONS or extension in XLSX_EXTENSIONS


def _should_skip_header_row(
    filename: str,
    row_index: int,
    source_text: str,
    target_text: str,
    skip_header: bool,
) -> bool:
    if row_index != 1:
        return False
    return _looks_like_header(source_text, target_text) or (skip_header and _is_table_import(filename))
