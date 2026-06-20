from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.models import GlossaryEntry
from app.services.language_pairs import require_language_pair
from app.services.normalizer import normalize_match_text, normalize_text


XLSX_EXTENSIONS = {".xlsx"}
GLOSSARY_STATUS_LOOKUP_CHUNK_SIZE = 10000
GLOSSARY_PREVIEW_MAX_SCAN_ROWS = 1000
SOURCE_HEADER_ALIASES = {"source", "source_text", "source_term", "term", "原文", "词汇", "词条", "术语"}
TARGET_HEADER_ALIASES = {"target", "target_text", "target_term", "translation", "译文", "翻译"}
NOTE_HEADER_ALIASES = {"note", "notes", "comment", "comments", "context", "remark", "remarks", "备注", "说明", "补充解释"}


@dataclass
class GlossaryImportSummary:
    filename: str
    created_rows: int
    updated_rows: int
    skipped_empty_rows: int
    skipped_header_rows: int

    @property
    def imported_rows(self) -> int:
        return self.created_rows + self.updated_rows


@dataclass
class GlossaryImportPreviewRow:
    row_index: int
    source_text: str
    target_text: str
    note: str
    status: str
    message: str


@dataclass
class GlossaryImportPreview:
    filename: str
    rows: list[GlossaryImportPreviewRow]
    total_rows: int
    valid_rows: int
    create_rows: int
    update_rows: int
    duplicate_rows: int
    skipped_empty_rows: int
    skipped_header_rows: int
    preview_limit: int
    scanned_rows: int = 0
    truncated: bool = False


def import_glossary_from_xlsx_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    *,
    glossary_base_id: UUID,
    creator_id: UUID | None = None,
    batch_size: int = 5000,
    skip_header: bool = False,
) -> GlossaryImportSummary:
    workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    return _import_workbook(
        db=db,
        workbook=workbook,
        filename=filename,
        batch_size=batch_size,
        glossary_base_id=glossary_base_id,
        source_language=source_language,
        target_language=target_language,
        creator_id=creator_id,
        skip_header=skip_header,
    )


def preview_glossary_from_xlsx_upload(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    target_language: str,
    *,
    glossary_base_id: UUID | None,
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = GLOSSARY_PREVIEW_MAX_SCAN_ROWS,
) -> GlossaryImportPreview:
    workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    return _preview_workbook(
        db=db,
        workbook=workbook,
        filename=filename,
        glossary_base_id=glossary_base_id,
        source_language=source_language,
        target_language=target_language,
        preview_limit=preview_limit,
        skip_header=skip_header,
        max_scan_rows=max_scan_rows,
    )


def preview_glossary_from_xlsx_path(
    db: Session,
    xlsx_path: str | Path,
    filename: str,
    source_language: str,
    target_language: str,
    *,
    glossary_base_id: UUID | None,
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = GLOSSARY_PREVIEW_MAX_SCAN_ROWS,
) -> GlossaryImportPreview:
    workbook = load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    return _preview_workbook(
        db=db,
        workbook=workbook,
        filename=filename,
        glossary_base_id=glossary_base_id,
        source_language=source_language,
        target_language=target_language,
        preview_limit=preview_limit,
        skip_header=skip_header,
        max_scan_rows=max_scan_rows,
    )


def import_glossary_from_xlsx_path(
    db: Session,
    xlsx_path: str | Path,
    source_language: str,
    target_language: str,
    *,
    filename: str | None = None,
    glossary_base_id: UUID,
    creator_id: UUID | None = None,
    batch_size: int = 5000,
    skip_header: bool = False,
) -> GlossaryImportSummary:
    workbook = load_workbook(Path(xlsx_path), read_only=True, data_only=True)
    return _import_workbook(
        db=db,
        workbook=workbook,
        filename=filename or Path(xlsx_path).name,
        batch_size=batch_size,
        glossary_base_id=glossary_base_id,
        source_language=source_language,
        target_language=target_language,
        creator_id=creator_id,
        skip_header=skip_header,
    )


def _preview_workbook(
    db: Session,
    workbook,
    filename: str,
    glossary_base_id: UUID | None,
    source_language: str,
    target_language: str,
    preview_limit: int = 100,
    skip_header: bool = False,
    max_scan_rows: int = GLOSSARY_PREVIEW_MAX_SCAN_ROWS,
) -> GlossaryImportPreview:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    worksheet = workbook.active
    preview_rows: list[GlossaryImportPreviewRow] = []
    source_normalized_values: set[str] = set()
    source_texts: set[str] = set()
    final_candidates: dict[str, dict] = {}
    skipped_empty_rows = 0
    skipped_header_rows = 0
    total_rows = 0
    duplicate_rows = 0
    column_indexes = (0, 1, 2)
    safe_max_scan_rows = max(1, int(max_scan_rows or GLOSSARY_PREVIEW_MAX_SCAN_ROWS))
    truncated = False

    try:
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_index > safe_max_scan_rows:
                truncated = True
                break
            total_rows += 1

            if row_index == 1:
                detected = _detect_header_indexes(row)
                if detected is not None:
                    column_indexes = detected
                    source_text, target_text, note = _read_glossary_preview_cells(row, column_indexes)
                    skipped_header_rows += 1
                    _append_preview_row(
                        preview_rows,
                        preview_limit,
                        GlossaryImportPreviewRow(
                            row_index=row_index,
                            source_text=source_text,
                            target_text=target_text,
                            note=note,
                            status="header",
                            message="表头行，导入时会跳过。",
                        ),
                    )
                    continue
                if skip_header:
                    source_text, target_text, note = _read_glossary_preview_cells(row, column_indexes)
                    skipped_header_rows += 1
                    _append_preview_row(
                        preview_rows,
                        preview_limit,
                        GlossaryImportPreviewRow(
                            row_index=row_index,
                            source_text=source_text,
                            target_text=target_text,
                            note=note,
                            status="header",
                            message="已选择跳过表头，导入时会跳过此行。",
                        ),
                    )
                    continue

            source_text, target_text, note = _read_glossary_preview_cells(row, column_indexes)
            if not source_text or not target_text:
                skipped_empty_rows += 1
                _append_preview_row(
                    preview_rows,
                    preview_limit,
                    GlossaryImportPreviewRow(
                        row_index=row_index,
                        source_text=source_text,
                        target_text=target_text,
                        note=note,
                        status="empty",
                        message="原文或译文为空，导入时会跳过。",
                    ),
                )
                continue

            glossary_row = _build_glossary_row(
                source_text=source_text,
                target_text=target_text,
                note=note,
                glossary_base_id=glossary_base_id,
                source_language=normalized_source_language,
                target_language=normalized_target_language,
            )
            source_normalized = glossary_row["source_normalized"]
            status = "pending"
            message = "待导入。"
            if source_normalized in source_normalized_values or source_text in source_texts:
                duplicate_rows += 1
                status = "duplicate"
                message = "文件内重复，实际导入时以后出现的这一行为准。"
            source_normalized_values.add(source_normalized)
            source_texts.add(source_text)
            final_candidates[source_normalized] = glossary_row
            _append_preview_row(
                preview_rows,
                preview_limit,
                GlossaryImportPreviewRow(
                    row_index=row_index,
                    source_text=source_text,
                    target_text=target_text,
                    note=note,
                    status=status,
                    message=message,
                ),
            )
    finally:
        workbook.close()

    existing_status = _load_existing_glossary_status(
        db=db,
        glossary_base_id=glossary_base_id,
        source_language=normalized_source_language,
        target_language=normalized_target_language,
        source_normalized_values=list(source_normalized_values),
        source_texts=list(source_texts),
    )

    create_rows = 0
    update_rows = 0
    for candidate in final_candidates.values():
        key = candidate["source_normalized"]
        source_text = candidate["source_text"]
        if key in existing_status or source_text in existing_status:
            update_rows += 1
        else:
            create_rows += 1

    for preview_row in preview_rows:
        if preview_row.status != "pending":
            continue
        normalized = normalize_match_text(preview_row.source_text) or normalize_text(preview_row.source_text)
        if normalized in existing_status or preview_row.source_text in existing_status:
            preview_row.status = "update"
            preview_row.message = "词汇表中已有相同原文，导入时会覆盖译文。"
        else:
            preview_row.status = "create"
            preview_row.message = "导入时会新增。"

    return GlossaryImportPreview(
        filename=filename,
        rows=preview_rows,
        total_rows=total_rows,
        valid_rows=len(final_candidates),
        create_rows=create_rows,
        update_rows=update_rows,
        duplicate_rows=duplicate_rows,
        skipped_empty_rows=skipped_empty_rows,
        skipped_header_rows=skipped_header_rows,
        preview_limit=preview_limit,
        scanned_rows=total_rows,
        truncated=truncated,
    )


def _import_workbook(
    db: Session,
    workbook,
    filename: str,
    batch_size: int,
    glossary_base_id: UUID,
    source_language: str,
    target_language: str,
    creator_id: UUID | None = None,
    skip_header: bool = False,
) -> GlossaryImportSummary:
    normalized_source_language, normalized_target_language = require_language_pair(
        source_language,
        target_language,
    )
    worksheet = workbook.active
    batch_rows: dict[str, dict] = {}
    created_rows = 0
    updated_rows = 0
    skipped_empty_rows = 0
    skipped_header_rows = 0
    column_indexes = (0, 1, 2)

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_index == 1:
            detected = _detect_header_indexes(row)
            if detected is not None:
                column_indexes = detected
                skipped_header_rows += 1
                continue
            if skip_header:
                skipped_header_rows += 1
                continue

        source_index, target_index, note_index = column_indexes
        source_text = normalize_text(_cell_to_text(row, source_index))
        target_text = normalize_text(_cell_to_text(row, target_index))
        note = normalize_text(_cell_to_text(row, note_index)) if note_index >= 0 else ""

        if not source_text or not target_text:
            skipped_empty_rows += 1
            continue

        glossary_row = _build_glossary_row(
            source_text=source_text,
            target_text=target_text,
            note=note,
            glossary_base_id=glossary_base_id,
            source_language=normalized_source_language,
            target_language=normalized_target_language,
            creator_id=creator_id,
        )
        batch_rows[glossary_row["source_normalized"]] = glossary_row

        if len(batch_rows) >= batch_size:
            created_in_batch, updated_in_batch = _flush_glossary_batch(
                db=db,
                batch_rows=list(batch_rows.values()),
                glossary_base_id=glossary_base_id,
                source_language=normalized_source_language,
                target_language=normalized_target_language,
            )
            created_rows += created_in_batch
            updated_rows += updated_in_batch
            batch_rows.clear()

    if batch_rows:
        created_in_batch, updated_in_batch = _flush_glossary_batch(
            db=db,
            batch_rows=list(batch_rows.values()),
            glossary_base_id=glossary_base_id,
            source_language=normalized_source_language,
            target_language=normalized_target_language,
        )
        created_rows += created_in_batch
        updated_rows += updated_in_batch

    workbook.close()
    return GlossaryImportSummary(
        filename=filename,
        created_rows=created_rows,
        updated_rows=updated_rows,
        skipped_empty_rows=skipped_empty_rows,
        skipped_header_rows=skipped_header_rows,
    )


def _read_glossary_preview_cells(row: tuple, column_indexes: tuple[int, int, int]) -> tuple[str, str, str]:
    source_index, target_index, note_index = column_indexes
    source_text = normalize_text(_cell_to_text(row, source_index))
    target_text = normalize_text(_cell_to_text(row, target_index))
    note = normalize_text(_cell_to_text(row, note_index)) if note_index >= 0 else ""
    return source_text, target_text, note


def _detect_header_indexes(row: tuple) -> tuple[int, int, int] | None:
    normalized_cells = [_normalize_header(_cell_to_text(row, index)) for index in range(len(row or ()))]
    source_index = _find_header_index(normalized_cells, SOURCE_HEADER_ALIASES)
    target_index = _find_header_index(normalized_cells, TARGET_HEADER_ALIASES)
    note_index = _find_header_index(normalized_cells, NOTE_HEADER_ALIASES)
    if source_index >= 0 and target_index >= 0:
        return source_index, target_index, note_index
    return None


def _find_header_index(cells: list[str], aliases: set[str]) -> int:
    for index, value in enumerate(cells):
        if value in aliases:
            return index
    return -1


def _normalize_header(value: str) -> str:
    return normalize_text(value).strip().lower().replace(" ", "_")


def _build_glossary_row(
    source_text: str,
    target_text: str,
    note: str,
    source_language: str,
    target_language: str,
    glossary_base_id: UUID,
    creator_id: UUID | None = None,
) -> dict:
    return {
        "glossary_base_id": glossary_base_id,
        "source_text": source_text,
        "target_text": target_text,
        "note": note or None,
        "source_normalized": normalize_match_text(source_text) or normalize_text(source_text),
        "source_language": source_language,
        "target_language": target_language,
        "creator_id": creator_id,
        "last_modified_by_id": creator_id,
    }


def _append_preview_row(
    rows: list[GlossaryImportPreviewRow],
    preview_limit: int,
    row: GlossaryImportPreviewRow,
) -> None:
    if len(rows) < preview_limit:
        rows.append(row)


def _load_existing_glossary_status(
    db: Session,
    glossary_base_id: UUID | None,
    source_language: str,
    target_language: str,
    source_normalized_values: list[str],
    source_texts: list[str],
) -> set[str]:
    if glossary_base_id is None or (not source_normalized_values and not source_texts):
        return set()

    keys: set[str] = set()
    source_normalized_values = list(dict.fromkeys(source_normalized_values))
    source_texts = list(dict.fromkeys(source_texts))
    max_lookup_count = max(len(source_normalized_values), len(source_texts))

    for offset in range(0, max_lookup_count, GLOSSARY_STATUS_LOOKUP_CHUNK_SIZE):
        normalized_chunk = source_normalized_values[offset : offset + GLOSSARY_STATUS_LOOKUP_CHUNK_SIZE]
        source_text_chunk = source_texts[offset : offset + GLOSSARY_STATUS_LOOKUP_CHUNK_SIZE]
        lookup_conditions = []
        if normalized_chunk:
            lookup_conditions.append(GlossaryEntry.source_normalized.in_(normalized_chunk))
        if source_text_chunk:
            lookup_conditions.append(GlossaryEntry.source_text.in_(source_text_chunk))

        existing_rows = (
            db.query(GlossaryEntry.source_normalized, GlossaryEntry.source_text)
            .filter(
                GlossaryEntry.glossary_base_id == glossary_base_id,
                GlossaryEntry.source_language == source_language,
                GlossaryEntry.target_language == target_language,
                or_(*lookup_conditions),
            )
            .all()
        )
        for source_normalized, source_text in existing_rows:
            if source_normalized:
                keys.add(source_normalized)
            if source_text:
                keys.add(source_text)
    return keys


def _flush_glossary_batch(
    db: Session,
    batch_rows: list[dict],
    source_language: str,
    target_language: str,
    glossary_base_id: UUID,
) -> tuple[int, int]:
    if not batch_rows:
        return 0, 0

    source_normalized_values = [row["source_normalized"] for row in batch_rows]
    source_texts = [row["source_text"] for row in batch_rows]
    existing_rows = (
        db.query(GlossaryEntry)
        .filter(
            GlossaryEntry.glossary_base_id == glossary_base_id,
            GlossaryEntry.source_language == source_language,
            GlossaryEntry.target_language == target_language,
            or_(
                GlossaryEntry.source_normalized.in_(source_normalized_values),
                GlossaryEntry.source_text.in_(source_texts),
            ),
        )
        .all()
    )

    existing_by_normalized: dict[str, GlossaryEntry] = {}
    existing_by_source_text: dict[str, GlossaryEntry] = {}
    for existing in existing_rows:
        if existing.source_normalized:
            existing_by_normalized.setdefault(existing.source_normalized, existing)
        existing_by_source_text.setdefault(existing.source_text, existing)

    created_rows = 0
    updated_rows = 0
    for row in batch_rows:
        existing = existing_by_normalized.get(row["source_normalized"]) or existing_by_source_text.get(
            row["source_text"]
        )
        if existing is None:
            db.add(GlossaryEntry(**row))
            created_rows += 1
            continue

        existing.source_text = row["source_text"]
        existing.target_text = row["target_text"]
        existing.note = row["note"]
        existing.source_normalized = row["source_normalized"]
        existing.source_language = row["source_language"]
        existing.target_language = row["target_language"]
        if existing.creator_id is None and row.get("creator_id"):
            existing.creator_id = row["creator_id"]
        if row.get("last_modified_by_id"):
            existing.last_modified_by_id = row["last_modified_by_id"]
        updated_rows += 1

    db.commit()
    return created_rows, updated_rows


def _cell_to_text(row: tuple, index: int) -> str:
    if index < 0 or index >= len(row or ()):
        return ""
    value = row[index]
    return "" if value is None else str(value)
