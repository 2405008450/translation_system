"""多语种 Excel 校对批次。

该模块只处理“表格型 XLSX”：用户显式映射原文列和一个或多个译文列，
每个目标语言仍落到现有 FileRecord/Segment，从而复用工作台、修订和项目同步。
"""
from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import re
from collections import defaultdict
from copy import copy
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from sqlalchemy.orm import Session

from app.config import get_settings
from app.database import SessionLocal
from app.models import (
    FileRecord,
    Project,
    ProofreadingBatch,
    ProofreadingColumnBinding,
    ProofreadingSegmentBaseline,
    Segment,
    SegmentRevision,
    TranslationReviewReport,
    TranslationReviewReportItem,
    User,
)
from app.services.analytics_service import count_source_words, record_translation_metric_event
from app.services.document_storage import save_source_file
from app.services.file_record_service import load_file_record_source, sync_file_record_status
from app.services.glossary_matcher import build_glossary_matches_by_text
from app.services.language_pairs import LANGUAGE_LABELS, normalize_language_code, require_language_pair
from app.services.llm_service import (
    LLMResponseValidationError,
    LLMTranslationTask,
    _validate_translation_output,
    request_chat_completion,
)
from app.services.normalizer import build_source_hash, normalize_text
from app.services.matcher import match_sentences
from app.services.reference_sync_service import attach_project_reference_bases_to_file
from app.services.translation_review.llm_gate import llm_gate

logger = logging.getLogger(__name__)

PROOFREADING_BATCH_SIZE = 20
PROOFREADING_BATCH_CHAR_LIMIT = 24000
PROOFREADING_SOURCE = "llm_review"
IMPORTED_TRANSLATION_SOURCE = "imported_translation"


def _load_batch_config(batch: ProofreadingBatch) -> dict[str, Any]:
    try:
        value = json.loads(batch.config_json or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return value if isinstance(value, dict) else {}


def _utcnow_naive() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


_HEADER_LANGUAGE_HINTS: dict[str, str] = {
    "english": "en-US", "英语": "en-US", "英文": "en-US",
    "french": "fr-FR", "法语": "fr-FR", "法文": "fr-FR",
    "german": "de-DE", "德语": "de-DE", "德文": "de-DE",
    "spanish": "es-ES", "西班牙语": "es-ES", "西语": "es-ES",
    "portuguese": "pt-BR", "葡萄牙语": "pt-BR", "葡语": "pt-BR",
    "polish": "pl-PL", "波兰语": "pl-PL",
    "japanese": "ja-JP", "日语": "ja-JP", "日文": "ja-JP",
    "korean": "ko-KR", "韩语": "ko-KR", "韩文": "ko-KR",
    "italian": "it-IT", "意大利语": "it-IT",
    "russian": "ru-RU", "俄语": "ru-RU",
    "chinese": "zh-CN", "中文": "zh-CN",
}


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _suggest_language(header: str) -> str | None:
    normalized = normalize_text(header).lower()
    if not normalized:
        return None
    for alias, language in _HEADER_LANGUAGE_HINTS.items():
        if alias in normalized:
            return language
    for code, label in LANGUAGE_LABELS.items():
        if code.lower() in normalized or label.lower() in normalized:
            return code
    for token in re.findall(r"[a-z]{2}(?:-[a-z0-9]{2,4})?", normalized):
        try:
            language = normalize_language_code(token, field_label="语言")
        except ValueError:
            continue
        if language:
            return language
    return None


def _detect_header_row(worksheet: Any) -> int:
    best_row = 1
    best_score = -1
    max_scan = min(int(worksheet.max_row or 1), 20)
    max_col = min(int(worksheet.max_column or 1), 100)
    for row_index in range(1, max_scan + 1):
        values = [_cell_text(worksheet.cell(row=row_index, column=col).value) for col in range(1, max_col + 1)]
        nonempty = [value for value in values if value]
        language_hits = sum(1 for value in nonempty if _suggest_language(value))
        source_hits = sum(1 for value in nonempty if any(token in value.lower() for token in ("原文", "source", "描述")))
        score = len(nonempty) + language_hits * 3 + source_hits * 2
        if score > best_score:
            best_row = row_index
            best_score = score
    return best_row


def _unsafe_sheet_structure_reasons(worksheet: Any) -> list[str]:
    formula_count = sum(
        1
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )
    reasons: list[str] = []
    if formula_count:
        reasons.append(f"包含 {formula_count} 个公式单元格")
    if worksheet.merged_cells.ranges:
        reasons.append(f"包含 {len(worksheet.merged_cells.ranges)} 个合并区域")
    if worksheet.tables:
        reasons.append(f"包含 {len(worksheet.tables)} 个 Excel Table")
    if getattr(worksheet, "_images", None) or getattr(worksheet, "_charts", None):
        reasons.append("包含图片或图表")
    if getattr(getattr(worksheet, "data_validations", None), "dataValidation", None):
        reasons.append("包含数据验证区域")
    if len(getattr(worksheet, "conditional_formatting", [])):
        reasons.append("包含条件格式区域")
    if getattr(worksheet, "_pivots", None):
        reasons.append("包含数据透视表")
    return reasons


def preview_workbook(raw_bytes: bytes, filename: str) -> dict[str, Any]:
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("校对工作流一期仅支持 .xlsx 文件。")
    try:
        workbook = load_workbook(BytesIO(raw_bytes), read_only=False, data_only=False)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"无法读取 Excel 文件：{exc}") from exc

    sheets: list[dict[str, Any]] = []
    try:
        for sheet_index, worksheet in enumerate(workbook.worksheets):
            header_row = _detect_header_row(worksheet)
            blocked_reasons = _unsafe_sheet_structure_reasons(worksheet)

            columns: list[dict[str, Any]] = []
            max_col = min(int(worksheet.max_column or 1), 200)
            sample_end = min(int(worksheet.max_row or header_row), header_row + 5)
            for column in range(1, max_col + 1):
                header = _cell_text(worksheet.cell(row=header_row, column=column).value)
                samples = [
                    _cell_text(worksheet.cell(row=row, column=column).value)
                    for row in range(header_row + 1, sample_end + 1)
                ]
                if not header and not any(samples):
                    continue
                columns.append({
                    "index": column,
                    "letter": get_column_letter(column),
                    "header": header,
                    "samples": samples,
                    "suggested_language": _suggest_language(header),
                    "suggested_role": (
                        "source"
                        if any(token in header.lower() for token in ("原文", "source", "英语", "english"))
                        and column == 1
                        else "target" if _suggest_language(header) else "other"
                    ),
                })
            sheets.append({
                "sheet_index": sheet_index,
                "name": worksheet.title,
                "header_row": header_row,
                "max_row": int(worksheet.max_row or 0),
                "max_column": int(worksheet.max_column or 0),
                "columns": columns,
                "header_candidates": [
                    {
                        "row_index": row_index,
                        "values": [
                            _cell_text(worksheet.cell(row=row_index, column=column).value)
                            for column in range(1, max_col + 1)
                        ],
                    }
                    for row_index in range(1, min(int(worksheet.max_row or 1), 20) + 1)
                ],
                "supported": not blocked_reasons,
                "blocked_reasons": blocked_reasons,
            })
    finally:
        workbook.close()

    return {
        "filename": filename,
        "file_hash": hashlib.sha256(raw_bytes).hexdigest(),
        "sheets": sheets,
    }


def _validate_mapping(workbook: Any, mappings: list[dict[str, Any]], source_language: str) -> list[dict[str, Any]]:
    normalized_mappings: list[dict[str, Any]] = []
    seen_targets: set[tuple[int, int]] = set()
    if not mappings:
        raise ValueError("请至少配置一个工作表和一个译文列。")
    for mapping in mappings:
        sheet_index = int(mapping.get("sheet_index", -1))
        if sheet_index < 0 or sheet_index >= len(workbook.worksheets):
            raise ValueError("工作表索引无效。")
        worksheet = workbook.worksheets[sheet_index]
        header_row = int(mapping.get("header_row", 0))
        source_column = int(mapping.get("source_column", 0))
        if (
            header_row < 1
            or header_row > int(worksheet.max_row or 1)
            or source_column < 1
            or source_column > int(worksheet.max_column or 1)
        ):
            raise ValueError(f"工作表“{worksheet.title}”的表头行或原文列无效。")
        unsafe_reasons = _unsafe_sheet_structure_reasons(worksheet)
        if unsafe_reasons:
            raise ValueError(f"工作表“{worksheet.title}”暂不支持安全插列：{'；'.join(unsafe_reasons)}。")

        targets: list[dict[str, Any]] = []
        seen_languages: set[str] = set()
        for target in mapping.get("targets") or []:
            target_column = int(target.get("target_column", 0))
            if (
                target_column < 1
                or target_column > int(worksheet.max_column or 1)
                or target_column == source_column
            ):
                raise ValueError(f"工作表“{worksheet.title}”的译文列无效。")
            target_language = normalize_language_code(target.get("target_language"), field_label="目标语言")
            if not target_language:
                raise ValueError("每个译文列都必须选择目标语言。")
            require_language_pair(source_language, target_language)
            if target_language in seen_languages:
                raise ValueError(f"工作表“{worksheet.title}”中同一目标语言只能映射一列。")
            binding_key = (sheet_index, target_column)
            if binding_key in seen_targets:
                raise ValueError("同一个译文列不能重复映射。")
            seen_languages.add(target_language)
            seen_targets.add(binding_key)
            targets.append({"target_column": target_column, "target_language": target_language})
        if not targets:
            raise ValueError(f"工作表“{worksheet.title}”至少需要一个译文列。")
        normalized_mappings.append({
            "sheet_index": sheet_index,
            "header_row": header_row,
            "source_column": source_column,
            "targets": targets,
        })
    return normalized_mappings


def create_batch_from_workbook(
    db: Session,
    *,
    project: Project,
    current_user: User,
    raw_bytes: bytes,
    filename: str,
    source_language: str,
    mappings: list[dict[str, Any]],
) -> ProofreadingBatch:
    source_language = normalize_language_code(source_language, field_label="源语言") or ""
    if not source_language:
        raise ValueError("请选择源语言。")
    if getattr(project, "workflow_template_id", "") != "proofread":
        raise ValueError("只有“校对”工作流项目可以创建多语种校对批次。")
    workbook = load_workbook(BytesIO(raw_bytes), read_only=False, data_only=False)
    try:
        normalized_mappings = _validate_mapping(workbook, mappings, source_language)
        batch = ProofreadingBatch(
            project_id=project.id,
            created_by_id=current_user.id,
            filename=filename,
            file_hash=hashlib.sha256(raw_bytes).hexdigest(),
            source_language=source_language,
            status="ready",
            progress=0,
            message="已完成列映射，等待开始校对。",
            config_json=json.dumps({"mappings": normalized_mappings}, ensure_ascii=False),
        )
        db.add(batch)
        db.flush()

        target_languages = list(dict.fromkeys(
            target["target_language"]
            for mapping in normalized_mappings
            for target in mapping["targets"]
        ))
        file_records: dict[str, FileRecord] = {}
        filename_path = Path(filename)
        for target_language in target_languages:
            language_filename = f"{filename_path.stem} - {target_language}{filename_path.suffix}"
            file_record = FileRecord(
                project_id=project.id,
                filename=language_filename,
                file_hash=batch.file_hash,
                status="in_progress",
                document_parse_mode="full",
                document_parse_options=json.dumps({"xlsx_mode": "proofread", "proofreading_batch_id": str(batch.id)}, ensure_ascii=False),
                creator_id=current_user.id,
                deadline=project.deadline,
                access_level=project.access_level,
                source_language=source_language,
                target_language=target_language,
            )
            db.add(file_record)
            db.flush()
            save_source_file(file_record.id, language_filename, raw_bytes)
            attach_project_reference_bases_to_file(db, file_record)
            file_records[target_language] = file_record

        total_segments = 0
        skipped_segments = 0
        first_step = (
            project.workflow_steps[0].id
            if getattr(project, "workflow_steps", None)
            else None
        )
        for mapping in normalized_mappings:
            worksheet = workbook.worksheets[mapping["sheet_index"]]
            header_row = mapping["header_row"]
            source_column = mapping["source_column"]
            source_header = _cell_text(worksheet.cell(row=header_row, column=source_column).value)
            for target in mapping["targets"]:
                target_column = target["target_column"]
                target_language = target["target_language"]
                file_record = file_records[target_language]
                target_header = _cell_text(worksheet.cell(row=header_row, column=target_column).value)
                binding = ProofreadingColumnBinding(
                    batch_id=batch.id,
                    file_record_id=file_record.id,
                    sheet_index=mapping["sheet_index"],
                    sheet_name=worksheet.title,
                    header_row=header_row,
                    source_column=source_column,
                    target_column=target_column,
                    output_column=target_column + 1,
                    source_header=source_header,
                    target_header=target_header,
                    target_language=target_language,
                )
                db.add(binding)
                db.flush()

                sequence_index = db.query(Segment).filter(Segment.file_record_id == file_record.id).count()
                for row_index in range(header_row + 1, int(worksheet.max_row or header_row) + 1):
                    source_text = _cell_text(worksheet.cell(row=row_index, column=source_column).value)
                    if not source_text:
                        continue
                    original_target = _cell_text(worksheet.cell(row=row_index, column=target_column).value)
                    if not original_target:
                        skipped_segments += 1
                    sentence_id = f"proof-{mapping['sheet_index'] + 1}-{row_index}-{source_column}-{target_column}"
                    metadata = {
                        "proofreading_batch_id": str(batch.id),
                        "proofreading_binding_id": str(binding.id),
                        "sheet_index": mapping["sheet_index"],
                        "sheet_name": worksheet.title,
                        "source_cell_ref": f"{get_column_letter(source_column)}{row_index}",
                        "target_cell_ref": f"{get_column_letter(target_column)}{row_index}",
                        "original_target_text": original_target,
                    }
                    segment = Segment(
                        file_record_id=file_record.id,
                        workflow_step_id=first_step,
                        sentence_id=sentence_id,
                        source_text=source_text,
                        source_hash=build_source_hash(source_text),
                        display_text=source_text,
                        target_text=original_target,
                        status="none",
                        source=IMPORTED_TRANSLATION_SOURCE if original_target else "none",
                        source_word_count=count_source_words(source_text),
                        block_type="table_cell",
                        block_index=mapping["sheet_index"],
                        row_index=row_index - 1,
                        cell_index=target_column - 1,
                        sequence_index=sequence_index,
                        display_index=sequence_index,
                        segment_metadata=json.dumps(metadata, ensure_ascii=False),
                    )
                    db.add(segment)
                    db.flush()
                    db.add(ProofreadingSegmentBaseline(
                        batch_id=batch.id,
                        binding_id=binding.id,
                        segment_id=segment.id,
                        sheet_index=mapping["sheet_index"],
                        row_index=row_index,
                        source_cell_ref=metadata["source_cell_ref"],
                        target_cell_ref=metadata["target_cell_ref"],
                        original_target_text=original_target,
                    ))
                    sequence_index += 1
                    total_segments += 1

        batch.total_segments = total_segments
        batch.skipped_segments = skipped_segments
        project.status = "in_progress"
        db.commit()
        db.refresh(batch)
        return batch
    except Exception:
        db.rollback()
        raise
    finally:
        workbook.close()


def _safe_json_array(content: str) -> list[dict[str, Any]]:
    cleaned = re.sub(r"```(?:json)?|```", "", content or "").strip()
    match = re.search(r"\[.*\]", cleaned, re.S)
    if match:
        cleaned = match.group(0)
    try:
        value = json.loads(cleaned)
    except (TypeError, ValueError):
        return []
    return [item for item in value if isinstance(item, dict)] if isinstance(value, list) else []


def _validate_target_language_script(text_value: str, target_language: str) -> None:
    """拦截明显输出成错误文字系统的结果；拉丁语种的细分仍由提示词与人工复核保证。"""
    letters = re.findall(r"[A-Za-z\u3400-\u9fff\u3040-\u30ff\uac00-\ud7af\u0400-\u04ff\u0600-\u06ff]", text_value)
    if len(letters) < 3:
        return
    language_base = target_language.split("-", 1)[0].lower()
    patterns = {
        "zh": r"[\u3400-\u9fff]",
        "ja": r"[\u3040-\u30ff\u3400-\u9fff]",
        "ko": r"[\uac00-\ud7af]",
        "ru": r"[\u0400-\u04ff]",
        "uk": r"[\u0400-\u04ff]",
        "ar": r"[\u0600-\u06ff]",
        "fa": r"[\u0600-\u06ff]",
        "ur": r"[\u0600-\u06ff]",
    }
    expected_pattern = patterns.get(language_base)
    if expected_pattern and not re.search(expected_pattern, text_value):
        raise LLMResponseValidationError("LLM 输出未使用目标语言对应的文字系统。")
    if not expected_pattern:
        non_latin = len(re.findall(r"[\u3400-\u9fff\u3040-\u30ff\uac00-\ud7af\u0400-\u04ff\u0600-\u06ff]", text_value))
        latin = len(re.findall(r"[A-Za-z]", text_value))
        if non_latin >= 3 and non_latin > latin:
            raise LLMResponseValidationError("LLM 输出疑似不是所选目标语言。")


def _pack_groups(groups: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    batches: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    current_chars = 0
    for group in groups:
        item_chars = (
            len(group["source_text"])
            + sum(len(value) for value in group["variants"])
            + len(json.dumps(group.get("glossary_references", []), ensure_ascii=False))
            + len(json.dumps(group.get("tm_reference", {}), ensure_ascii=False))
            + 200
        )
        if current and (len(current) >= PROOFREADING_BATCH_SIZE or current_chars + item_chars > PROOFREADING_BATCH_CHAR_LIMIT):
            batches.append(current)
            current = []
            current_chars = 0
        current.append(group)
        current_chars += item_chars
    if current:
        batches.append(current)
    return batches


def _parse_uuid_list(raw_value: str | None) -> list[UUID]:
    try:
        values = json.loads(raw_value or "[]")
    except (TypeError, ValueError):
        return []
    result: list[UUID] = []
    for value in values if isinstance(values, list) else []:
        try:
            result.append(UUID(str(value)))
        except (TypeError, ValueError):
            continue
    return list(dict.fromkeys(result))


def _attach_reference_context(
    db: Session,
    file_record: FileRecord,
    groups: list[dict[str, Any]],
    *,
    source_language: str,
    target_language: str,
) -> None:
    source_texts = [group["source_text"] for group in groups]
    glossary_ids = _parse_uuid_list(getattr(file_record, "glossary_base_ids", "[]"))
    if glossary_ids:
        matches_by_text = build_glossary_matches_by_text(
            db,
            source_texts,
            glossary_ids,
            source_language=source_language,
            target_language=target_language,
        )
        for group in groups:
            group["glossary_references"] = [
                {"source_text": item.source_text, "target_text": item.target_text, "note": item.note}
                for item in matches_by_text.get(group["source_text"], [])
            ]

    collection_ids = _parse_uuid_list(getattr(file_record, "collection_ids_json", "[]"))
    if collection_ids:
        try:
            tm_results = match_sentences(
                db,
                source_texts,
                float(getattr(file_record, "tm_match_threshold", 0.8) or 0.8),
                collection_ids=collection_ids,
                source_language=source_language,
                target_language=target_language,
                include_fuzzy=True,
            )
        except Exception as exc:  # noqa: BLE001 - TM 不可用不应阻断校对
            logger.warning("proofreading TM lookup failed file=%s: %s", file_record.id, exc)
            return
        for group, match in zip(groups, tm_results):
            if match.status != "none" and normalize_text(match.target_text):
                group["tm_reference"] = {
                    "source_text": match.matched_source_text or group["source_text"],
                    "target_text": match.target_text,
                    "score": round(float(match.score or 0), 4),
                }


def _build_prompt(
    groups: list[dict[str, Any]],
    source_language: str,
    target_language: str,
    rules_text: str,
    user_instructions: str = "",
) -> str:
    items = []
    for seq, group in enumerate(groups):
        variants = "\n".join(f"    - {value}" for value in group["variants"])
        glossary = "\n".join(
            f"    - {item['source_text']} → {item['target_text']}"
            for item in group.get("glossary_references", [])
        )
        tm_reference = group.get("tm_reference")
        references = ""
        if glossary:
            references += f"\n  术语库命中（必须保持）：\n{glossary}"
        if tm_reference:
            references += (
                "\n  TM 参考（仅供参考，不得直接覆盖现有译文）："
                f"\n    - {tm_reference['source_text']} → {tm_reference['target_text']}"
            )
        items.append(
            f"[{seq}] <sid={group['sid']}>\n"
            f"  原文：{group['source_text']}\n"
            f"  现有译文候选：\n{variants}{references}"
        )
    rules = rules_text.strip() or "准确、自然、简洁；保留数字、占位符、标签和专有名词；同一原文必须输出同一译文。"
    user_rules = user_instructions.strip()
    user_section = f"\n【本批次用户校对要求】\n{user_rules}\n" if user_rules else ""
    return (
        f"源语言：{source_language}\n目标语言：{target_language}\n"
        f"【项目/系统校对规则】\n{rules}\n{user_section}\n"
        "请为每组原文及现有译文候选生成一个唯一、完整的校对版译文。只输出 JSON 数组，顺序和 sid 必须保持不变。\n"
        "字段：seq、sid、reviewed_target_text、changed、reason、category、confidence。"
        "reason 必须使用简体中文，便于中文审校人员复核；category 使用简短的问题类别。"
        "confidence 只能是 high/medium/low。不得省略任何输入项。\n\n"
        + "\n\n".join(items)
    )


async def _review_group_batch(
    groups: list[dict[str, Any]],
    *,
    source_language: str,
    target_language: str,
    rules_text: str,
    user_instructions: str = "",
    provider: str,
    model: str | None,
) -> tuple[dict[str, dict[str, Any]], str, str]:
    prompt = _build_prompt(groups, source_language, target_language, rules_text, user_instructions)
    expected = {index: group["sid"] for index, group in enumerate(groups)}
    last_error = "LLM 返回格式无效。"
    for attempt in range(2):
        messages = [
            {"role": "system", "content": "你是资深本地化校对专家。严格返回 JSON，不要解释。"},
            {"role": "user", "content": prompt + (f"\n\n上次错误：{last_error}，请修正后重试。" if attempt else "")},
        ]
        try:
            async with llm_gate():
                completion = await request_chat_completion(
                    messages=messages,
                    provider=provider,
                    model_override=model,
                    temperature=0,
                    allow_fallback=provider == "auto",
                )
        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)
            continue
        parsed = _safe_json_array(completion.content)
        results: dict[str, dict[str, Any]] = {}
        validation_errors: list[str] = []
        for item in parsed:
            try:
                seq = int(item.get("seq"))
            except (TypeError, ValueError):
                continue
            sid = str(item.get("sid") or "")
            if expected.get(seq) != sid or not isinstance(item.get("changed"), bool):
                continue
            text_value = str(item.get("reviewed_target_text") or "").strip()
            group = groups[seq]
            try:
                _validate_translation_output(
                    LLMTranslationTask(
                        sentence_id=sid,
                        status="none",
                        source_text=group["source_text"],
                        source_language=source_language,
                        target_language=target_language,
                    ),
                    text_value,
                )
                _validate_target_language_script(text_value, target_language)
            except LLMResponseValidationError as exc:
                validation_errors.append(f"{sid}: {exc}")
                continue
            confidence = str(item.get("confidence") or "medium").lower()
            if confidence not in {"high", "medium", "low"}:
                confidence = "medium"
            results[sid] = {
                "reviewed_target_text": text_value,
                "changed": item["changed"],
                "reason": str(item.get("reason") or "")[:1000],
                "category": str(item.get("category") or "校对改写")[:40],
                "confidence": confidence,
            }
        if len(results) == len(groups):
            return results, completion.provider, completion.model
        last_error = "; ".join(validation_errors) or f"仅返回 {len(results)}/{len(groups)} 项有效结果。"
    raise ValueError(last_error)


async def generate_batch(
    db: Session,
    batch: ProofreadingBatch,
    *,
    current_user: User | None,
    provider: str = "auto",
    model: str | None = None,
    user_instructions: str = "",
) -> None:
    batch.status = "running"
    batch.progress = 1
    batch.message = "正在整理重复原文。"
    batch.error_message = ""
    db.commit()

    bindings = (
        db.query(ProofreadingColumnBinding)
        .filter(ProofreadingColumnBinding.batch_id == batch.id)
        .order_by(ProofreadingColumnBinding.target_language, ProofreadingColumnBinding.sheet_index, ProofreadingColumnBinding.target_column)
        .all()
    )
    file_ids = list(dict.fromkeys(binding.file_record_id for binding in bindings))
    report = TranslationReviewReport(
        project_id=batch.project_id,
        created_by_id=current_user.id if current_user else batch.created_by_id,
        scope="proofreading_batch",
        report_mode="proofread_generate",
        proofreading_batch_id=batch.id,
        file_ids=json.dumps([str(value) for value in file_ids]),
        total_files=len(file_ids),
        provider=provider,
        model=model or "",
        status="running",
        progress="{}",
        category_counts="{}",
        file_counts="{}",
        failed_categories="[]",
    )
    db.add(report)
    db.flush()

    rules_text = batch.project.translation_rules or ""
    changed_count = 0
    failed_count = 0
    checked_count = 0
    category_counts: dict[str, int] = defaultdict(int)

    bindings_by_language: dict[str, list[ProofreadingColumnBinding]] = defaultdict(list)
    for binding in bindings:
        bindings_by_language[binding.target_language].append(binding)

    total_groups = 0
    grouped_by_language: dict[str, list[dict[str, Any]]] = {}
    for target_language, language_bindings in bindings_by_language.items():
        binding_ids = [binding.id for binding in language_bindings]
        rows = (
            db.query(ProofreadingSegmentBaseline, Segment)
            .join(Segment, Segment.id == ProofreadingSegmentBaseline.segment_id)
            .filter(ProofreadingSegmentBaseline.binding_id.in_(binding_ids))
            .order_by(ProofreadingSegmentBaseline.sheet_index, ProofreadingSegmentBaseline.row_index)
            .all()
        )
        groups_by_hash: dict[str, dict[str, Any]] = {}
        for baseline, segment in rows:
            if not normalize_text(baseline.original_target_text) or segment.status == "confirmed":
                continue
            source_hash = segment.source_hash or build_source_hash(segment.source_text)
            group = groups_by_hash.setdefault(source_hash, {
                "sid": str(segment.id),
                "source_text": segment.source_text,
                "variants": [],
                "segments": [],
            })
            if baseline.original_target_text not in group["variants"]:
                group["variants"].append(baseline.original_target_text)
            group["segments"].append((baseline, segment))
        groups = list(groups_by_hash.values())
        file_record = db.get(FileRecord, language_bindings[0].file_record_id)
        if file_record and groups:
            _attach_reference_context(
                db,
                file_record,
                groups,
                source_language=batch.source_language,
                target_language=target_language,
            )
        grouped_by_language[target_language] = groups
        total_groups += len(groups)

    processed_groups = 0
    for target_language, groups in grouped_by_language.items():
        for packed in _pack_groups(groups):
            try:
                results, actual_provider, actual_model = await _review_group_batch(
                    packed,
                    source_language=batch.source_language,
                    target_language=target_language,
                    rules_text=rules_text,
                    user_instructions=user_instructions,
                    provider=provider,
                    model=model,
                )
            except Exception as exc:  # noqa: BLE001
                logger.warning("proofreading batch=%s language=%s group failed: %s", batch.id, target_language, exc)
                packed_failed_count = sum(len(group["segments"]) for group in packed)
                failed_count += packed_failed_count
                category_counts["generation_error"] += packed_failed_count
                error_text = str(exc)[:1000]
                for group in packed:
                    for baseline, segment in group["segments"]:
                        db.add(TranslationReviewReportItem(
                            report_id=report.id,
                            project_id=batch.project_id,
                            file_record_id=segment.file_record_id,
                            segment_id=segment.id,
                            sentence_id=segment.sentence_id,
                            file_name=segment.file_record.filename,
                            display_index=segment.display_index,
                            sequence_index=segment.sequence_index,
                            category_key="generation_error",
                            severity="error",
                            origin="ai",
                            source_text=segment.source_text,
                            target_text=baseline.original_target_text,
                            reason=error_text,
                            confidence="low",
                            apply_mode="manual",
                            locate_status="invalid",
                            original_target_text=baseline.original_target_text,
                            applied=False,
                            status="open",
                            block_index=segment.block_index,
                            row_index=segment.row_index,
                            cell_index=segment.cell_index,
                        ))
                processed_groups += len(packed)
                batch.progress = min(99, round(processed_groups / max(total_groups, 1) * 100))
                batch.message = f"{target_language} 部分句段校对失败，继续处理其余内容。"
                db.commit()
                continue

            report.provider = actual_provider
            report.model = actual_model
            for group in packed:
                result = results[group["sid"]]
                reviewed_text = result["reviewed_target_text"]
                for baseline, segment in group["segments"]:
                    before_text = segment.target_text or ""
                    checked_count += 1
                    segment.source = PROOFREADING_SOURCE
                    segment.llm_provider = actual_provider
                    segment.llm_model = actual_model
                    segment.last_modified_by_id = current_user.id if current_user else None
                    if before_text == reviewed_text:
                        continue
                    segment.target_text = reviewed_text
                    segment.target_html = None
                    segment.version = int(segment.version or 1) + 1
                    segment.last_modified_by_id = current_user.id if current_user else None
                    db.add(SegmentRevision(
                        file_record_id=segment.file_record_id,
                        segment_id=segment.id,
                        sentence_id=segment.sentence_id,
                        before_text=before_text,
                        after_text=reviewed_text,
                        source=PROOFREADING_SOURCE,
                        status="pending",
                        author_id=current_user.id if current_user else None,
                    ))
                    record_translation_metric_event(
                        db,
                        segment=segment,
                        before_text=before_text,
                        after_text=reviewed_text,
                        source=PROOFREADING_SOURCE,
                        current_user=current_user,
                    )
                    category = result["category"] or "校对改写"
                    category_counts[category] += 1
                    db.add(TranslationReviewReportItem(
                        report_id=report.id,
                        project_id=batch.project_id,
                        file_record_id=segment.file_record_id,
                        segment_id=segment.id,
                        sentence_id=segment.sentence_id,
                        file_name=segment.file_record.filename,
                        display_index=segment.display_index,
                        sequence_index=segment.sequence_index,
                        category_key=category,
                        severity="warning",
                        origin="ai",
                        source_text=segment.source_text,
                        target_text=before_text,
                        suggested_target_text=reviewed_text,
                        reason=result["reason"],
                        confidence=result["confidence"],
                        apply_mode="full",
                        locate_status="ok",
                        original_target_text=before_text,
                        applied=True,
                        applied_at=_utcnow_naive(),
                        status="applied",
                        block_index=segment.block_index,
                        row_index=segment.row_index,
                        cell_index=segment.cell_index,
                    ))
                    changed_count += 1
            processed_groups += len(packed)
            batch.progress = min(99, round(processed_groups / max(total_groups, 1) * 100))
            batch.message = f"正在校对 {target_language}。"
            report.progress = json.dumps({"overall_percent": batch.progress, "target_language": target_language}, ensure_ascii=False)
            db.commit()

    for file_id in file_ids:
        sync_file_record_status(db, file_id)
    current_rows = (
        db.query(ProofreadingSegmentBaseline, Segment)
        .join(Segment, Segment.id == ProofreadingSegmentBaseline.segment_id)
        .filter(ProofreadingSegmentBaseline.batch_id == batch.id)
        .all()
    )
    batch.changed_segments = sum(
        1
        for baseline, segment in current_rows
        if (segment.target_text or "") != (baseline.original_target_text or "")
    )
    batch.failed_segments = failed_count
    batch.progress = 100
    batch.status = "partial_failed" if failed_count else "completed"
    batch.message = "校对部分完成。" if failed_count else "校对完成，可复核或导出。"
    batch.finished_at = _utcnow_naive()
    report.total_segments = checked_count
    report.checked_segments = checked_count
    report.issue_count = changed_count + failed_count
    report.active_issue_count = failed_count
    report.applied_count = changed_count
    report.category_counts = json.dumps(dict(category_counts), ensure_ascii=False)
    report.enabled_categories = json.dumps(sorted(category_counts), ensure_ascii=False)
    report.failed_categories = json.dumps(["generation_error"] if failed_count else [])
    report.status = (
        "partial_failed"
        if failed_count and checked_count
        else "failed"
        if failed_count
        else "completed"
    )
    report.finished_at = _utcnow_naive()
    db.commit()


def run_generate_batch(
    batch_id: UUID,
    current_user_id: UUID | None,
    provider: str,
    model: str | None,
    user_instructions: str = "",
) -> None:
    with SessionLocal() as db:
        batch = db.query(ProofreadingBatch).filter(ProofreadingBatch.id == batch_id).first()
        if not batch:
            return
        current_user = db.query(User).filter(User.id == current_user_id).first() if current_user_id else None
        try:
            asyncio.run(generate_batch(
                db,
                batch,
                current_user=current_user,
                provider=provider,
                model=model,
                user_instructions=user_instructions,
            ))
        except Exception as exc:  # noqa: BLE001
            db.rollback()
            batch = db.query(ProofreadingBatch).filter(ProofreadingBatch.id == batch_id).first()
            if batch:
                batch.status = "failed"
                batch.error_message = str(exc)
                batch.message = "校对失败。"
                batch.finished_at = _utcnow_naive()
                db.commit()
            logger.exception("proofreading generation failed batch_id=%s", batch_id)


def export_batch_xlsx(db: Session, batch: ProofreadingBatch) -> tuple[bytes, str]:
    bindings = (
        db.query(ProofreadingColumnBinding)
        .filter(ProofreadingColumnBinding.batch_id == batch.id)
        .order_by(ProofreadingColumnBinding.sheet_index, ProofreadingColumnBinding.target_column.desc())
        .all()
    )
    if not bindings:
        raise ValueError("校对批次没有可导出的列映射。")
    source_record = db.query(FileRecord).filter(FileRecord.id == bindings[0].file_record_id).first()
    raw_bytes = load_file_record_source(source_record) if source_record else None
    if not raw_bytes:
        raise ValueError("原始 Excel 文件缺失。")
    workbook = load_workbook(BytesIO(raw_bytes), read_only=False, data_only=False)
    try:
        for binding in bindings:
            worksheet = workbook.worksheets[binding.sheet_index]
            insert_at = binding.target_column + 1
            worksheet.insert_cols(insert_at, 1)
            source_letter = get_column_letter(binding.target_column)
            output_letter = get_column_letter(insert_at)
            source_dimension = worksheet.column_dimensions[source_letter]
            output_dimension = worksheet.column_dimensions[output_letter]
            output_dimension.width = source_dimension.width
            output_dimension.hidden = source_dimension.hidden
            output_dimension.bestFit = source_dimension.bestFit

            for row_index in range(1, int(worksheet.max_row or 1) + 1):
                source_cell = worksheet.cell(row=row_index, column=binding.target_column)
                output_cell = worksheet.cell(row=row_index, column=insert_at)
                if source_cell.has_style:
                    output_cell._style = copy(source_cell._style)  # noqa: SLF001 - openpyxl 的公开复制惯例
                output_cell.number_format = source_cell.number_format
                output_cell.alignment = copy(source_cell.alignment)
                output_cell.protection = copy(source_cell.protection)

            header = binding.target_header or f"{LANGUAGE_LABELS.get(binding.target_language, binding.target_language)}译文"
            worksheet.cell(row=binding.header_row, column=insert_at).value = f"{header}（校对版）"
            baseline_rows = (
                db.query(ProofreadingSegmentBaseline, Segment)
                .join(Segment, Segment.id == ProofreadingSegmentBaseline.segment_id)
                .filter(ProofreadingSegmentBaseline.binding_id == binding.id)
                .all()
            )
            for baseline, segment in baseline_rows:
                output_cell = worksheet.cell(row=baseline.row_index, column=insert_at)
                output_cell.value = segment.target_text or ""
                if (segment.target_text or "") != (baseline.original_target_text or ""):
                    changed_font = copy(output_cell.font)
                    changed_font.color = "FF0563C1"
                    changed_font.bold = True
                    output_cell.font = changed_font
                    output_cell.fill = PatternFill(fill_type="solid", fgColor="FFDDEBFF")

            if worksheet.auto_filter.ref:
                min_col, min_row, max_col, max_row = range_boundaries(worksheet.auto_filter.ref)
                if insert_at <= max_col:
                    max_col += 1
                worksheet.auto_filter.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

        output = BytesIO()
        workbook.save(output)
    finally:
        workbook.close()
    filename_path = Path(batch.filename)
    return output.getvalue(), f"{filename_path.stem}_校对版.xlsx"


def run_export_batch(batch_id: UUID) -> None:
    with SessionLocal() as db:
        batch = db.query(ProofreadingBatch).filter(ProofreadingBatch.id == batch_id).first()
        if not batch:
            return
        batch.export_status = "running"
        batch.export_progress = 10
        batch.export_error_message = ""
        db.commit()
        try:
            if getattr(batch, "batch_kind", "xlsx_columns") == "document_pair":
                from app.services.document_alignment.export import export_document_pair_xlsx
                content, filename = export_document_pair_xlsx(db, batch)
            else:
                content, filename = export_batch_xlsx(db, batch)
            export_dir = Path(get_settings().file_storage_dir) / "proofreading_exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            output_path = export_dir / f"{batch.id}.xlsx"
            temporary_path = export_dir / f"{batch.id}.tmp"
            temporary_path.write_bytes(content)
            temporary_path.replace(output_path)
            batch.export_status = "completed"
            batch.export_progress = 100
            batch.export_filename = filename
            batch.export_path = str(output_path)
            db.commit()
        except Exception as exc:  # noqa: BLE001
            db.rollback()
            batch = db.query(ProofreadingBatch).filter(ProofreadingBatch.id == batch_id).first()
            if batch:
                batch.export_status = "failed"
                batch.export_progress = 0
                batch.export_error_message = str(exc)
                db.commit()
            logger.exception("proofreading export failed batch_id=%s", batch_id)


def load_exported_batch(batch: ProofreadingBatch) -> tuple[bytes, str]:
    if batch.export_status != "completed" or not batch.export_path:
        raise ValueError("校对版 Excel 尚未生成完成。")
    path = Path(batch.export_path)
    if not path.is_file():
        raise ValueError("校对版 Excel 文件已失效，请重新生成。")
    return path.read_bytes(), batch.export_filename or f"{Path(batch.filename).stem}_校对版.xlsx"


def serialize_batch(db: Session, batch: ProofreadingBatch) -> dict[str, Any]:
    bindings = (
        db.query(ProofreadingColumnBinding)
        .filter(ProofreadingColumnBinding.batch_id == batch.id)
        .order_by(ProofreadingColumnBinding.sheet_index, ProofreadingColumnBinding.target_column)
        .all()
    )
    config = _load_batch_config(batch)
    generation = config.get("generation") if isinstance(config.get("generation"), dict) else {}
    latest_report = (
        db.query(TranslationReviewReport)
        .filter(TranslationReviewReport.proofreading_batch_id == batch.id)
        .order_by(TranslationReviewReport.created_at.desc())
        .first()
    )
    latest_llm_segment = (
        db.query(Segment)
        .join(ProofreadingSegmentBaseline, ProofreadingSegmentBaseline.segment_id == Segment.id)
        .filter(
            ProofreadingSegmentBaseline.batch_id == batch.id,
            Segment.llm_provider.is_not(None),
        )
        .order_by(Segment.updated_at.desc())
        .first()
    )
    actual_provider = (
        latest_llm_segment.llm_provider
        if latest_llm_segment and latest_llm_segment.llm_provider
        else latest_report.provider if latest_report else ""
    )
    actual_model = (
        latest_llm_segment.llm_model
        if latest_llm_segment and latest_llm_segment.llm_model
        else latest_report.model if latest_report else ""
    )
    return {
        "id": str(batch.id),
        "project_id": str(batch.project_id),
        "filename": batch.filename,
        "source_language": batch.source_language,
        "target_language": getattr(batch, "target_language", ""),
        "batch_kind": getattr(batch, "batch_kind", "xlsx_columns"),
        "alignment_status": getattr(batch, "alignment_status", "not_applicable"),
        "status": batch.status,
        "progress": batch.progress,
        "message": batch.message,
        "error_message": batch.error_message,
        "total_segments": batch.total_segments,
        "changed_segments": batch.changed_segments,
        "skipped_segments": batch.skipped_segments,
        "failed_segments": batch.failed_segments,
        "export_status": batch.export_status,
        "export_progress": batch.export_progress,
        "export_error_message": batch.export_error_message,
        "export_filename": batch.export_filename,
        "created_at": batch.created_at.isoformat() if batch.created_at else None,
        "finished_at": batch.finished_at.isoformat() if batch.finished_at else None,
        "generation_settings": {
            "provider": str(generation.get("provider") or "auto"),
            "model": str(generation.get("model") or ""),
            "user_instructions": str(generation.get("user_instructions") or ""),
            "actual_provider": str(actual_provider or ""),
            "actual_model": str(actual_model or ""),
        },
        "bindings": [
            {
                "id": str(binding.id),
                "file_record_id": str(binding.file_record_id),
                "sheet_index": binding.sheet_index,
                "sheet_name": binding.sheet_name,
                "header_row": binding.header_row,
                "source_column": binding.source_column,
                "target_column": binding.target_column,
                "output_column": binding.output_column,
                "source_header": binding.source_header,
                "target_header": binding.target_header,
                "target_language": binding.target_language,
            }
            for binding in bindings
        ],
    }
